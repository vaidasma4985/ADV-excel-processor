from __future__ import annotations

import re
from io import BytesIO
from typing import Any, Dict, List, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


# -----------------------------------------------------------------------------
# Type normalization/classification constants (single source of truth helpers)
# -----------------------------------------------------------------------------
_REMOVE_PREFIXES: Tuple[str, ...] = ("-B", "-C", "-R", "-M", "-P", "-Q", "-S", "-W", "-T")

_ALLOWED_RAW_TYPES: set[str] = {
    "2002-1611/1000-541",
    "2002-1611/1000-836",
    "2002-3201",
    "2002-3207",
    "RXZE2S114M",
    "RXM4GB2BD",
    "RXG22P7",
    "RXG22BD",
    "RGZE1S48M",
    "A9F04604",
    "RE17LCBM",
    "RE22R1AMR",
    "39.00.8.230.8240",
}

_WAGO_RAW_TYPES: set[str] = {
    "2002-1611/1000-541",
    "2002-1611/1000-836",
    "2002-3201",
    "2002-3207",
    "249-116",
    "2002-991",
    "2002-3292",
}

_SE_TYPE_MAP: Dict[str, str] = {
    "RGZE1S48M": "SE.RGZE1S48M",
    "RXG22P7": "SE.RXG22P7",
    "RXG22BD": "SE.RXG22BD",
    "RXM4GB2BD": "SE.RXM4GB2BD",
    "RXZE2S114M": "SE.RXZE2S114M",
    "A9F04604": "SE.A9F04604",
}

_ADV_TYPE_MAP: Dict[str, str] = {
    "WAGO.2002-3207": "WAGO.2002-3207_ADV",
    "WAGO.2002-3201": "WAGO.2002-3201_ADV",
    "WAGO.2002-3292": "WAGO.2002-3292_ADV",
    "WAGO.2002-991": "WAGO.2002-991_ADV",
    "WAGO.249-116": "WAGO.249-116_ADV",
    "WAGO.2002-1611/1000-541": "WAGO.2002-1611/1000-541_ADV",
    "WAGO.2002-1611/1000-836": "WAGO.2002-1611/1000-836_ADV",
    "SE.A9F04604": "SE.A9F04604_ADV",
}

_RELAY_TYPE_MAP: Dict[str, str] = {
    "RE17LCBM": "SE.RE17LCBM_ADV",
    "RE22R1AMR": "SE.RE22R1AMR_ADV",
    "39.00.8.230.8240": "FIN.39.00.8.230.8240_ADV",
}

_TERMINAL_TYPES: set[str] = {"WAGO.2002-3201_ADV", "WAGO.2002-3207_ADV"}
_FUSE_TYPES: set[str] = {
    "WAGO.2002-1611/1000-541_ADV",
    "WAGO.2002-1611/1000-836_ADV",
    "SE.A9F04604_ADV",
}


def normalize_type(raw_type: str) -> str:
    """Normalize a raw component type using the same mapping sequence as STEP 4."""
    if raw_type is None or pd.isna(raw_type):
        return ""

    type_value = str(raw_type).strip()
    if not type_value:
        return ""

    if type_value in _WAGO_RAW_TYPES:
        type_value = f"WAGO.{type_value}"

    type_value = _SE_TYPE_MAP.get(type_value, type_value)
    type_value = _ADV_TYPE_MAP.get(type_value, type_value)
    type_value = _RELAY_TYPE_MAP.get(type_value, type_value)

    return str(type_value)


def classify_component(name: str, raw_type: str, group_sorting: Any) -> Dict[str, Any]:
    """Classify one component row without mutating external state."""
    name_str = "" if name is None or pd.isna(name) else str(name)
    raw_type_str = "" if raw_type is None or pd.isna(raw_type) else str(raw_type).strip()
    normalized_type = normalize_type(raw_type)

    gs_num = pd.to_numeric(group_sorting, errors="coerce")
    gs_numeric = None if pd.isna(gs_num) else float(gs_num)
    gs_is_non_numeric = bool(
        not pd.isna(group_sorting)
        and str(group_sorting).strip() != ""
        and pd.isna(gs_num)
    )

    removed_by_name_prefix = name_str.startswith(_REMOVE_PREFIXES)
    allowed_raw_type = raw_type_str in _ALLOWED_RAW_TYPES
    would_be_processed = (not removed_by_name_prefix) and (not gs_is_non_numeric) and allowed_raw_type

    if normalized_type in _TERMINAL_TYPES:
        domain = "terminal"
    elif normalized_type in _FUSE_TYPES:
        domain = "fuse"
    elif allowed_raw_type:
        domain = "relay"
    else:
        domain = "other"

    reason = "ok"
    if removed_by_name_prefix:
        reason = "removed_name_prefix"
    elif gs_is_non_numeric:
        reason = "gs_non_numeric"
    elif not allowed_raw_type:
        reason = "type_not_allowed"

    return {
        "name": name_str,
        "raw_type": raw_type_str,
        "normalized_type": normalized_type,
        "gs_numeric": gs_numeric,
        "removed_by_name_prefix": removed_by_name_prefix,
        "gs_is_non_numeric": gs_is_non_numeric,
        "allowed_raw_type": allowed_raw_type,
        "would_be_processed": would_be_processed,
        "domain": domain,
        "reason": reason,
    }


def _sanity_check_type_helpers() -> None:
    """Manual smoke checks (not executed automatically).

    Example expectations:
    - normalize_type("2002-3201") -> "WAGO.2002-3201_ADV"
    - normalize_type("SE.A9F04604") -> "SE.A9F04604_ADV"
    - normalize_type("RE22R1AMR") -> "SE.RE22R1AMR_ADV"
    """

# -----------------------------------------------------------------------------
# Helpers
# -----------------------------------------------------------------------------
def _drop_unnamed_cols(df: pd.DataFrame) -> pd.DataFrame:
    return df.loc[:, ~df.columns.astype(str).str.match(r"^Unnamed")].copy()


def _ensure_cols(df: pd.DataFrame, cols: List[str]) -> None:
    missing = [c for c in cols if c not in df.columns]
    if missing:
        raise ValueError(f"Trūksta privalomų stulpelių: {', '.join(missing)}")


def _append_removed(parts: List[pd.DataFrame], df_part: pd.DataFrame, reason: str) -> None:
    if df_part.empty:
        return
    tmp = df_part.copy()
    tmp["Removed Reason"] = reason
    parts.append(tmp)


def _to_num(v: Any) -> float | None:
    x = pd.to_numeric(v, errors="coerce")
    return None if pd.isna(x) else float(x)


def _gs_int(v: Any) -> int | None:
    n = _to_num(v)
    return None if n is None else int(n)


def _gs_str(v: Any) -> str:
    n = _gs_int(v)
    return "" if n is None else str(n)


def _starts_with_any(s: pd.Series, prefixes: Tuple[str, ...]) -> pd.Series:
    return s.astype(str).str.startswith(prefixes, na=False)


def _next_free_gs(existing: set[int], preferred: int) -> int:
    g = preferred
    while g in existing:
        g += 1
    existing.add(g)
    return g


def _add_gs_prefix(df: pd.DataFrame) -> pd.DataFrame:
    """Prefix Name with +GS where missing."""
    gss = df["Group Sorting"].apply(_gs_str)
    has_gs = gss.ne("")
    not_prefixed = ~df["Name"].astype(str).str.startswith("+", na=False)
    df.loc[has_gs & not_prefixed, "Name"] = "+" + gss[has_gs & not_prefixed] + df.loc[
        has_gs & not_prefixed, "Name"
    ].astype(str)
    return df


def _apply_function_designation(df: pd.DataFrame) -> pd.DataFrame:
    def type_to_func(t: str) -> str:
        m = {
            "SE.A9F04604_ADV": "POWER",
            "WAGO.2002-1611/1000-541_ADV": "FUSES",
            "WAGO.2002-1611/1000-836_ADV": "FUSES",
            "SE.RGZE1S48M + SE.RXG22P7": "2POLE",
            "SE.RXZE2S114M + SE.RXM4GB2BD": "4POLE",
            "WAGO.2002-3201_ADV": "CONTROL",
            "WAGO.2002-3207_ADV": "CONTROL",
            "SE.RE22R1AMR_ADV": "2POLE",
            "FIN.39.00.8.230.8240_ADV": "1POLE",
            "SE.RE17LCBM_ADV": "TIMED_RELAYS",
        }
        return m.get(t, "")

    func_col = df["Type"].astype(str).map(type_to_func).fillna("")
    # Use original names for K-based detection because Name is prefixed later.
    name_src = df["_name_orig"] if "_name_orig" in df.columns else df["Name"]
    timed_name_mask = name_src.astype(str).str.contains(r"-K192A?\d+\b", regex=True, na=False)
    func_col.loc[timed_name_mask] = "TIMED_RELAYS"
    hasf = func_col.astype(str).ne("")
    df.loc[hasf, "Name"] = "=" + func_col[hasf].astype(str) + df.loc[hasf, "Name"].astype(str)
    return df


def _allocate_category_gs(
    df: pd.DataFrame, missing_default: int | pd.Series | None = None, fallback_col: str | None = None
) -> pd.DataFrame:
    """
    Fill missing Group Sorting values without overwriting existing ones.

    `missing_default` supplies the value (or per-row series) for rows where
    Group Sorting is missing. If `fallback_col` is provided, its numeric value
    is used when both Group Sorting and `missing_default` are missing.
    """
    df = df.copy()

    gs_numeric = pd.to_numeric(df["Group Sorting"], errors="coerce")

    if isinstance(missing_default, pd.Series):
        default_series = missing_default.reindex(df.index)
    elif missing_default is None:
        default_series = pd.Series([None] * len(df), index=df.index, dtype="float")
    else:
        default_series = pd.Series(missing_default, index=df.index)

    if fallback_col is not None and fallback_col in df.columns:
        fallback_series = pd.to_numeric(df[fallback_col], errors="coerce")
        default_series = default_series.combine_first(fallback_series)

    resolved = gs_numeric.combine_first(default_series)
    df["Group Sorting"] = resolved
    df["_gs_sort"] = pd.to_numeric(resolved, errors="coerce").fillna(10**9).astype(int)

    return df


def _terminal_base_name(name: str) -> str:
    s = str(name)
    s = re.sub(r"^=[A-Z]+", "", s, flags=re.IGNORECASE)
    s = s.lstrip("=")
    s = re.sub(r"^\+\d+", "", s)
    return s


def _validate_terminal_uniqueness(df: pd.DataFrame) -> None:
    if df.empty:
        return

    base = df["Name"].astype(str).apply(_terminal_base_name)
    combo = df["Type"].astype(str)
    gs_num = pd.to_numeric(df["Group Sorting"], errors="coerce")

    dup_mask = (
        pd.concat({"base": base, "type": combo, "gs": gs_num}, axis=1)
        .groupby(["base", "type"])["gs"]
        .transform("nunique")
        > 1
    )
    if dup_mask.any():
        bad = df.loc[dup_mask, ["Name", "Type", "Group Sorting"]]
        raise ValueError(
            "Terminal invariant violated: multiple GS values for the same terminal (Name+Type):\n"
            f"{bad.to_string(index=False)}"
        )


def _extract_x_digits(name: str) -> str | None:
    m = re.search(r"-X(\d+)", str(name))
    if not m:
        return None
    digits = m.group(1)
    return digits if len(digits) >= 2 else None


def _extract_x_subgroup(name: str) -> str | None:
    digits = _extract_x_digits(name)
    return digits[:2] if digits is not None else None


def _x_number_endswith_14(name: str) -> bool:
    digits = _extract_x_digits(name)
    return bool(digits and digits.endswith("14"))


def _terminal_sort_key(name: str) -> int:
    """
    Rikiavimo raktas, kad X192A* galėtume laikyti "tarp" X1922 ir X1953.
    Prielaida: X192A3 = 1923 (t.y. 192 + A3).
    """
    s = str(name)

    m = re.search(r"-X(\d{4})\b", s)
    if m:
        return int(m.group(1))

    m = re.search(r"-X(\d{3})A(\d+)\b", s)  # X192A3
    if m:
        return int(m.group(1) + m.group(2))  # "192"+"3" => 1923

    m = re.search(r"-X(\d+)", s)
    if m:
        return int(m.group(1))

    return 10**9


def _load_terminal_list_pe_requirements(terminal_list_bytes: bytes | None) -> Dict[int, int]:
    if not terminal_list_bytes:
        return {}
    try:
        term_df = pd.read_excel(BytesIO(terminal_list_bytes), sheet_name=0, engine="openpyxl")
    except Exception:
        return {}

    term_df = _drop_unnamed_cols(term_df)
    term_df.columns = term_df.columns.astype(str).str.strip()

    if "Conns." not in term_df.columns or "GROUP SORTING" not in term_df.columns or "Name" not in term_df.columns:
        return {}

    conns = term_df["Conns."].astype(str).str.strip()
    names = term_df["Name"].astype(str).str.strip()
    gs_numeric = pd.to_numeric(term_df["GROUP SORTING"], errors="coerce")
    gs_ok = gs_numeric.notna()
    pe_gs_ok = (gs_numeric % 10) == 5
    name_ok = names.str.match(r"^-X\d+\b")
    pe_mask = conns.eq("⏚") & name_ok & gs_ok & pe_gs_ok
    if not pe_mask.any():
        return {}

    gs_int = gs_numeric[pe_mask].astype(int)
    counts = gs_int.value_counts()
    return {int(gs): int((count + 2) // 3) for gs, count in counts.items()}


def _normalize_terminal_name(name: str) -> str:
    """
    Normalizuoja terminalo pavadinimą deduplikavimui, nuimdama
    automatiškai pridėtą +GS prefiksą ir lygybės ženklą.
    """
    s = str(name).lstrip("=")
    return re.sub(r"^\+\d+", "", s)


# -----------------------------------------------------------------------------
# Terminal rule: X192A insertion inside one base GS group
# -----------------------------------------------------------------------------
def apply_x192a_terminal_gs_rules(df: pd.DataFrame) -> pd.DataFrame:
    """
    Taisyklė:

    Bazė = ORIGINALUS GS iš įkelto failo (df['_gs_orig']).

    Vienoje bazinėje GS (pvz. 2010), jei yra X192A*:
      - iki X192A paliekam esamam GS
      - visi X192A* perkelti į naują GS = next_free(base+1) (pvz. 2011)
      - visi terminalai po X192A (pagal X numerį) perkelti į dar sekantį GS (pvz. 2012)
    """
    terminal_types = {"WAGO.2002-3201_ADV"}
    is_terminal = df["Type"].astype(str).isin(terminal_types)

    gs_num = pd.to_numeric(df["Group Sorting"], errors="coerce")
    existing_gs: set[int] = set(gs_num.dropna().astype(int).tolist())

    # Bazė = originalus GS (tik terminalams)
    if "_gs_orig" not in df.columns:
        df["_gs_orig"] = gs_num

    work = df[is_terminal & df["_gs_orig"].notna()].copy()
    if work.empty:
        return df

    work["_base_gs"] = work["_gs_orig"].astype(int)
    work["_xkey"] = work["Name"].astype(str).apply(_terminal_sort_key)
    work["_is_x192a"] = work["Name"].astype(str).str.contains(r"-X192A\d+\b", regex=True, na=False)

    for base_gs, grp in work.groupby("_base_gs", sort=True):
        if not grp["_is_x192a"].any():
            continue

        gs_for_x192a = _next_free_gs(existing_gs, int(base_gs) + 1)
        max_x192a_key = int(grp.loc[grp["_is_x192a"], "_xkey"].max())
        after_mask = (grp["_xkey"] > max_x192a_key) & (~grp["_is_x192a"])

        gs_for_after = None
        if after_mask.any():
            gs_for_after = _next_free_gs(existing_gs, gs_for_x192a + 1)

        df.loc[grp.loc[grp["_is_x192a"]].index, "Group Sorting"] = gs_for_x192a
        if gs_for_after is not None:
            df.loc[grp.loc[after_mask].index, "Group Sorting"] = gs_for_after

    return df


# -----------------------------------------------------------------------------
# Main processing
# -----------------------------------------------------------------------------
def process_excel(
    file_bytes: bytes, terminal_list_bytes: bytes | None = None
) -> Tuple[pd.DataFrame, pd.DataFrame, bytes, Dict[str, int]]:
    df = pd.read_excel(BytesIO(file_bytes), sheet_name=0, engine="openpyxl")
    df = _drop_unnamed_cols(df)
    input_rows = len(df)

    required = ["Name", "Type", "Quantity", "Group Sorting"]
    _ensure_cols(df, required)

    df["Name"] = df["Name"].astype(str)
    df["Type"] = df["Type"].astype(str)
    df["_name_orig"] = df["Name"]

    # ensure columns exist
    for col, default in [
        ("Accessories", ""),
        ("Quantity of accessories", 0),
        ("Accessories2", ""),
        ("Quantity of accessories2", 0),
        ("Designation", ""),
    ]:
        if col not in df.columns:
            df[col] = default

    removed_parts: List[pd.DataFrame] = []

    # -------------------------------------------------------------------------
    # STEP 0 – remove non-numeric Group Sorting (but not empty)
    # -------------------------------------------------------------------------
    gs_raw = df["Group Sorting"]
    gs_num = pd.to_numeric(gs_raw, errors="coerce")
    non_numeric = gs_raw.notna() & (gs_raw.astype(str).str.strip() != "") & gs_num.isna()
    _append_removed(removed_parts, df[non_numeric], "Removed: Group Sorting is not numeric")
    df = df[~non_numeric].copy()

    # Įsimenam ORIGINALŲ GS (tik skaičiai)
    df["_gs_orig"] = pd.to_numeric(df["Group Sorting"], errors="coerce")

    # -------------------------------------------------------------------------
    # STEP 1 – remove by Name prefixes
    # -------------------------------------------------------------------------
    remove_prefixes = ("-B", "-C", "-R", "-M", "-P", "-Q", "-S", "-W", "-T")
    m1 = _starts_with_any(df["Name"], remove_prefixes)
    _append_removed(removed_parts, df[m1], "Removed by Name prefix (-B/-C/-R/-M/-P/-Q/-S/-W/-T)")
    df = df[~m1].copy()

    # -------------------------------------------------------------------------
    # STEP 2 – invalid prefix highlight (not removed)
    # -------------------------------------------------------------------------
    valid_prefixes = ("-F", "-K", "-X")
    df["_highlight_invalid_prefix"] = ~_starts_with_any(df["Name"], valid_prefixes)

    # -------------------------------------------------------------------------
    # STEP 3 – Type allowlist (palik kaip pas tave projekte)
    # -------------------------------------------------------------------------
    allowed_types = {
        "2002-1611/1000-541",
        "2002-1611/1000-836",
        "2002-3201",
        "2002-3207",
        "RXZE2S114M",
        "RXM4GB2BD",
        "RXG22P7",
        "RXG22BD",
        "RGZE1S48M",
        "A9F04604",
        "RE17LCBM",
        "RE22R1AMR",
        "39.00.8.230.8240",
    }
    type_for_allow = df["Type"].astype(str).str.strip()
    type_for_allow = type_for_allow.str.replace(r"^(WAGO|SE|ABB)\.", "", regex=True)
    type_for_allow = type_for_allow.str.replace(r"_ADV$", "", regex=True)
    m3 = type_for_allow.isin(allowed_types)
    _append_removed(removed_parts, df[~m3], "Removed by Type filter (not in allowed list)")
    df = df[m3].copy()

    # -------------------------------------------------------------------------
    # STEP 4 – map types (WAGO/SE) + _ADV
    # -------------------------------------------------------------------------
    wago_types = {
        "2002-1611/1000-541",
        "2002-1611/1000-836",
        "2002-3201",
        "2002-3207",
        "249-116",
        "2002-991",
        "2002-3292",
    }
    mw = df["Type"].isin(wago_types)
    df.loc[mw, "Type"] = "WAGO." + df.loc[mw, "Type"].astype(str)

    se_map = {
        "RGZE1S48M": "SE.RGZE1S48M",
        "RXG22P7": "SE.RXG22P7",
        "RXG22BD": "SE.RXG22BD",
        "RXM4GB2BD": "SE.RXM4GB2BD",
        "RXZE2S114M": "SE.RXZE2S114M",
        "A9F04604": "SE.A9F04604",
    }
    df["Type"] = df["Type"].replace(se_map)

    adv_map = {
        "WAGO.2002-3207": "WAGO.2002-3207_ADV",
        "WAGO.2002-3201": "WAGO.2002-3201_ADV",
        "WAGO.2002-3292": "WAGO.2002-3292_ADV",
        "WAGO.2002-991": "WAGO.2002-991_ADV",
        "WAGO.249-116": "WAGO.249-116_ADV",
        "WAGO.2002-1611/1000-541": "WAGO.2002-1611/1000-541_ADV",
        "WAGO.2002-1611/1000-836": "WAGO.2002-1611/1000-836_ADV",
        "SE.A9F04604": "SE.A9F04604_ADV",
    }
    df["Type"] = df["Type"].replace(adv_map)

    relay_map = {
        "RE17LCBM": "SE.RE17LCBM_ADV",
        "RE22R1AMR": "SE.RE22R1AMR_ADV",
        "39.00.8.230.8240": "FIN.39.00.8.230.8240_ADV",
    }
    df["Type"] = df["Type"].replace(relay_map)

    required_pe_qty_by_gs = _load_terminal_list_pe_requirements(terminal_list_bytes)

    # -------------------------------------------------------------------------
    # STEP 4b – relay merge by Name (2POLE/4POLE)
    # -------------------------------------------------------------------------
    def _merge_relay(base_type: str, combined_type: str) -> None:
        nonlocal df
        mb = df["Type"].astype(str).eq(base_type)
        if not mb.any():
            return

        drop_idxs: List[int] = []
        # grupuojam tik pagal Name, kur bent viena eilutė yra base_type
        names_with_base = set(df.loc[mb, "Name"].astype(str).tolist())

        for name_val, grp in df[df["Name"].astype(str).isin(names_with_base)].groupby("Name", sort=False):
            idxs = grp.index.to_list()
            keep = None
            for i in idxs:
                if df.at[i, "Type"] == base_type:
                    keep = i
                    break
            if keep is None:
                keep = idxs[0]

            for i in idxs:
                if i != keep:
                    drop_idxs.append(i)

            df.at[keep, "Type"] = combined_type
            df.at[keep, "Quantity"] = 1

        if drop_idxs:
            _append_removed(removed_parts, df.loc[drop_idxs], "Removed: merged into relay combined row")
            df = df.drop(index=drop_idxs).copy()

    _merge_relay("SE.RGZE1S48M", "SE.RGZE1S48M + SE.RXG22P7")
    _merge_relay("SE.RXZE2S114M", "SE.RXZE2S114M + SE.RXM4GB2BD")

    # -------------------------------------------------------------------------
    # Category pipelines
    # -------------------------------------------------------------------------
    terminal_types = {"WAGO.2002-3201_ADV", "WAGO.2002-3207_ADV"}
    fuse_types = {"WAGO.2002-1611/1000-541_ADV", "WAGO.2002-1611/1000-836_ADV", "SE.A9F04604_ADV"}
    relay_types = set(df["Type"].unique()) - terminal_types - fuse_types

    terminal_df = df[df["Type"].isin(terminal_types)].copy()
    fuse_df = df[df["Type"].isin(fuse_types)].copy()
    relay_df = df[df["Type"].isin(relay_types)].copy()

    def process_relays(relay_data: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
        if relay_data.empty:
            return relay_data.copy(), pd.DataFrame(columns=list(relay_data.columns) + ["Removed Reason"])

        relay_data = relay_data.copy()

        relay_combo_types = {
            "SE.RGZE1S48M + SE.RXG22P7",
            "SE.RXZE2S114M + SE.RXM4GB2BD",
        }
        type_2pole = "SE.RGZE1S48M + SE.RXG22P7"
        type_4pole = "SE.RXZE2S114M + SE.RXM4GB2BD"
        type_1pole = "FIN.39.00.8.230.8240_ADV"

        gs_numeric_all = pd.to_numeric(relay_data["Group Sorting"], errors="coerce")
        existing_gs: set[int] = set(int(x) for x in gs_numeric_all.dropna())

        combo_mask = relay_data["Type"].astype(str).isin(relay_combo_types)

        all_2pole = relay_data["Type"].astype(str).isin([type_2pole, "SE.RE22R1AMR_ADV"])
        all_4pole = relay_data["Type"].astype(str).eq(type_4pole)

        def _assign_missing(
            mask: pd.Series, preferred_start: int, min_start: int | None = None, current_value: int | None = None
        ) -> int | None:
            missing_numeric = pd.to_numeric(relay_data.loc[mask, "Group Sorting"], errors="coerce").isna()
            if not missing_numeric.any():
                return None
            start = preferred_start if min_start is None else max(preferred_start, min_start)
            if current_value is not None and start < current_value:
                start = current_value
            if current_value is not None and start == current_value:
                existing_gs.add(int(current_value))
                relay_data.loc[missing_numeric[missing_numeric].index, "Group Sorting"] = int(current_value)
                return int(current_value)
            gs_value = _next_free_gs(existing_gs, start)
            relay_data.loc[missing_numeric[missing_numeric].index, "Group Sorting"] = gs_value
            return gs_value

        def _first_group_gs(mask: pd.Series) -> int | None:
            s = pd.to_numeric(relay_data.loc[mask, "Group Sorting"], errors="coerce")
            if s.notna().any():
                return int(s.dropna().iloc[0])
            return None

        gs_2_existing = _first_group_gs(all_2pole)
        gs_2pole_val = _assign_missing(
            all_2pole, preferred_start=gs_2_existing or 1, min_start=None, current_value=gs_2_existing
        )
        gs_2_final = _first_group_gs(all_2pole)

        gs_4_existing = _first_group_gs(all_4pole)
        min_for_4pole = (gs_2_final + 1) if gs_2_final is not None else 2
        gs_4pole_val = _assign_missing(
            all_4pole, preferred_start=gs_4_existing or 2, min_start=min_for_4pole, current_value=gs_4_existing
        )
        gs_4_final = _first_group_gs(all_4pole)

        # --- TIMED RELAYS GS (K192* sequential, K192A* one shared) -------------------
        k192_mask = relay_data["Name"].astype(str).str.contains(r"-K192(?!A)\d+\b", regex=True, na=False)
        k192a_mask = relay_data["Name"].astype(str).str.contains(r"-K192A\d+\b", regex=True, na=False)

        def _k192_num(name: str) -> int:
            m = re.search(r"-K192(?!A)(\d+)\b", str(name))
            return int(m.group(1)) if m else 10**9

        def _k192a_num(name: str) -> int:
            m = re.search(r"-K192A(\d+)\b", str(name))
            return int(m.group(1)) if m else 10**9

        # start GS for timed relays must be AFTER 2POLE and 4POLE (real numbers)
        gs_2_final = _first_group_gs(all_2pole)
        gs_4_final = _first_group_gs(all_4pole)
        start_gs = max([g for g in [gs_2_final, gs_4_final] if g is not None], default=0) + 1

        # ensure we respect any already used GS numbers
        gs_numeric_all = pd.to_numeric(relay_data["Group Sorting"], errors="coerce")
        existing_gs = set(int(x) for x in gs_numeric_all.dropna())

        # 1) K192* (no A): assign sequential GS per item, sorted by K-number
        k192_idxs = relay_data[k192_mask].index.tolist()
        k192_sorted = sorted(k192_idxs, key=lambda i: _k192_num(relay_data.at[i, "Name"]))

        gs_cursor = _next_free_gs(existing_gs, start_gs) if k192_sorted else start_gs

        for i, idx in enumerate(k192_sorted):
            # gs_cursor already free; assign and move cursor to next free
            relay_data.at[idx, "Group Sorting"] = gs_cursor
            if i < len(k192_sorted) - 1:
                gs_cursor = _next_free_gs(existing_gs, gs_cursor + 1)

        # 2) K192A*: ALL share one GS, which is the next GS after last K192*
        k192a_idxs = relay_data[k192a_mask].index.tolist()
        gs_k192 = gs_cursor if k192_sorted else None
        if k192a_idxs:
            # if there were K192*, next after last assigned; else start_gs
            preferred_a = (gs_cursor + 1) if k192_sorted else start_gs
            gs_k192a = _next_free_gs(existing_gs, preferred_a)
            for idx in k192a_idxs:
                relay_data.at[idx, "Group Sorting"] = gs_k192a
        else:
            gs_k192a = None

        one_pole_mask = relay_data["Type"].astype(str).eq(type_1pole)
        prev_for_1pole = max([g for g in [gs_2_final, gs_4_final, gs_k192, gs_k192a] if g is not None], default=0) + 1
        gs_one_pole_val = _assign_missing(one_pole_mask, preferred_start=prev_for_1pole, min_start=prev_for_1pole)
        gs_one_pole_final = _first_group_gs(one_pole_mask)

        def _relay_sort_value(name: str) -> int:
            m = re.search(r"-K192A?(\d+)", str(name))
            if m:
                digits = m.group(1)
                return int(f"192{digits}")
            return 10**9

        def _gs_int_safe(idx: int) -> int:
            gs_val = pd.to_numeric(relay_data.at[idx, "Group Sorting"], errors="coerce")
            return 10**9 if pd.isna(gs_val) else int(gs_val)

        timed_idxs = relay_data[k192_mask | k192a_mask].index.tolist()
        if timed_idxs:
            relay_data.loc[timed_idxs, "_terminal_sort"] = [_relay_sort_value(relay_data.at[i, "Name"]) for i in timed_idxs]

        name_src = relay_data["_name_orig"] if "_name_orig" in relay_data.columns else relay_data["Name"]
        k_num = name_src.astype(str).str.extract(r"-K(\d+)", expand=False)
        k_num_int = pd.to_numeric(k_num, errors="coerce")

        def _apply_relay_accessory(group_mask: pd.Series) -> None:
            if not group_mask.any():
                return
            group_idxs = relay_data[group_mask].index
            relay_data.loc[group_idxs, "Accessories"] = ""
            relay_data.loc[group_idxs, "Quantity of accessories"] = 0
            valid_k = k_num_int.loc[group_idxs].dropna()
            if valid_k.empty:
                return
            max_idx = valid_k.idxmax()
            # Use max K-number for deterministic “last” selection; DF order is unstable.
            relay_data.at[max_idx, "Accessories"] = "WAGO.249-116_ADV"
            relay_data.at[max_idx, "Quantity of accessories"] = 1

        _apply_relay_accessory(k192_mask | k192a_mask)
        _apply_relay_accessory(one_pole_mask)
        _apply_relay_accessory(all_2pole)
        _apply_relay_accessory(all_4pole)

        relay_data = _allocate_category_gs(relay_data, missing_default=1)
        return relay_data, pd.DataFrame(columns=list(relay_data.columns) + ["Removed Reason"])

    def process_fuses(fuse_data: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
        if fuse_data.empty:
            return fuse_data.copy(), pd.DataFrame(columns=list(fuse_data.columns) + ["Removed Reason"])

        fuse_data = fuse_data.copy()
        fuse_data["_orig_order"] = range(len(fuse_data))
        fuse_data["Accessories"] = fuse_data["Accessories"].fillna("")
        fuse_data["Accessories2"] = fuse_data["Accessories2"].fillna("")
        fuse_data["Quantity of accessories"] = (
            pd.to_numeric(fuse_data["Quantity of accessories"], errors="coerce").fillna(0).astype(int)
        )
        fuse_data["Quantity of accessories2"] = (
            pd.to_numeric(fuse_data["Quantity of accessories2"], errors="coerce").fillna(0).astype(int)
        )

        def _parse_fuse_key(name: str) -> dict | None:
            s = str(name)
            m_a = re.search(r"-F(\d{3})A(\d+)", s)
            if m_a:
                main = int(m_a.group(1))
                sub = int(m_a.group(2))
                return {"main": main, "sub": sub, "is_a": True}
            m = re.search(r"-F(\d+)(?:\.(\d+))?", s)
            if not m:
                return None
            return {"main": int(m.group(1)), "sub": int(m.group(2) or 0), "is_a": False}

        def _fuse_order_value(key: dict | None) -> float | None:
            if key is None:
                return None
            if key.get("is_a"):
                # F192A5 -> 1925 (matches terminal-style insertion)
                return int(f"{key['main']}{key['sub']}")
            return key["main"] + key["sub"] / 10

        fuse_data["_fuse_key"] = fuse_data["Name"].astype(str).apply(_parse_fuse_key)
        fuse_data["_fuse_order_val"] = fuse_data["_fuse_key"].apply(_fuse_order_value)

        def _fuse_sort_key(idx: int) -> Tuple[int, int, int]:
            order_val = fuse_data.at[idx, "_fuse_order_val"]
            orig = int(fuse_data.at[idx, "_orig_order"])
            if pd.isna(order_val):
                return (1, 10**9, orig)
            return (0, int(order_val), orig)

        fuse_types_target = {"WAGO.2002-1611/1000-541_ADV", "WAGO.2002-1611/1000-836_ADV"}
        is_541 = fuse_data["Type"].astype(str).eq("WAGO.2002-1611/1000-541_ADV")
        is_836 = fuse_data["Type"].astype(str).eq("WAGO.2002-1611/1000-836_ADV")

        existing_gs = pd.to_numeric(fuse_data["Group Sorting"], errors="coerce")
        existing_fuse_gs = sorted(set(int(x) for x in existing_gs.dropna() if 31 <= x <= 50))
        existing_non541_gs = sorted(
            set(int(x) for x in existing_gs[~is_541].dropna() if 31 <= x <= 50)
        )

        gs_541 = 31
        if 31 in existing_non541_gs:
            alternatives = [g for g in existing_fuse_gs if g != 31]
            if alternatives:
                gs_541 = min(alternatives)

        if is_541.any():
            fuse_data.loc[is_541, "Group Sorting"] = gs_541

        fuse_main = fuse_data["_fuse_key"].apply(lambda k: k.get("main") if isinstance(k, dict) else None)
        block32 = is_836 & fuse_main.apply(lambda m: m is not None and 900 <= m <= 999)
        fuse_data.loc[block32, "Group Sorting"] = 32

        base_gs = 33
        non_f9_mask = is_836 & (~block32)
        fuse_data.loc[non_f9_mask, "Group Sorting"] = base_gs

        f192a_mask = non_f9_mask & fuse_data["Name"].astype(str).str.contains(r"-F192A", regex=True, na=False)
        order_series = fuse_data["_fuse_order_val"]
        after_mask = pd.Series([False] * len(fuse_data), index=fuse_data.index)
        if f192a_mask.any():
            first_key = order_series[f192a_mask].min()
            last_key = order_series[f192a_mask].max()
            before_mask = non_f9_mask & (~f192a_mask) & order_series.notna() & (order_series < first_key)
            after_mask = non_f9_mask & (~f192a_mask) & order_series.notna() & (order_series > last_key)

            fuse_data.loc[f192a_mask, "Group Sorting"] = base_gs + 1
            if after_mask.any():
                fuse_data.loc[after_mask, "Group Sorting"] = base_gs + 2

        def _apply_accessories(block_mask: pd.Series, accessories: str | None, accessories2: str | None) -> None:
            if not block_mask.any():
                return
            idxs = block_mask[block_mask].index.tolist()
            last_idx = sorted(idxs, key=_fuse_sort_key)[-1]
            if accessories and str(fuse_data.at[last_idx, "Accessories"]).strip() == "":
                fuse_data.at[last_idx, "Accessories"] = accessories
                fuse_data.at[last_idx, "Quantity of accessories"] = 1
            if accessories2 and str(fuse_data.at[last_idx, "Accessories2"]).strip() == "":
                fuse_data.at[last_idx, "Accessories2"] = accessories2
                fuse_data.at[last_idx, "Quantity of accessories2"] = 1

        last_836_block = None
        if f192a_mask.any():
            if after_mask.any():
                last_836_block = after_mask
            else:
                last_836_block = f192a_mask
        else:
            last_836_block = non_f9_mask

        _apply_accessories(is_541, "WAGO.2002-991_ADV", "WAGO.249-116_ADV")
        _apply_accessories(block32, "WAGO.2002-991_ADV", None)
        _apply_accessories(last_836_block, "WAGO.2002-991_ADV", None)

        name_src = fuse_data["_name_orig"] if "_name_orig" in fuse_data.columns else fuse_data["Name"]
        crankcase_token = name_src.astype(str).str.extract(r"(-F\d{3})", expand=False)
        any_fxx8 = crankcase_token.fillna("").str.match(r"^-F\d{2}8$")
        valid_fxx8 = crankcase_token.fillna("").str.match(r"^-F[1-5][1-9]8$")
        # Exempt WAGO fuse variants that legitimately use F**8 codes.
        invalid_fxx8 = any_fxx8 & (~valid_fxx8) & ~(is_541 | is_836)
        if invalid_fxx8.any():
            to_remove = fuse_data.loc[invalid_fxx8].copy()
            to_remove = to_remove.drop(columns=[c for c in to_remove.columns if str(c).startswith("_")], errors="ignore")
            _append_removed(removed_parts, to_remove, "Removed: invalid F**8 fuse code")
            fuse_data = fuse_data[~invalid_fxx8].copy()
            valid_fxx8 = valid_fxx8.reindex(fuse_data.index, fill_value=False)
        missing_defaults = pd.Series(None, index=fuse_data.index, dtype="float")
        # Strict crankcase matching keeps unrelated F*** devices out of GS=0 (EPLAN ordering relies on GS).
        missing_defaults.loc[valid_fxx8] = 0
        fuse_data["_terminal_sort"] = fuse_data["_fuse_order_val"].where(
            fuse_data["_fuse_order_val"].notna(), 10**9 + fuse_data["_orig_order"]
        )
        fuse_data = _allocate_category_gs(fuse_data, missing_default=missing_defaults)
        fuse_data["_gs_sort"] = pd.to_numeric(fuse_data["Group Sorting"], errors="coerce").fillna(10**9).astype(int)
        fuse_data = fuse_data.drop(columns=["_orig_order", "_fuse_key", "_fuse_order_val"], errors="ignore")

        return fuse_data, pd.DataFrame(columns=list(fuse_data.columns) + ["Removed Reason"])

    def process_terminals(
        term_data: pd.DataFrame, pe_requirements: Dict[int, int]
    ) -> Tuple[pd.DataFrame, pd.DataFrame]:
        removed_local: List[pd.DataFrame] = []
        if term_data.empty:
            empty_removed = pd.DataFrame(columns=list(term_data.columns) + ["Removed Reason"])
            return term_data.copy(), empty_removed

        term_data = apply_x192a_terminal_gs_rules(term_data)

        term_data["Quantity"] = pd.to_numeric(term_data["Quantity"], errors="coerce").fillna(0)

        rows: List[dict] = []
        for _, r in term_data.iterrows():
            d = r.to_dict()
            t = str(d.get("Type", ""))
            qty = pd.to_numeric(d.get("Quantity", 0), errors="coerce")

            if t in terminal_types and pd.notna(qty) and qty > 1:
                n = int(qty)
                base = d.copy()
                base["Quantity"] = 1
                base["Designation"] = ""
                rows.append(base)

                for i in range(1, n):
                    rr = d.copy()
                    rr["Quantity"] = 1
                    rr["Designation"] = str(i)
                    rr["Accessories"] = ""
                    rr["Quantity of accessories"] = 0
                    rr["Accessories2"] = ""
                    rr["Quantity of accessories2"] = 0
                    rows.append(rr)
            else:
                d["Designation"] = d.get("Designation", "")
                rows.append(d)

        term_data = pd.DataFrame(rows)

        # PE mapping and /3 reduction
        pe_mask = term_data["Type"].astype(str).str.contains("2002-3207", na=False)
        pe_with_base = pe_mask & term_data["_gs_orig"].notna()

        if pe_with_base.any():
            base_values = term_data.loc[pe_with_base, "_gs_orig"].astype(int)
            unique_bases = sorted(base_values.unique())
            base_to_pe = {b: i + 1 for i, b in enumerate(unique_bases)}

            for idx in term_data.loc[pe_with_base].index:
                base = int(term_data.at[idx, "_gs_orig"])
                pe_id = base_to_pe[base]
                term_data.at[idx, "Name"] = f"+{base}-PE{pe_id}"
                term_data.at[idx, "Group Sorting"] = base

        name_s = term_data["Name"].astype(str)
        pe_name_mask = name_s.str.contains(r"-PE\d+\b", regex=True, na=False)
        if pe_name_mask.any():
            keep: List[int] = []
            for pe_name, grp in term_data[pe_name_mask].groupby("Name", sort=False):
                idxs = grp.index.to_list()
                cnt = len(idxs)
                base_gs = term_data.at[idxs[0], "_gs_orig"]
                base_gs_int = int(base_gs) if pd.notna(base_gs) else None
                need = (cnt + 2) // 3  # ceil(cnt/3)
                if base_gs_int is not None and base_gs_int in pe_requirements:
                    need = int(pe_requirements[base_gs_int])
                if need > cnt:
                    prototype = term_data.loc[idxs[0]].copy()
                    for extra_idx in range(cnt, need):
                        new_row = prototype.copy()
                        new_row["Quantity"] = 1
                        new_row["Accessories"] = ""
                        new_row["Quantity of accessories"] = 0
                        new_row["Accessories2"] = ""
                        new_row["Quantity of accessories2"] = 0
                        new_row["Designation"] = str(extra_idx) if extra_idx > 0 else ""
                        term_data = pd.concat([term_data, pd.DataFrame([new_row])], ignore_index=True)
                        idxs.append(term_data.index[-1])
                    cnt = len(idxs)

                for offset, idx in enumerate(idxs[:need]):
                    term_data.at[idx, "Designation"] = "" if offset == 0 else str(offset)

                removed_idxs = idxs[need:]
                if removed_idxs:
                    _append_removed(
                        removed_local,
                        term_data.loc[removed_idxs],
                        "Removed: PE reduction ceil(count/3)",
                    )

                keep.extend(idxs[:need])

            term_data = term_data[(~pe_name_mask) | (term_data.index.isin(keep))].copy()

        # Accessories
        term_data["Accessories"] = term_data["Accessories"].fillna("")
        term_data["Accessories2"] = term_data["Accessories2"].fillna("")
        term_data["Quantity of accessories"] = (
            pd.to_numeric(term_data["Quantity of accessories"], errors="coerce").fillna(0).astype(int)
        )
        term_data["Quantity of accessories2"] = (
            pd.to_numeric(term_data["Quantity of accessories2"], errors="coerce").fillna(0).astype(int)
        )

        # current GS for ordering
        gs_now = pd.to_numeric(term_data["Group Sorting"], errors="coerce").fillna(10**9).astype(int)
        term_data["_gs_sort"] = gs_now

        gs_base = pd.to_numeric(term_data.get("_gs_orig", pd.Series([None] * len(term_data))), errors="coerce")

        def _place_cover(candidate_idxs: List[int], gs_orig_value: int | None, subgroup_key: str | None) -> None:
            for idx in reversed(candidate_idxs):
                if str(term_data.at[idx, "Accessories"]).strip() == "":
                    term_data.at[idx, "Accessories"] = "WAGO.2002-3292_ADV"
                    term_data.at[idx, "Quantity of accessories"] = 1
                    return
            raise ValueError(
                f"Cannot place terminal cover (gs_orig={gs_orig_value}, subgroup={subgroup_key}): no empty Accessories slot"
            )

        for base in (1010, 1110):
            m = term_data["Type"].astype(str).isin(terminal_types) & gs_base.eq(base)
            if not m.any():
                continue
            tmp = term_data[m].copy()
            tmp["_sub"] = tmp["_name_orig"].astype(str).apply(_extract_x_subgroup)
            for _, grp in tmp.groupby("_sub", sort=False, dropna=False):
                idxs = grp.index.to_list()
                _place_cover(idxs, base, grp["_sub"].iloc[0])

        m1030 = term_data["Type"].astype(str).isin(terminal_types) & gs_base.eq(1030)
        if m1030.any():
            tmp = term_data[m1030].copy()
            tmp["_sub"] = tmp["_name_orig"].astype(str).apply(_extract_x_subgroup)
            for _, grp in tmp.groupby("_sub", sort=False, dropna=False):
                target_idx = None
                idxs = grp.index.to_list()
                for idx in idxs:
                    if _x_number_endswith_14(term_data.at[idx, "_name_orig"]):
                        target_idx = idx
                        break
                if target_idx is None:
                    target_idx = idxs[-1]
                target_pos = idxs.index(target_idx)
                candidate_idxs = idxs[: target_pos + 1]
                _place_cover(candidate_idxs, 1030, grp["_sub"].iloc[0])

        other_mask = (
            term_data["Type"].astype(str).isin(terminal_types) & gs_base.notna() & (~gs_base.isin([1010, 1110, 1030]))
        )
        if other_mask.any():
            xkey = term_data["Name"].astype(str).apply(_terminal_sort_key)
            for _, grp_idxs in term_data[other_mask].groupby(gs_base[other_mask], sort=True).groups.items():
                idxs = list(grp_idxs)
                sorted_idxs = sorted(
                    idxs,
                    key=lambda i: (
                        int(term_data.at[i, "_gs_sort"]),
                        int(xkey.loc[i]),
                        str(term_data.at[i, "Designation"]),
                    ),
                )
                gs_val = term_data.at[sorted_idxs[0], "_gs_orig"]
                _place_cover(sorted_idxs, gs_val if pd.notna(gs_val) else None, None)

        # remove duplicate terminals within same original GS
        term_data["_terminal_sort"] = term_data["Name"].astype(str).apply(_terminal_sort_key)

        is_terminal_mask = term_data["Type"].astype(str).isin(terminal_types)
        if is_terminal_mask.any():
            dup_idxs: List[int] = []
            term_df = term_data[is_terminal_mask & term_data["_gs_orig"].notna()].copy()
            term_df["_norm_term_name"] = term_df["Name"].apply(_normalize_terminal_name)

            group_keys = ["_gs_orig", "Type", "_norm_term_name", "Designation"]
            for _, grp in term_df.groupby(group_keys, sort=False):
                keep_idx = grp.sort_values(["_gs_sort", "_terminal_sort", "Name"], kind="stable").index[0]
                dup_idxs.extend(i for i in grp.index if i != keep_idx)

            if dup_idxs:
                _append_removed(
                    removed_local, term_data.loc[dup_idxs], "Removed: duplicate terminal within original GS"
                )
                term_data = term_data.drop(index=dup_idxs).copy()

        term_data = _allocate_category_gs(term_data, missing_default=None, fallback_col="_gs_orig")
        term_data = _add_gs_prefix(term_data)

        # Terminal duplicate bugfix across GS
        term_data["_terminal_base"] = term_data["Name"].astype(str).apply(_terminal_base_name)
        term_data["_gs_num_final"] = pd.to_numeric(term_data["Group Sorting"], errors="coerce")
        duplicate_idxs: List[int] = []
        for (_, _type), grp in term_data.groupby(["_terminal_base", "Type"], sort=False):
            if grp["_gs_num_final"].nunique() <= 1:
                continue
            keep_idx = grp.sort_values(["_gs_num_final", "_terminal_sort", "Name"], kind="stable").index[0]
            duplicate_idxs.extend(i for i in grp.index if i != keep_idx)

        if duplicate_idxs:
            _append_removed(
                removed_local,
                term_data.loc[duplicate_idxs],
                "Removed: duplicate terminal (Name+Type) across multiple GS",
            )
            term_data = term_data.drop(index=duplicate_idxs).copy()

        term_data.drop(columns=["_terminal_base", "_gs_num_final"], inplace=True, errors="ignore")
        _validate_terminal_uniqueness(term_data)

        return term_data, (pd.concat(removed_local, ignore_index=True) if removed_local else pd.DataFrame())

    relay_clean, relay_removed = process_relays(relay_df)
    fuse_clean, fuse_removed = process_fuses(fuse_df)
    terminal_clean, terminal_removed = process_terminals(terminal_df, required_pe_qty_by_gs)

    cleaned_df = pd.concat([relay_clean, fuse_clean, terminal_clean], ignore_index=True)
    removed_df_parts = [p for p in [relay_removed, fuse_removed, terminal_removed] if p is not None and not p.empty]
    removed_df = pd.concat(removed_parts + removed_df_parts, ignore_index=True) if (removed_parts or removed_df_parts) else pd.DataFrame()
    if removed_df.empty:
        removed_df = pd.DataFrame(columns=list(df.columns) + ["Removed Reason"])
    elif "Removed Reason" not in removed_df.columns:
        removed_df["Removed Reason"] = ""

    cleaned_df = _add_gs_prefix(cleaned_df)
    cleaned_df = _apply_function_designation(cleaned_df)

    _validate_terminal_uniqueness(cleaned_df[cleaned_df["Type"].astype(str).isin(terminal_types)])

    # Final sort
    is_terminal_sorted = cleaned_df["Type"].astype(str).isin(terminal_types)
    if "_terminal_sort" in cleaned_df.columns:
        cleaned_df["_terminal_sort"] = cleaned_df["_terminal_sort"].fillna(
            cleaned_df["Name"].astype(str).apply(_terminal_sort_key)
        )
    else:
        cleaned_df["_terminal_sort"] = cleaned_df["Name"].astype(str).apply(_terminal_sort_key)
    cleaned_df["_terminal_sort_order"] = is_terminal_sorted.map({True: 0, False: 1})
    cleaned_df["_gs_sort"] = pd.to_numeric(cleaned_df["Group Sorting"], errors="coerce").fillna(10**9).astype(int)

    relay_type_priority = {
        "SE.RGZE1S48M + SE.RXG22P7": 0,  # 2POLE before 4POLE when GS is equal
        "SE.RE22R1AMR_ADV": 0,
        "SE.RXZE2S114M + SE.RXM4GB2BD": 1,
        "FIN.39.00.8.230.8240_ADV": 3,
    }
    timed_mask = cleaned_df["Name"].astype(str).str.contains(r"-K192A?\d+", regex=True, na=False)
    type_order_series = cleaned_df["Type"].map(relay_type_priority)
    type_order_series.loc[timed_mask] = 2
    cleaned_df["_relay_type_order"] = type_order_series.fillna(5).astype(int)

    cleaned_df = cleaned_df.sort_values(
        ["_gs_sort", "_terminal_sort_order", "_relay_type_order", "_terminal_sort", "Name"],
        kind="stable",
    )

    # Single secondary accessory on the globally last WAGO.2002-3201_ADV terminal
    mask_primary_terminal = cleaned_df["Type"].astype(str).eq("WAGO.2002-3201_ADV")
    if mask_primary_terminal.any():
        last_idx = cleaned_df[mask_primary_terminal].index[-1]
        if str(cleaned_df.at[last_idx, "Accessories2"]).strip() == "":
            cleaned_df.at[last_idx, "Accessories2"] = "WAGO.249-116_ADV"
            cleaned_df.at[last_idx, "Quantity of accessories2"] = 1

    cleaned_df = cleaned_df.drop(
        columns=["_gs_sort", "_terminal_sort", "_terminal_sort_order", "_relay_type_order"], errors="ignore"
    )

    # Workbook output
    wb = Workbook()
    ws_c = wb.active
    ws_c.title = "Cleaned"
    ws_r = wb.create_sheet("Removed")

    out_clean = cleaned_df.drop(columns=["_highlight_invalid_prefix"], errors="ignore")
    out_clean = out_clean.drop(columns=["_gs_orig", "_name_orig"], errors="ignore")
    for row in dataframe_to_rows(out_clean, index=False, header=True):
        ws_c.append(row)

    for r in range(2, ws_c.max_row + 1):
        c = ws_c.cell(row=r, column=1)
        c.number_format = "@"
        c.data_type = "s"

    if "_highlight_invalid_prefix" in cleaned_df.columns:
        flags = cleaned_df["_highlight_invalid_prefix"].tolist()
        for excel_row, bad in enumerate(flags, start=2):
            if bool(bad):
                for col in range(1, ws_c.max_column + 1):
                    ws_c.cell(row=excel_row, column=col).fill = YELLOW_FILL

    removed_out = removed_df.drop(
        columns=["_gs_orig", "_name_orig", "_terminal_sort", "_terminal_sort_order"], errors="ignore"
    )
    for row in dataframe_to_rows(removed_out, index=False, header=True):
        ws_r.append(row)

    if ws_r.max_row >= 1:
        headers = [ws_r.cell(row=1, column=c).value for c in range(1, ws_r.max_column + 1)]
        if "Name" in headers:
            name_col = headers.index("Name") + 1
            for r in range(2, ws_r.max_row + 1):
                c = ws_r.cell(row=r, column=name_col)
                c.number_format = "@"
                c.data_type = "s"

    bio = BytesIO()
    wb.save(bio)
    out_bytes = bio.getvalue()

    stats = {
        "input_rows": input_rows,
        "cleaned_rows": len(out_clean),
        "removed_rows": len(removed_df),
    }

    return out_clean, removed_df, out_bytes, stats
