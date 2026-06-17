from __future__ import annotations

from copy import copy
from io import BytesIO
from typing import Any

import pandas as pd


_MARKINGS_BLOCK_HEADER_ROW_COUNT = 2
_BLOCK_HEADER_TITLES = {
    "FUSE STRIP",
    "RELAYS STRIP",
    "Components",
    "FUSES",
    "RELAYS",
    "TERMINALS STRIP",
    "TERMINALS TMB",
    "CABLES",
    "POWER WIRES",
}
_TERMINAL_TMB_BLOCK_SUBTITLE = "2009-115"
_TERMINAL_STRIP_BLOCK_SUBTITLE = "Wago 2009-110"
_WIRE_CABLES_BLOCK_TITLE = "CABLES"
_WIRE_CABLES_BLOCK_SUBTITLE = "Phoenix WML 6"
_WIRE_POWER_WIRES_BLOCK_TITLE = "POWER WIRES"
_WIRE_POWER_WIRES_BLOCK_SUBTITLE = "Phoenix UC-WMT (15x4)"
_COMPONENT_MARKING_HEADER_BLOCK_RANGES = ("A1:B2", "E1:F2", "J1:L2", "O1:O2", "R1:R2")
_TERMINAL_MARKING_HEADER_BLOCK_RANGES = ("A1:D2", "G1:H2")
_WIRE_MARKING_HEADER_BLOCK_RANGES = ("A1:E2", "H1:L2")
_COMPONENT_MARKING_HEADER_MERGE_RANGES = ("A1:B1", "A2:B2", "E1:F1", "E2:F2", "J1:L1", "J2:L2")
_TERMINAL_MARKING_HEADER_MERGE_RANGES = ("A1:D1", "A2:D2", "G1:H1", "G2:H2")
_WIRE_MARKING_HEADER_MERGE_RANGES = ("A1:E1", "A2:E2", "H1:L1", "H2:L2")


def _stringify_cell(value: Any) -> str:
    """Normalize Excel-ish cell values to stripped strings for comparisons/export."""
    if pd.isna(value):
        return ""
    return str(value).strip()


def _make_excel_text_safe(value: Any) -> str:
    """Keep exported values as plain strings for text-formatted Excel cells."""
    return _stringify_cell(value)


def _set_excel_font(cell: Any, *, name: str = "Arial", bold: bool | None = None) -> None:
    """Update one cell font while disabling theme-font fallback in Excel."""
    updated_font = copy(cell.font)
    updated_font.name = name
    updated_font.scheme = None
    updated_font.family = 2
    if bold is not None:
        updated_font.bold = bold
    cell.font = updated_font


def _apply_worksheet_arial_font_formatting(worksheet: Any) -> None:
    """Render every visible worksheet cell in Arial with a non-bold base style."""
    if worksheet.max_row < 1 or worksheet.max_column < 1:
        return

    for row in worksheet.iter_rows(
        min_row=1,
        max_row=worksheet.max_row,
        min_col=1,
        max_col=worksheet.max_column,
    ):
        for cell in row:
            if cell.__class__.__name__ == "MergedCell":
                continue
            _set_excel_font(cell, name="Arial", bold=False)


def _apply_workbook_arial_font_formatting(workbook: Any) -> None:
    """Apply Arial to every used cell across the whole exported workbook."""
    for worksheet in workbook.worksheets:
        _apply_worksheet_arial_font_formatting(worksheet)


def _find_worksheet_header_columns(
    worksheet: Any,
    *,
    header_row: int = 1,
) -> dict[str, list[int]]:
    """Map header labels from one worksheet row to one or more column indexes."""
    if worksheet.max_row < header_row or worksheet.max_column < 1:
        return {}

    header_columns: dict[str, list[int]] = {}
    for column_index in range(1, worksheet.max_column + 1):
        header_value = _stringify_cell(worksheet.cell(row=header_row, column=column_index).value)
        if header_value == "":
            continue
        header_columns.setdefault(header_value, []).append(column_index)
    return header_columns


def _set_non_empty_column_cells_bold(
    worksheet: Any,
    column_indexes: list[int],
    *,
    data_start_row: int = 2,
) -> None:
    """Bold non-empty data cells in the requested worksheet columns."""
    if worksheet.max_row < data_start_row:
        return

    for column_index in column_indexes:
        for row_index in range(data_start_row, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            if _stringify_cell(cell.value) == "":
                continue
            _set_excel_font(cell, name="Arial", bold=True)


_NON_BOLD_CM_SECTION_LABELS = {
    "Fuses 24VDC",
    "Fuses 230VAC",
    "Relays 2_Pole",
    "Relays 4_Pole",
    "Relays Timed",
    "Relays 1_Pole",
    "24VDC_2_pole",
    "230VAC_2_pole",
    "24VDC_4_pole",
    "230VAC_4_pole",
    "230VAC_4A_pole",
    "Timed relays",
    "Buttons",
    "Other",
}

_NON_BOLD_COMPONENT_STRIP_LABELS = {
    "24VDC",
    "230VAC",
    "START",
    "STOP",
}


def _set_matching_non_empty_column_cells_bold(
    worksheet: Any,
    column_indexes: list[int],
    matcher: Any,
    *,
    data_start_row: int = 2,
) -> None:
    """Bold non-empty data cells in the requested columns when they match one rule."""
    if worksheet.max_row < data_start_row:
        return

    for column_index in column_indexes:
        for row_index in range(data_start_row, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            text_value = _stringify_cell(cell.value)
            if text_value == "" or not matcher(text_value):
                continue
            _set_excel_font(cell, name="Arial", bold=True)


def _is_component_marking_sheet_name(sheet_name: str) -> bool:
    """Return whether one worksheet title represents a Component Marking sheet."""
    normalized_name = _stringify_cell(sheet_name)
    return normalized_name == "Component Marking" or normalized_name.endswith(" Component Marking")


def _is_component_strip_sheet_name(sheet_name: str) -> bool:
    """Return whether one worksheet title represents a Component Strip sheet."""
    normalized_name = _stringify_cell(sheet_name)
    return normalized_name == "Component Strip" or normalized_name.endswith(" Component Strip")


def _is_component_cm_sheet_name(sheet_name: str) -> bool:
    """Return whether one worksheet title represents a CM sheet."""
    normalized_name = _stringify_cell(sheet_name)
    return normalized_name == "CM" or normalized_name.endswith(" CM")


def _apply_component_marking_label_bold_formatting(worksheet: Any) -> None:
    """Bold raw or final-layout Component Marking labels while skipping non-marking section text."""
    header_columns = _find_worksheet_header_columns(worksheet, header_row=1)
    has_raw_name_format = bool(header_columns.get("Name"))
    final_header_columns = _find_worksheet_header_columns(worksheet, header_row=3)
    has_block_header_layout = bool(final_header_columns.get("Text")) and any(
        final_header_columns.get(column_name)
        for column_name in ("Mounting plate", "Component", "Door")
    )
    has_combined_final_format = bool(header_columns.get("Text")) and any(
        header_columns.get(column_name)
        for column_name in ("Mounting plate", "Component", "Door")
    )

    if has_raw_name_format:
        _set_non_empty_column_cells_bold(worksheet, header_columns.get("Name", []))
        return

    if has_block_header_layout:
        _set_matching_non_empty_column_cells_bold(
            worksheet,
            final_header_columns.get("Text", []),
            lambda text_value: text_value not in _NON_BOLD_COMPONENT_STRIP_LABELS,
            data_start_row=4,
        )
        _set_non_empty_column_cells_bold(
            worksheet,
            final_header_columns.get("Mounting plate", []),
            data_start_row=4,
        )
        _set_matching_non_empty_column_cells_bold(
            worksheet,
            final_header_columns.get("Component", []),
            lambda text_value: text_value not in _NON_BOLD_CM_SECTION_LABELS,
            data_start_row=4,
        )
        _set_non_empty_column_cells_bold(
            worksheet,
            final_header_columns.get("Door", []),
            data_start_row=4,
        )
        _set_matching_non_empty_column_cells_bold(
            worksheet,
            header_columns.get("FUSES", []),
            lambda text_value: text_value not in _NON_BOLD_CM_SECTION_LABELS,
            data_start_row=3,
        )
        _set_matching_non_empty_column_cells_bold(
            worksheet,
            header_columns.get("RELAYS", []),
            lambda text_value: text_value not in _NON_BOLD_CM_SECTION_LABELS,
            data_start_row=3,
        )
        return

    if not has_combined_final_format:
        return

    _set_matching_non_empty_column_cells_bold(
        worksheet,
        header_columns.get("Text", []),
        lambda text_value: text_value not in _NON_BOLD_COMPONENT_STRIP_LABELS,
    )
    _set_non_empty_column_cells_bold(worksheet, header_columns.get("Mounting plate", []))
    _set_matching_non_empty_column_cells_bold(
        worksheet,
        header_columns.get("Component", []),
        lambda text_value: text_value not in _NON_BOLD_CM_SECTION_LABELS,
    )
    _set_non_empty_column_cells_bold(worksheet, header_columns.get("Door", []))
    _set_matching_non_empty_column_cells_bold(
        worksheet,
        header_columns.get("FUSES", []),
        lambda text_value: text_value not in _NON_BOLD_CM_SECTION_LABELS,
    )
    _set_matching_non_empty_column_cells_bold(
        worksheet,
        header_columns.get("RELAYS", []),
        lambda text_value: text_value not in _NON_BOLD_CM_SECTION_LABELS,
    )


def _apply_component_cm_label_bold_formatting(worksheet: Any) -> None:
    """Bold CM marking entries while keeping section labels non-bold."""
    header_columns = _find_worksheet_header_columns(worksheet)
    _set_non_empty_column_cells_bold(worksheet, header_columns.get("Mounting plate", []))
    _set_non_empty_column_cells_bold(worksheet, header_columns.get("Door", []))
    _set_matching_non_empty_column_cells_bold(
        worksheet,
        header_columns.get("Component", []),
        lambda text_value: text_value not in _NON_BOLD_CM_SECTION_LABELS,
    )


def _apply_component_strip_label_bold_formatting(worksheet: Any) -> None:
    """Bold non-empty Text-column values in Component Strip sheets."""
    header_columns = _find_worksheet_header_columns(worksheet)
    _set_non_empty_column_cells_bold(worksheet, header_columns.get("Text", []))


def _apply_workbook_component_label_bold_formatting(workbook: Any) -> None:
    """Bold only the component-related marking label cells in the workbook."""
    for worksheet in workbook.worksheets:
        if _is_component_marking_sheet_name(worksheet.title):
            _apply_component_marking_label_bold_formatting(worksheet)
        elif _is_component_strip_sheet_name(worksheet.title):
            _apply_component_strip_label_bold_formatting(worksheet)
        elif _is_component_cm_sheet_name(worksheet.title):
            _apply_component_cm_label_bold_formatting(worksheet)


def _write_markings_block_header(
    worksheet: Any,
    *,
    startrow: int,
    startcol: int,
    block_width: int,
    title: str,
    subtitle: str,
) -> None:
    """Write one two-row merged header block before pandas writes data below it."""
    from openpyxl.styles import Alignment, Border, Side

    start_row_number = startrow + 1
    end_column_number = startcol + max(block_width, 1)
    title_cell = worksheet.cell(row=start_row_number, column=startcol + 1)
    subtitle_cell = worksheet.cell(row=start_row_number + 1, column=startcol + 1)
    title_cell.value = title
    subtitle_cell.value = subtitle
    for cell in (title_cell, subtitle_cell):
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=Side(style="thin"))
    if end_column_number > startcol + 1:
        worksheet.merge_cells(
            start_row=start_row_number,
            start_column=startcol + 1,
            end_row=start_row_number,
            end_column=end_column_number,
        )
        worksheet.merge_cells(
            start_row=start_row_number + 1,
            start_column=startcol + 1,
            end_row=start_row_number + 1,
            end_column=end_column_number,
        )


def _write_terminal_markings_sheet(
    writer: Any,
    sheet_name: str,
    terminal_tmb_df: pd.DataFrame,
    terminal_strip_df: pd.DataFrame,
) -> None:
    """Write Terminal TMB and Terminal Strip tables side by side on one sheet."""
    tmb_export_df = terminal_tmb_df.copy()
    strip_export_df = terminal_strip_df.copy()
    for export_df in (tmb_export_df, strip_export_df):
        for column_name in export_df.columns:
            export_df[column_name] = export_df[column_name].apply(_make_excel_text_safe)

    strip_start_col = len(tmb_export_df.columns) + 2
    data_start_row = _MARKINGS_BLOCK_HEADER_ROW_COUNT
    worksheet = writer.book.create_sheet(title=sheet_name)
    writer.sheets[sheet_name] = worksheet
    _write_markings_block_header(
        worksheet,
        startrow=0,
        startcol=0,
        block_width=max(1, len(tmb_export_df.columns)),
        title="TERMINALS TMB",
        subtitle=_TERMINAL_TMB_BLOCK_SUBTITLE,
    )
    _write_markings_block_header(
        worksheet,
        startrow=0,
        startcol=strip_start_col,
        block_width=max(1, len(strip_export_df.columns)),
        title="TERMINALS STRIP",
        subtitle=_TERMINAL_STRIP_BLOCK_SUBTITLE,
    )
    tmb_export_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=data_start_row, startcol=0)
    strip_export_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=data_start_row, startcol=strip_start_col)

    for row in worksheet.iter_rows():
        for cell in row:
            if cell.__class__.__name__ == "MergedCell":
                continue
            cell.number_format = "@"
            if cell.value is None:
                cell.value = ""
            else:
                cell.value = str(cell.value)


def _write_wire_markings_sheet(
    writer: Any,
    sheet_name: str,
    cables_df: pd.DataFrame,
    power_wires_df: pd.DataFrame,
) -> None:
    """Write Cable and Power Wires tables side by side on one sheet."""
    cables_export_df = cables_df.copy()
    power_wires_export_df = power_wires_df.copy()
    for export_df in (cables_export_df, power_wires_export_df):
        for column_name in export_df.columns:
            export_df[column_name] = export_df[column_name].apply(_make_excel_text_safe)

    power_wires_start_col = len(cables_export_df.columns) + 2
    data_start_row = _MARKINGS_BLOCK_HEADER_ROW_COUNT
    worksheet = writer.book.create_sheet(title=sheet_name)
    writer.sheets[sheet_name] = worksheet
    _write_markings_block_header(
        worksheet,
        startrow=0,
        startcol=0,
        block_width=max(1, len(cables_export_df.columns)),
        title=_WIRE_CABLES_BLOCK_TITLE,
        subtitle=_WIRE_CABLES_BLOCK_SUBTITLE,
    )
    _write_markings_block_header(
        worksheet,
        startrow=0,
        startcol=power_wires_start_col,
        block_width=max(1, len(power_wires_export_df.columns)),
        title=_WIRE_POWER_WIRES_BLOCK_TITLE,
        subtitle=_WIRE_POWER_WIRES_BLOCK_SUBTITLE,
    )
    cables_export_df.to_excel(
        writer,
        sheet_name=sheet_name,
        index=False,
        startrow=data_start_row,
        startcol=0,
    )
    power_wires_export_df.to_excel(
        writer,
        sheet_name=sheet_name,
        index=False,
        startrow=data_start_row,
        startcol=power_wires_start_col,
    )

    for row in worksheet.iter_rows():
        for cell in row:
            if cell.__class__.__name__ == "MergedCell":
                continue
            cell.number_format = "@"
            if cell.value is None:
                cell.value = ""
            else:
                cell.value = str(cell.value)


def _apply_workbook_block_header_formatting(workbook: Any) -> None:
    """Bold the first header row for known block titles in the main markings layouts."""
    for worksheet in workbook.worksheets:
        for column_index in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=column_index)
            if _stringify_cell(cell.value) not in _BLOCK_HEADER_TITLES:
                continue
            _set_excel_font(cell, name="Arial", bold=True)


def _copy_worksheet_cell_style_and_value(source_cell: Any, target_cell: Any) -> None:
    """Copy one worksheet cell value and visible style to another cell."""
    target_cell.value = source_cell.value
    target_cell.font = copy(source_cell.font)
    target_cell.fill = copy(source_cell.fill)
    target_cell.border = copy(source_cell.border)
    target_cell.alignment = copy(source_cell.alignment)
    target_cell.number_format = source_cell.number_format
    target_cell.protection = copy(source_cell.protection)


def _clear_worksheet_cell(cell: Any) -> None:
    """Reset one worksheet cell to an empty visible value without removing layout."""
    from openpyxl.styles import PatternFill

    cell.value = ""
    cell.fill = PatternFill(fill_type=None)


def _shift_worksheet_column_down_one(
    worksheet: Any,
    *,
    column_index: int,
    start_row: int,
) -> None:
    """Shift one worksheet column down by one row from one fixed start point."""
    if worksheet.max_row < start_row:
        return
    if _stringify_cell(worksheet.cell(row=start_row, column=column_index).value) == "":
        return

    for row_index in range(worksheet.max_row, start_row - 1, -1):
        source_cell = worksheet.cell(row=row_index, column=column_index)
        target_cell = worksheet.cell(row=row_index + 1, column=column_index)
        _copy_worksheet_cell_style_and_value(source_cell, target_cell)

    _clear_worksheet_cell(worksheet.cell(row=start_row, column=column_index))


def _merge_worksheet_ranges_if_needed(worksheet: Any, merge_ranges: tuple[str, ...]) -> None:
    """Ensure the requested worksheet ranges are merged before header styling runs."""
    merged_range_values = {str(merged_range) for merged_range in worksheet.merged_cells.ranges}
    for cell_range in merge_ranges:
        if cell_range in merged_range_values:
            continue
        worksheet.merge_cells(cell_range)


def _apply_center_alignment_to_range(worksheet: Any, cell_range: str) -> None:
    """Center visible header cells inside one fixed worksheet range."""
    from openpyxl.styles import Alignment

    for row in worksheet[cell_range]:
        for cell in row:
            if cell.__class__.__name__ == "MergedCell":
                continue
            cell.alignment = Alignment(horizontal="center")


def _apply_outer_thin_border_to_range(worksheet: Any, cell_range: str) -> None:
    """Apply one continuous thin outer border around one fixed worksheet range."""
    from openpyxl.styles import Border, Side
    from openpyxl.utils.cell import range_boundaries

    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    thin_side = Side(style="thin")
    for row_index in range(min_row, max_row + 1):
        for column_index in range(min_col, max_col + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            cell.border = Border(
                left=thin_side if column_index == min_col else Side(style=None),
                right=thin_side if column_index == max_col else Side(style=None),
                top=thin_side if row_index == min_row else Side(style=None),
                bottom=thin_side if row_index == max_row else Side(style=None),
            )


def _is_final_component_marking_sheet(worksheet: Any) -> bool:
    """Return whether one worksheet is the final combined Component Marking layout."""
    return (
        _stringify_cell(worksheet["A1"].value) == "FUSE STRIP"
        and _stringify_cell(worksheet["J1"].value) == "Components"
        and _stringify_cell(worksheet["O1"].value) == "FUSES"
        and _stringify_cell(worksheet["R1"].value) == "RELAYS"
    )


def _is_final_terminal_marking_sheet(worksheet: Any) -> bool:
    """Return whether one worksheet is the final combined Terminal markings layout."""
    return (
        _stringify_cell(worksheet["A1"].value) == "TERMINALS TMB"
        and _stringify_cell(worksheet["G1"].value) == "TERMINALS STRIP"
    )


def _is_final_wire_marking_sheet(worksheet: Any) -> bool:
    """Return whether one worksheet is the header-formatted Cable Marking layout."""
    return (
        _stringify_cell(worksheet["A1"].value) == _WIRE_CABLES_BLOCK_TITLE
        and _stringify_cell(worksheet["H1"].value) == _WIRE_POWER_WIRES_BLOCK_TITLE
    )


def _apply_component_marking_header_polish(worksheet: Any) -> None:
    """Apply fixed final Component Marking header alignment, borders, and width tweaks."""
    _shift_worksheet_column_down_one(worksheet, column_index=15, start_row=3)
    _shift_worksheet_column_down_one(worksheet, column_index=18, start_row=3)
    worksheet["O3"] = ""
    worksheet["R3"] = ""
    worksheet.column_dimensions["O"].width = 17
    worksheet.column_dimensions["R"].width = 17
    worksheet.column_dimensions["R"].hidden = True
    _merge_worksheet_ranges_if_needed(worksheet, _COMPONENT_MARKING_HEADER_MERGE_RANGES)
    for cell_range in _COMPONENT_MARKING_HEADER_BLOCK_RANGES:
        _apply_center_alignment_to_range(worksheet, cell_range)
        _apply_outer_thin_border_to_range(worksheet, cell_range)


def _apply_terminal_marking_header_polish(worksheet: Any) -> None:
    """Apply fixed final Terminal markings header alignment, borders, and width tweaks."""
    worksheet.column_dimensions["A"].width = 14
    worksheet.column_dimensions["G"].width = 10
    worksheet.column_dimensions["H"].width = 10
    _merge_worksheet_ranges_if_needed(worksheet, _TERMINAL_MARKING_HEADER_MERGE_RANGES)
    for cell_range in _TERMINAL_MARKING_HEADER_BLOCK_RANGES:
        _apply_center_alignment_to_range(worksheet, cell_range)
        _apply_outer_thin_border_to_range(worksheet, cell_range)


def _apply_wire_marking_header_polish(worksheet: Any) -> None:
    """Apply fixed final Cable Marking header alignment and borders."""
    _merge_worksheet_ranges_if_needed(worksheet, _WIRE_MARKING_HEADER_MERGE_RANGES)
    for cell_range in _WIRE_MARKING_HEADER_BLOCK_RANGES:
        _apply_center_alignment_to_range(worksheet, cell_range)
        _apply_outer_thin_border_to_range(worksheet, cell_range)


def _apply_workbook_markings_header_polish(workbook: Any) -> None:
    """Apply fixed visual header polish only to the final main markings layouts."""
    for worksheet in workbook.worksheets:
        if _is_final_component_marking_sheet(worksheet):
            _apply_component_marking_header_polish(worksheet)
            continue
        if _is_final_terminal_marking_sheet(worksheet):
            _apply_terminal_marking_header_polish(worksheet)
            continue
        if _is_final_wire_marking_sheet(worksheet):
            _apply_wire_marking_header_polish(worksheet)


def _write_component_strip_sheet(
    writer: Any,
    sheet_name: str,
    fuse_strip_df: pd.DataFrame,
    relay_strip_df: pd.DataFrame,
) -> None:
    """Write the Component Strip sheet with fuse and relay strips side-by-side."""
    fuse_strip_export_df = fuse_strip_df.copy()
    relay_strip_export_df = relay_strip_df.copy()
    for export_df in (fuse_strip_export_df, relay_strip_export_df):
        for column_name in export_df.columns:
            export_df[column_name] = export_df[column_name].apply(_make_excel_text_safe)

    relay_start_col = len(fuse_strip_export_df.columns) + 2
    worksheet = writer.book.create_sheet(title=sheet_name)
    writer.sheets[sheet_name] = worksheet
    fuse_strip_export_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0, startcol=0)
    relay_strip_export_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0, startcol=relay_start_col)

    for row in worksheet.iter_rows():
        for cell in row:
            if cell.__class__.__name__ == "MergedCell":
                continue
            cell.number_format = "@"
            if cell.value is None:
                cell.value = ""
            else:
                cell.value = str(cell.value)


def export_placeholder_workbook(sheets: dict[str, Any]) -> bytes:
    """Write available placeholder sheets to one Excel workbook in memory."""
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, sheet_content in sheets.items():
            if (
                isinstance(sheet_content, dict)
                and sheet_content.get("layout") == "terminal_markings"
            ):
                _write_terminal_markings_sheet(
                    writer,
                    sheet_name,
                    sheet_content.get("terminal_tmb_df", pd.DataFrame()),
                    sheet_content.get("terminal_strip_df", pd.DataFrame()),
                )
                continue

            if (
                isinstance(sheet_content, dict)
                and sheet_content.get("layout") == "component_strip"
            ):
                _write_component_strip_sheet(
                    writer,
                    sheet_name,
                    sheet_content.get("fuse_strip_df", pd.DataFrame(columns=["Space", "Text"])),
                    sheet_content.get("relay_strip_df", pd.DataFrame(columns=["Space", "Text"])),
                )
                continue

            if (
                isinstance(sheet_content, dict)
                and sheet_content.get("layout") == "wire_markings"
            ):
                _write_wire_markings_sheet(
                    writer,
                    sheet_name,
                    sheet_content.get("cables_df", pd.DataFrame()),
                    sheet_content.get("power_wires_df", pd.DataFrame()),
                )
                continue

            export_df = sheet_content.copy()
            for column_name in export_df.columns:
                export_df[column_name] = export_df[column_name].apply(_make_excel_text_safe)
            export_df.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.book[sheet_name]
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    if cell.__class__.__name__ == "MergedCell":
                        continue
                    cell.number_format = "@"
                    if cell.value is None:
                        cell.value = ""
                    else:
                        cell.value = str(cell.value)
        _apply_workbook_markings_header_polish(writer.book)
        _apply_workbook_arial_font_formatting(writer.book)
        _apply_workbook_block_header_formatting(writer.book)
        _apply_workbook_component_label_bold_formatting(writer.book)
    output.seek(0)
    return output.getvalue()
