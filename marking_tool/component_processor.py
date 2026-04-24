from __future__ import annotations

from io import BytesIO
from pathlib import Path
import re
from typing import Any

import pandas as pd


_COMPONENT_EXPECTED_COLUMNS = {
    "name": "Name",
    "type": "TYPE",
    "quantity": "Quantity",
    "total quantity": "Total quantity",
}

_COMPONENT_OPTIONAL_COLUMNS = {
    "description": "Description",
    "article no": "Article No.",
}

_FUSE_TYPES = {
    "2002-1611/1000-541",
    "2002-1611/1000-836",
}

_RELAY_4P_TYPES = {
    "RXM4GB2P7",
    "RXM4GB2BD",
    "RXZE2S114M",
}

_RELAY_RE22_TYPES = {
    "RE22R1AMR",
    "RE22R1KMR",
}

_RELAY_2P_TYPES = {
    "RXG22BD",
    "RXG22P7",
    "RGZE1S48M",
    "RE17LCBM",
    *_RELAY_RE22_TYPES,
}

_RELAY_1P_TYPES = {
    "39.00.8.230.8240",
}

_BUTTON_P_TYPES = {
    "XB4BVM3",
    "XB4BVM4",
    "XB4BVB3",
}

_COMPONENT_CM_COLUMNS = ["Mounting plate", "Component", "Door"]
_COMPONENT_CM_COLUMN_WIDTH = 14.8
_COMPONENT_STRIP_INTERNAL_GAP_COLUMNS = 2
_COMPONENT_FINAL_CM_GAP_COLUMNS = 3
_COMPONENT_CM_FUSE_GROUP_LABELS = {
    "24VDC": "Fuses 24VDC",
    "230VAC": "Fuses 230VAC",
}
_COMPONENT_CM_RELAY_GROUP_LABELS = {
    "2_pole": "Relays 2_Pole",
    "4_pole": "Relays 4_Pole",
    "timed": "Relays Timed",
    "1_pole": "Relays 1_Pole",
}
_COMPONENT_CM_BUTTONS_LABEL = "Buttons"
_COMPONENT_FINAL_SIDE_BLOCK_GAP_COLUMNS = 2
_COMPONENT_FINAL_FUSES_BLOCK_TITLE = "FUSES"
_COMPONENT_FINAL_RELAYS_BLOCK_TITLE = "RELAYS"
_COMPONENT_FINAL_BLOCK_HEADER_ROW_COUNT = 2
_COMPONENT_FINAL_STRIP_BLOCK_SUBTITLE = "Wago 210-872"
_COMPONENT_FINAL_COMPONENT_BLOCK_TITLE = "Components"
_COMPONENT_FINAL_COMPONENT_BLOCK_SUBTITLE = "Phoenix EML-16,5x5"
_COMPONENT_FINAL_FUSES_BLOCK_SUBTITLE = "Wago 2009-115"
_COMPONENT_FINAL_RELAYS_BLOCK_SUBTITLE = "Phoenix UC1-TM5"
_COMPONENT_CM_DISPLAY_RELAY_GROUP_ORDER = (
    "24VDC_2_pole",
    "230VAC_2_pole",
    "24VDC_4_pole",
    "230VAC_4_pole",
    "Timed relays",
)
_COMPONENT_CM_DISPLAY_RELAY_GROUP_BY_TYPE = {
    "RXG22BD": "24VDC_2_pole",
    "RXG22P7": "230VAC_2_pole",
    "RXM4GB2BD": "24VDC_4_pole",
    "RXM4GB2P7": "230VAC_4_pole",
    "RE22R1AMR": "Timed relays",
    "RE22R1KMR": "Timed relays",
    "RE17LCBM": "Timed relays",
}
_COMPONENT_CM_DISPLAY_RELAY_COLOR_BY_LABEL = {
    "24VDC_2_pole": "C6EFCE",
    "230VAC_2_pole": "F4CCCC",
    "24VDC_4_pole": "C6EFCE",
    "230VAC_4_pole": "F4CCCC",
    "230VAC_4A_pole": "F4CCCC",
}
_COMPONENT_FINAL_MOVED_CM_FUSE_GROUP_LABELS = tuple(_COMPONENT_CM_FUSE_GROUP_LABELS.values())
_COMPONENT_FINAL_MOVED_CM_RELAY_GROUP_LABEL_ALIASES = {
    "24VDC_2_pole": "24VDC_2_pole",
    "230VAC_2_pole": "230VAC_2_pole",
    "24VDC_4_pole": "24VDC_4_pole",
    "230VAC_4_pole": "230VAC_4A_pole",
    "230VAC_4A_pole": "230VAC_4A_pole",
}
_COMPONENT_CM_SECTION_LABELS = {
    *_COMPONENT_CM_FUSE_GROUP_LABELS.values(),
    *_COMPONENT_CM_DISPLAY_RELAY_GROUP_ORDER,
    "230VAC_4A_pole",
    _COMPONENT_CM_BUTTONS_LABEL,
    "Other",
}
_COMPONENT_CM_MARKING_MULTIPLICITY_RULES = {
    "mounting_plate": {
        "default_max_occurrences": 1,
        "name_overrides": {"Q81": 1},
        "type_overrides": {},
    },
    "component": {
        "default_max_occurrences": 1,
        "name_overrides": {"Q81": 1},
        "type_overrides": {"LA1KN40": 2},
    },
}
_COMPONENT_STRIP_SIDE_COLUMNS = ["Space", "Text"]
_COMPONENT_STRIP_GROUP_ORDER = ("24VDC", "230VAC")
_COMPONENT_CABINET_NAME_PATTERN = re.compile(
    r"^\+(?P<cabinet_id>A\d+)\b(?:[^A-Za-z0-9-]*)?(?P<normalized_name>-.*)$",
    re.IGNORECASE,
)
_COMPONENT_FILTERED_S_SUFFIX_NAME_PATTERN = re.compile(r"^-S.*\.S$")
_COMPONENT_CM_P92_NAME_PATTERN = re.compile(r"^-P92", re.IGNORECASE)
_COMPONENT_INVALID_EXCEL_SHEET_CHAR_PATTERN = re.compile(r"[\\/\?\*\[\]:]")
_FUSE_TYPE_TO_VOLTAGE_GROUP = {
    "2002-1611/1000-541": "24VDC",
    "2002-1611/1000-836": "230VAC",
}
_FUSE_A_SUFFIX_SORT_PATTERN = re.compile(
    r"^-F(?P<family>\d+)A(?P<suffix_number>\d*)(?P<suffix_text>.*)$",
    re.IGNORECASE,
)
_FUSE_NAME_SORT_PATTERN = re.compile(r"^-F(?P<number>\d+)(?P<suffix>.*)$", re.IGNORECASE)
_F92_FUSE_PATTERN = re.compile(r"^-F92", re.IGNORECASE)
_FUSE_STRIP_WIDTH = 6.2
_FUSE_STRIP_COVERED_WIDTH = 8.3
_FUSE_STRIP_SEPARATE_COVER_WIDTH = 2.1
_FUSE_STRIP_230VAC_SEPARATOR_SPACE = 13.45
_RELAY_STRIP_START_STOP_SPACE = 6.2
_RELAY_STRIP_1POLE_WIDTH = 6.2
_RELAY_STRIP_2POLE_WIDTH = 15.8
_RELAY_STRIP_4POLE_WIDTH = 27
_RELAY_STRIP_RE22_WIDTH = 22.5
_RELAY_STRIP_CLIPFIX_SPACE = 5.15
_RELAY_STRIP_START_TEXT = "START"
_RELAY_STRIP_STOP_TEXT = "STOP"
_RELAY_STRIP_GROUP_SEQUENCE = ("2_pole", "4_pole", "timed", "1_pole")
_RELAY_STRIP_GROUP_BY_TYPE = {
    "RXG22BD": "2_pole",
    "RXG22P7": "2_pole",
    "RGZE1S48M": "2_pole",
    "RE17LCBM": "2_pole",
    **{relay_type: "2_pole" for relay_type in _RELAY_RE22_TYPES},
    "RXM4GB2P7": "4_pole",
    "RXM4GB2BD": "4_pole",
    "RXZE2S114M": "4_pole",
    "39.00.8.230.8240": "1_pole",
}
_RELAY_STRIP_TYPE_PRIORITY = {
    "RE22R1AMR": 0,
    "RE22R1KMR": 0,
    "RE17LCBM": 1,
    "RGZE1S48M": 2,
    "RXZE2S114M": 3,
    "39.00.8.230.8240": 4,
    "RXG22BD": 10,
    "RXG22P7": 11,
    "RXM4GB2BD": 12,
    "RXM4GB2P7": 13,
}
_RELAY_NAME_A_SUFFIX_SORT_PATTERN = re.compile(
    r"^-K(?P<family>\d+)A(?P<suffix_number>\d*)(?P<suffix_text>.*)$",
    re.IGNORECASE,
)
_RELAY_NAME_SORT_PATTERN = re.compile(r"^-K(?P<number>\d+)(?P<suffix>.*)$", re.IGNORECASE)
_TIMED_RELAY_PATTERN = re.compile(r"^-K192(?!A)(?P<suffix_number>\d+)\b", re.IGNORECASE)
_TIMED_RELAY_A_PATTERN = re.compile(r"^-K192A(?P<suffix_number>\d+)\b", re.IGNORECASE)

_PRODUCTION_COLUMNS = ["Name", "Article No.", "TYPE", "Quantity", "Marked", "Description", "Comments"]
_RELAY_SECTION_LABEL = "Relays"
_FUSE_SECTION_LABEL = "Fuses"
_BUTTON_SECTION_LABEL = "Buttons"
_OTHER_SECTION_LABEL = "Other"
_PRODUCTION_TECHNICAL_FLAG_COLUMN = "_IncludeInCalculation"
_PRODUCTION_ONLY_COMPONENT_COLUMNS = ("Article No.",)
_GROUPED_COMPONENT_SECTIONS = (
    (_RELAY_SECTION_LABEL, {"RELAY_1P", "RELAY_4P", "RELAY_2P"}, "relay_rows"),
    (_FUSE_SECTION_LABEL, {"FUSE"}, "fuse_rows"),
    (_BUTTON_SECTION_LABEL, {"BUTTON"}, "button_rows"),
)


class _ComponentCmSheetDataFrame(pd.DataFrame):
    """Small DataFrame subtype that applies fixed CM column widths during export."""

    @property
    def _constructor(self) -> type["_ComponentCmSheetDataFrame"]:
        return _ComponentCmSheetDataFrame

    def to_excel(self, excel_writer: Any, *args: Any, **kwargs: Any) -> Any:
        """Write the CM skeleton sheet and keep all three columns at the requested width."""
        result = super().to_excel(excel_writer, *args, **kwargs)
        sheet_name = kwargs.get("sheet_name")
        if sheet_name and hasattr(excel_writer, "book"):
            from openpyxl.utils import get_column_letter

            worksheet = excel_writer.book[sheet_name]
            startrow = int(kwargs.get("startrow", 0) or 0)
            startcol = int(kwargs.get("startcol", 0) or 0)
            for column_offset, _ in enumerate(self.columns, start=1):
                worksheet.column_dimensions[get_column_letter(startcol + column_offset)].width = (
                    _COMPONENT_CM_COLUMN_WIDTH
                )
            _apply_component_cm_component_relay_color_formatting(
                worksheet,
                startrow=startrow,
                startcol=startcol,
                data_row_count=len(self),
            )
        return result


class _ComponentFinalMarkingSheetDataFrame(pd.DataFrame):
    """Self-rendering final component sheet with Strip on the left and CM on the right."""

    _metadata = [
        "layout",
        "component_strip_layout",
        "component_cm_df",
        "strip_internal_gap_columns",
        "cm_gap_columns",
    ]

    def __init__(
        self,
        *args: Any,
        component_strip_layout: dict[str, Any] | None = None,
        component_cm_df: pd.DataFrame | None = None,
        strip_internal_gap_columns: int = _COMPONENT_STRIP_INTERNAL_GAP_COLUMNS,
        cm_gap_columns: int = _COMPONENT_FINAL_CM_GAP_COLUMNS,
        **kwargs: Any,
    ) -> None:
        super().__init__(*args, **kwargs)
        self.layout = "component_markings_final"
        self.component_strip_layout = component_strip_layout or {
            "layout": "component_strip",
            "fuse_strip_df": pd.DataFrame(columns=_COMPONENT_STRIP_SIDE_COLUMNS),
            "relay_strip_df": pd.DataFrame(columns=_COMPONENT_STRIP_SIDE_COLUMNS),
        }
        self.component_cm_df = (
            component_cm_df.copy().reset_index(drop=True)
            if component_cm_df is not None
            else _build_component_cm_sheet_df()
        )
        self.strip_internal_gap_columns = strip_internal_gap_columns
        self.cm_gap_columns = cm_gap_columns

    @property
    def _constructor(self) -> type["_ComponentFinalMarkingSheetDataFrame"]:
        return _ComponentFinalMarkingSheetDataFrame

    def copy(self, deep: bool = True) -> "_ComponentFinalMarkingSheetDataFrame":
        """Keep the layout metadata attached when pandas copies the export frame."""
        copied_df = super().copy(deep=deep)
        copied_df.layout = self.layout
        copied_df.component_strip_layout = {
            "layout": self.component_strip_layout.get("layout", "component_strip"),
            "fuse_strip_df": self.component_strip_layout.get(
                "fuse_strip_df",
                pd.DataFrame(columns=_COMPONENT_STRIP_SIDE_COLUMNS),
            ).copy(deep=deep),
            "relay_strip_df": self.component_strip_layout.get(
                "relay_strip_df",
                pd.DataFrame(columns=_COMPONENT_STRIP_SIDE_COLUMNS),
            ).copy(deep=deep),
        }
        copied_df.component_cm_df = self.component_cm_df.copy(deep=deep)
        copied_df.strip_internal_gap_columns = self.strip_internal_gap_columns
        copied_df.cm_gap_columns = self.cm_gap_columns
        return copied_df

    def to_excel(self, excel_writer: Any, *args: Any, **kwargs: Any) -> Any:
        """Write one final component sheet with Strip blocks and CM side-by-side."""
        sheet_name = kwargs.get("sheet_name")
        if not sheet_name:
            raise ValueError("component final marking export requires a sheet_name")

        from openpyxl.utils import get_column_letter

        startrow = int(kwargs.get("startrow", 0) or 0)
        startcol = int(kwargs.get("startcol", 0) or 0)

        fuse_strip_df = self.component_strip_layout.get(
            "fuse_strip_df",
            pd.DataFrame(columns=_COMPONENT_STRIP_SIDE_COLUMNS),
        ).copy().reset_index(drop=True)
        relay_strip_df = self.component_strip_layout.get(
            "relay_strip_df",
            pd.DataFrame(columns=_COMPONENT_STRIP_SIDE_COLUMNS),
        ).copy().reset_index(drop=True)
        component_cm_export_df, fuse_block_export_df, relay_block_export_df = _split_final_component_cm_layout(
            self.component_cm_df
        )

        for export_df in (fuse_strip_df, relay_strip_df, component_cm_export_df, fuse_block_export_df, relay_block_export_df):
            for column_name in export_df.columns:
                export_df[column_name] = export_df[column_name].map(_stringify_cell)

        relay_start_col = startcol + len(fuse_strip_df.columns) + self.strip_internal_gap_columns
        cm_start_col = relay_start_col + len(relay_strip_df.columns) + self.cm_gap_columns
        side_block_start_col = cm_start_col + len(component_cm_export_df.columns) + _COMPONENT_FINAL_SIDE_BLOCK_GAP_COLUMNS
        block_data_startrow = startrow + _COMPONENT_FINAL_BLOCK_HEADER_ROW_COUNT

        worksheet = excel_writer.book.create_sheet(title=sheet_name)
        excel_writer.sheets[sheet_name] = worksheet
        _write_component_final_block_header(
            worksheet,
            startrow=startrow,
            startcol=startcol,
            block_width=max(1, len(fuse_strip_df.columns)),
            title="FUSE STRIP",
            subtitle=_COMPONENT_FINAL_STRIP_BLOCK_SUBTITLE,
        )
        fuse_strip_df.to_excel(
            excel_writer,
            sheet_name=sheet_name,
            index=False,
            startrow=block_data_startrow,
            startcol=startcol,
        )
        _write_component_final_block_header(
            worksheet,
            startrow=startrow,
            startcol=relay_start_col,
            block_width=max(1, len(relay_strip_df.columns)),
            title="RELAYS STRIP",
            subtitle=_COMPONENT_FINAL_STRIP_BLOCK_SUBTITLE,
        )
        relay_strip_df.to_excel(
            excel_writer,
            sheet_name=sheet_name,
            index=False,
            startrow=block_data_startrow,
            startcol=relay_start_col,
        )
        _write_component_final_block_header(
            worksheet,
            startrow=startrow,
            startcol=cm_start_col,
            block_width=max(1, len(component_cm_export_df.columns)),
            title=_COMPONENT_FINAL_COMPONENT_BLOCK_TITLE,
            subtitle=_COMPONENT_FINAL_COMPONENT_BLOCK_SUBTITLE,
        )
        component_cm_export_df.to_excel(
            excel_writer,
            sheet_name=sheet_name,
            index=False,
            startrow=block_data_startrow,
            startcol=cm_start_col,
        )
        current_side_block_start_col = side_block_start_col
        _write_component_final_block_header(
            worksheet,
            startrow=startrow,
            startcol=current_side_block_start_col,
            block_width=1,
            title=_COMPONENT_FINAL_FUSES_BLOCK_TITLE,
            subtitle=_COMPONENT_FINAL_FUSES_BLOCK_SUBTITLE,
        )
        if not fuse_block_export_df.empty:
            fuse_block_export_df.to_excel(
                excel_writer,
                sheet_name=sheet_name,
                index=False,
                startrow=block_data_startrow,
                startcol=current_side_block_start_col,
                header=False,
            )
            worksheet.column_dimensions[
                get_column_letter(current_side_block_start_col + 1)
            ].width = _COMPONENT_CM_COLUMN_WIDTH
        else:
            worksheet.column_dimensions[
                get_column_letter(current_side_block_start_col + 1)
            ].width = _COMPONENT_CM_COLUMN_WIDTH
        current_side_block_start_col += 1 + _COMPONENT_FINAL_SIDE_BLOCK_GAP_COLUMNS
        _write_component_final_block_header(
            worksheet,
            startrow=startrow,
            startcol=current_side_block_start_col,
            block_width=1,
            title=_COMPONENT_FINAL_RELAYS_BLOCK_TITLE,
            subtitle=_COMPONENT_FINAL_RELAYS_BLOCK_SUBTITLE,
        )
        if not relay_block_export_df.empty:
            relay_block_export_df.to_excel(
                excel_writer,
                sheet_name=sheet_name,
                index=False,
                startrow=block_data_startrow,
                startcol=current_side_block_start_col,
                header=False,
            )
            relay_column_index = current_side_block_start_col + 1
            worksheet.column_dimensions[get_column_letter(relay_column_index)].width = _COMPONENT_CM_COLUMN_WIDTH
            _apply_component_cm_relay_color_formatting_for_column(
                worksheet,
                column_index=relay_column_index,
                data_start_row=block_data_startrow + 1,
                data_row_count=len(relay_block_export_df),
            )
        else:
            relay_column_index = current_side_block_start_col + 1
            worksheet.column_dimensions[get_column_letter(relay_column_index)].width = _COMPONENT_CM_COLUMN_WIDTH

        for row in worksheet.iter_rows():
            for cell in row:
                if cell.__class__.__name__ == "MergedCell":
                    continue
                cell.number_format = "@"
                if cell.value is None:
                    cell.value = ""
                else:
                    cell.value = str(cell.value)
        return None


def _normalize_column_name(value: Any) -> str:
    """Return a simple normalized column label for matching."""
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    normalized = " ".join(text.replace("\n", " ").split()).lower()
    return normalized.replace(".", "")


def _stringify_cell(value: Any) -> str:
    """Convert a cell value to a simple trimmed string."""
    if pd.isna(value):
        return ""
    return str(value).strip()


def _write_component_final_block_header(
    worksheet: Any,
    *,
    startrow: int,
    startcol: int,
    block_width: int,
    title: str,
    subtitle: str,
) -> None:
    """Write one two-row block header above one final Component Marking block."""
    start_row_index = startrow + 1
    start_column_index = startcol + 1
    end_column_index = start_column_index + max(0, block_width - 1)
    if block_width > 1:
        worksheet.merge_cells(
            start_row=start_row_index,
            start_column=start_column_index,
            end_row=start_row_index,
            end_column=end_column_index,
        )
        worksheet.merge_cells(
            start_row=start_row_index + 1,
            start_column=start_column_index,
            end_row=start_row_index + 1,
            end_column=end_column_index,
        )
    worksheet.cell(row=start_row_index, column=start_column_index).value = title
    worksheet.cell(row=start_row_index + 1, column=start_column_index).value = subtitle


def _apply_component_cm_component_relay_color_formatting(
    worksheet: Any,
    *,
    startrow: int,
    startcol: int,
    data_row_count: int,
) -> None:
    """Color CM Component relay marking cells by the active CM-only relay section."""
    if data_row_count <= 0:
        return

    from openpyxl.styles import PatternFill

    component_column_index = startcol + _COMPONENT_CM_COLUMNS.index("Component") + 1
    data_start_row = startrow + 2
    data_end_row = data_start_row + data_row_count - 1
    active_fill_color = ""
    for row_index in range(data_start_row, data_end_row + 1):
        cell = worksheet.cell(row=row_index, column=component_column_index)
        text_value = _stringify_cell(cell.value)
        if text_value == "":
            active_fill_color = ""
            continue
        if text_value in _COMPONENT_CM_DISPLAY_RELAY_COLOR_BY_LABEL:
            active_fill_color = _COMPONENT_CM_DISPLAY_RELAY_COLOR_BY_LABEL[text_value]
            continue
        if text_value in _COMPONENT_CM_SECTION_LABELS:
            active_fill_color = ""
            continue
        if active_fill_color == "":
            continue
        cell.fill = PatternFill(
            fill_type="solid",
            start_color=active_fill_color,
            end_color=active_fill_color,
        )


def _apply_component_cm_relay_color_formatting_for_column(
    worksheet: Any,
    *,
    column_index: int,
    data_start_row: int,
    data_row_count: int,
) -> None:
    """Color relay marking cells in one final-sheet column by the active relay section label."""
    if data_row_count <= 0:
        return

    from openpyxl.styles import PatternFill

    data_end_row = data_start_row + data_row_count - 1
    active_fill_color = ""
    for row_index in range(data_start_row, data_end_row + 1):
        cell = worksheet.cell(row=row_index, column=column_index)
        text_value = _stringify_cell(cell.value)
        if text_value == "":
            active_fill_color = ""
            continue
        if text_value in _COMPONENT_CM_DISPLAY_RELAY_COLOR_BY_LABEL:
            active_fill_color = _COMPONENT_CM_DISPLAY_RELAY_COLOR_BY_LABEL[text_value]
            continue
        if text_value in _COMPONENT_CM_SECTION_LABELS:
            active_fill_color = ""
            continue
        if active_fill_color == "":
            continue
        cell.fill = PatternFill(
            fill_type="solid",
            start_color=active_fill_color,
            end_color=active_fill_color,
        )


def _is_filtered_component_name(value: Any) -> bool:
    """Remove global -S*.S component rows before they enter downstream processing."""
    text = _stringify_cell(value)
    evaluation_text = text
    if text.startswith("+A"):
        cabinet_split_index = text.find("-", 2)
        if cabinet_split_index != -1:
            evaluation_text = text[cabinet_split_index:]

    return bool(_COMPONENT_FILTERED_S_SUFFIX_NAME_PATTERN.fullmatch(evaluation_text.upper()))


def _extract_component_cabinet_number(cabinet_id_value: Any) -> int | None:
    """Extract the numeric part from one cabinet id such as A2."""
    cabinet_id = _stringify_cell(cabinet_id_value).upper()
    if not cabinet_id.startswith("A"):
        return None
    cabinet_number_text = cabinet_id[1:]
    if not cabinet_number_text.isdigit():
        return None
    return int(cabinet_number_text)


def _is_cabinet_door_m_component(name_value: Any, *, cabinet_id: Any | None = None) -> bool:
    """Return whether one name represents an A2+ cabinet-prefixed local -M* Door component."""
    local_name = _normalize_component_local_name(name_value)
    if not _normalize_component_name(local_name).startswith("-M"):
        return False

    cabinet_number = _extract_component_cabinet_number(cabinet_id) if cabinet_id is not None else None
    if cabinet_number is None:
        cabinet_parts = _extract_component_cabinet_parts(name_value)
        if not cabinet_parts:
            return False
        cabinet_number = _extract_component_cabinet_number(cabinet_parts[0])
    return cabinet_number is not None and cabinet_number >= 2


def _load_component_input(file_bytes: bytes) -> tuple[pd.DataFrame, list[str], list[str]]:
    """Read the first sheet, drop fully empty rows, and retain expected columns if present."""
    developer_debug_messages: list[str] = []
    raw_df = pd.read_excel(BytesIO(file_bytes), sheet_name=0, dtype=object)
    raw_df = raw_df.dropna(axis=0, how="all").reset_index(drop=True)
    developer_debug_messages.append(f"component parser: loaded {len(raw_df)} non-empty rows from first sheet")

    normalized_columns: dict[str, list[Any]] = {}
    for column_name in raw_df.columns:
        normalized_name = _normalize_column_name(column_name)
        if not normalized_name:
            continue
        normalized_columns.setdefault(normalized_name, []).append(column_name)

    found_columns = [
        canonical_name
        for normalized_name, canonical_name in _COMPONENT_EXPECTED_COLUMNS.items()
        if normalized_name in normalized_columns
    ]
    missing_columns = [
        canonical_name
        for normalized_name, canonical_name in _COMPONENT_EXPECTED_COLUMNS.items()
        if normalized_name not in normalized_columns
    ]
    developer_debug_messages.append(
        "component parser: found expected columns -> "
        + (", ".join(found_columns) if found_columns else "none")
    )
    developer_debug_messages.append(
        "component parser: missing expected columns -> "
        + (", ".join(missing_columns) if missing_columns else "none")
    )

    found_optional_columns = [
        canonical_name
        for normalized_name, canonical_name in _COMPONENT_OPTIONAL_COLUMNS.items()
        if normalized_name in normalized_columns
    ]
    developer_debug_messages.append(
        "component parser: found optional columns -> "
        + (", ".join(found_optional_columns) if found_optional_columns else "none")
    )

    for normalized_name, canonical_name in {**_COMPONENT_EXPECTED_COLUMNS, **_COMPONENT_OPTIONAL_COLUMNS}.items():
        matching_columns = normalized_columns.get(normalized_name, [])
        if len(matching_columns) > 1:
            ignored_columns = ", ".join(f"`{column_name}`" for column_name in matching_columns[1:])
            developer_debug_messages.append(
                f"component parser: multiple columns matched {canonical_name} -> "
                f"using `{matching_columns[0]}`, ignoring {ignored_columns}"
            )

    selected_columns = [
        normalized_columns[normalized_name][0]
        for normalized_name in _COMPONENT_EXPECTED_COLUMNS
        if normalized_name in normalized_columns
    ]
    selected_columns.extend(
        normalized_columns[normalized_name][0]
        for normalized_name in _COMPONENT_OPTIONAL_COLUMNS
        if normalized_name in normalized_columns
    )
    column_aliases = {**_COMPONENT_EXPECTED_COLUMNS, **_COMPONENT_OPTIONAL_COLUMNS}
    component_df = raw_df.loc[:, selected_columns].copy()
    component_df = component_df.rename(
        columns={
            normalized_columns[normalized_name][0]: canonical_name
            for normalized_name, canonical_name in column_aliases.items()
            if normalized_name in normalized_columns
        }
    )

    for column_name in component_df.columns:
        if component_df[column_name].dtype == object:
            component_df[column_name] = component_df[column_name].map(
                lambda value: _stringify_cell(value) if pd.notna(value) else value
            )

    filtered_s_suffix_rows = 0
    if "Name" in component_df.columns:
        filtered_s_suffix_mask = component_df["Name"].map(_is_filtered_component_name)
        filtered_s_suffix_rows = int(filtered_s_suffix_mask.sum())
        component_df = component_df.loc[~filtered_s_suffix_mask].reset_index(drop=True)
    developer_debug_messages.append(
        f"component filter: removed {filtered_s_suffix_rows} rows matching -S*.S pattern"
    )

    return component_df, found_columns, developer_debug_messages


def _is_unused_component_name(name: Any) -> bool:
    """Apply the requested conservative Unused split rules."""
    text = _stringify_cell(name)
    if text == "":
        return True

    evaluation_text = text
    if text.startswith("+A"):
        cabinet_split_index = text.find("-", 2)
        if cabinet_split_index != -1:
            evaluation_text = text[cabinet_split_index:]

    if text.startswith("+") and not text.startswith("+A"):
        return True
    if evaluation_text.startswith("-B"):
        return True
    if evaluation_text.startswith("-W"):
        return True
    if _is_cabinet_door_m_component(text):
        return False
    if evaluation_text.startswith("-M") and not evaluation_text.startswith("-M92"):
        return True
    if evaluation_text.startswith("-X") and not evaluation_text.startswith("-X921"):
        return True
    return False


def _normalize_component_type(value: Any) -> str:
    """Normalize TYPE values conservatively for classification."""
    return _stringify_cell(value).upper()


def _normalize_component_name(value: Any) -> str:
    """Normalize Name values conservatively for category checks."""
    return _stringify_cell(value).upper()


def _sanitize_component_cabinet_sheet_name(cabinet_id: Any) -> str:
    """Keep cabinet ids safe for Excel worksheet names."""
    sanitized_name = _stringify_cell(cabinet_id).upper()
    if not sanitized_name:
        return ""
    sanitized_name = _COMPONENT_INVALID_EXCEL_SHEET_CHAR_PATTERN.sub("", sanitized_name)
    sanitized_name = sanitized_name.strip().strip("'")
    return sanitized_name[:31]


def _build_component_markings_workbook_sheet_name(cabinet_id: Any, suffix: str) -> str:
    """Build a valid main Markings workbook sheet name for one cabinet component sheet."""
    safe_cabinet_id = _sanitize_component_cabinet_sheet_name(cabinet_id) or "CABINET"
    safe_suffix = _stringify_cell(suffix)
    max_cabinet_length = max(1, 31 - len(f" {safe_suffix}"))
    if len(safe_cabinet_id) > max_cabinet_length:
        safe_cabinet_id = safe_cabinet_id[:max_cabinet_length]
    return f"{safe_cabinet_id} {safe_suffix}"


def _build_component_debug_workbook(
    debug_sheets: dict[str, Any],
) -> bytes | None:
    """Reuse the shared workbook exporter for component debug support sheets."""
    if not debug_sheets:
        return None
    from .terminal_processor import export_placeholder_workbook

    return export_placeholder_workbook(debug_sheets)


def _build_component_debug_messages_sheet(messages: list[str]) -> pd.DataFrame:
    """Build a readable component debug sheet with one message per row."""
    return pd.DataFrame(
        [{"Index": index + 1, "Message": message} for index, message in enumerate(messages)],
        columns=["Index", "Message"],
    )


def _determine_component_routing_mode(
    sorted_cabinet_ids: list[str],
    use_single_cabinet_local_dataset: bool,
) -> str:
    """Choose the component workbook routing mode from detected cabinet context."""
    cabinet_count = len(sorted_cabinet_ids)
    if cabinet_count > 1:
        return "multi_cabinet"
    if cabinet_count == 1 or use_single_cabinet_local_dataset:
        return "single_cabinet"
    return "no_cabinet"


def _build_component_local_name_source_df(component_df: pd.DataFrame) -> pd.DataFrame:
    """Create one component dataset with cabinet prefixes removed from Name values only."""
    localized_df = component_df.copy().reset_index(drop=True)
    if localized_df.empty or "Name" not in localized_df.columns:
        return localized_df
    localized_df["Name"] = localized_df["Name"].map(_normalize_component_local_name)
    return localized_df


def _build_component_general_information_sheet(
    *,
    routing_mode: str,
    detected_cabinet_ids: list[str],
    total_input_rows: int,
    component_marking_rows: int,
    unused_rows: int,
    cabinet_source_row_counts: dict[str, int],
    final_main_sheet_names: list[str],
    final_debug_sheet_names: list[str],
) -> pd.DataFrame:
    """Build one flat routing/debug summary sheet for the component debug workbook."""
    info_rows: list[dict[str, Any]] = [
        {"Field": "Routing mode", "Value": routing_mode},
        {
            "Field": "Detected cabinet ids",
            "Value": ", ".join(detected_cabinet_ids) if detected_cabinet_ids else "none",
        },
        {"Field": "Input rows", "Value": total_input_rows},
        {"Field": "Component Marking source rows", "Value": component_marking_rows},
        {"Field": "Unused source rows", "Value": unused_rows},
        {
            "Field": "Final main sheets",
            "Value": ", ".join(final_main_sheet_names) if final_main_sheet_names else "none",
        },
        {
            "Field": "Final debug sheets",
            "Value": ", ".join(final_debug_sheet_names) if final_debug_sheet_names else "none",
        },
    ]

    if cabinet_source_row_counts:
        for cabinet_id, row_count in cabinet_source_row_counts.items():
            info_rows.append(
                {
                    "Field": f"Cabinet {cabinet_id} source rows",
                    "Value": row_count,
                }
            )
    else:
        info_rows.append({"Field": "Cabinet source rows", "Value": "none"})

    return pd.DataFrame(info_rows, columns=["Field", "Value"])


def _validate_component_routing_sheet_sets(
    routing_mode: str,
    final_main_sheet_names: list[str],
    final_debug_sheet_names: list[str],
    expected_final_component_marking_sheet_names: list[str] | None = None,
) -> None:
    """Validate that main/debug component sheet routing matches the derived mode."""
    main_sheet_set = set(final_main_sheet_names)
    debug_sheet_set = set(final_debug_sheet_names)
    expected_final_sheet_set = set(expected_final_component_marking_sheet_names or [])

    if routing_mode == "no_cabinet":
        expected_main_sheet_set = {"Component Marking", "Component Strip", "Unused", "CM"}
        missing_main_sheets = expected_main_sheet_set - main_sheet_set
        unexpected_main_sheets = main_sheet_set - expected_main_sheet_set
        if missing_main_sheets or unexpected_main_sheets:
            details: list[str] = []
            if missing_main_sheets:
                details.append("missing " + ", ".join(sorted(missing_main_sheets)))
            if unexpected_main_sheets:
                details.append("unexpected " + ", ".join(sorted(unexpected_main_sheets)))
            raise ValueError(
                "component workbook routing: no_cabinet main workbook mismatch -> "
                + "; ".join(details)
            )

        unexpected_debug_sheets = debug_sheet_set - {"Developer Debug"}
        if unexpected_debug_sheets:
            raise ValueError(
                "component workbook routing: no_cabinet debug workbook unexpectedly includes "
                + ", ".join(sorted(unexpected_debug_sheets))
            )
        return

    if routing_mode == "single_cabinet":
        expected_main_sheet_set = expected_final_sheet_set or {"Component Marking"}
        missing_main_sheets = expected_main_sheet_set - main_sheet_set
        unexpected_main_sheets = main_sheet_set - expected_main_sheet_set
        forbidden_intermediate_sheet_names = sorted(
            sheet_name
            for sheet_name in final_main_sheet_names
            if _stringify_cell(sheet_name) in {"Component Strip", "CM"}
            or _stringify_cell(sheet_name).endswith(" Component Strip")
            or _stringify_cell(sheet_name).endswith(" CM")
        )
        if missing_main_sheets or unexpected_main_sheets:
            details = []
            if missing_main_sheets:
                details.append("missing " + ", ".join(sorted(missing_main_sheets)))
            if unexpected_main_sheets:
                details.append("unexpected " + ", ".join(sorted(unexpected_main_sheets)))
            raise ValueError(
                "component workbook routing: single_cabinet main workbook mismatch -> "
                + "; ".join(details)
            )
        if forbidden_intermediate_sheet_names:
            raise ValueError(
                "component workbook routing: single_cabinet main workbook unexpectedly includes intermediate sheets "
                + ", ".join(forbidden_intermediate_sheet_names)
            )

        expected_debug_sheet_set = {
            "General information",
            "Component Marking",
            "Unused",
            "Developer Debug",
        }
        missing_debug_sheets = expected_debug_sheet_set - debug_sheet_set
        unexpected_debug_sheets = debug_sheet_set - expected_debug_sheet_set
        if missing_debug_sheets or unexpected_debug_sheets:
            details = []
            if missing_debug_sheets:
                details.append("missing " + ", ".join(sorted(missing_debug_sheets)))
            if unexpected_debug_sheets:
                details.append("unexpected " + ", ".join(sorted(unexpected_debug_sheets)))
            raise ValueError(
                "component workbook routing: single_cabinet debug workbook mismatch -> "
                + "; ".join(details)
            )
        return

    if routing_mode == "multi_cabinet":
        missing_main_sheets = expected_final_sheet_set - main_sheet_set
        unexpected_main_sheets = main_sheet_set - expected_final_sheet_set
        forbidden_intermediate_sheet_names = sorted(
            sheet_name
            for sheet_name in final_main_sheet_names
            if _stringify_cell(sheet_name) in {"Component Strip", "CM", "Unused"}
            or _stringify_cell(sheet_name).endswith(" Component Strip")
            or _stringify_cell(sheet_name).endswith(" CM")
        )
        if missing_main_sheets or unexpected_main_sheets:
            details = []
            if missing_main_sheets:
                details.append("missing " + ", ".join(sorted(missing_main_sheets)))
            if unexpected_main_sheets:
                details.append("unexpected " + ", ".join(sorted(unexpected_main_sheets)))
            raise ValueError(
                "component workbook routing: multi_cabinet main workbook mismatch -> "
                + "; ".join(details)
            )
        if forbidden_intermediate_sheet_names:
            raise ValueError(
                "component workbook routing: multi_cabinet main workbook unexpectedly includes intermediate sheets "
                + ", ".join(forbidden_intermediate_sheet_names)
            )
        if not final_main_sheet_names:
            raise ValueError("component workbook routing: multi_cabinet main workbook is empty")

        required_debug_sheet_set = {
            "General information",
            "Component Marking",
            "Component Marking Raw",
            "Unused",
            "Developer Debug",
        }
        missing_debug_sheets = required_debug_sheet_set - debug_sheet_set
        if missing_debug_sheets:
            raise ValueError(
                "component workbook routing: multi_cabinet debug workbook missing "
                + ", ".join(sorted(missing_debug_sheets))
            )

        unexpected_debug_sheets = {"Component Strip"} & debug_sheet_set
        if unexpected_debug_sheets:
            raise ValueError(
                "component workbook routing: multi_cabinet debug workbook unexpectedly includes "
                + ", ".join(sorted(unexpected_debug_sheets))
            )
        return

    raise ValueError(f"component workbook routing: unsupported routing mode {routing_mode!r}")


def _normalize_component_local_name(name_value: Any) -> str:
    """Return one cabinet-local component name, removing any +A* prefix when present."""
    text = _stringify_cell(name_value)
    cabinet_parts = _extract_component_cabinet_parts(text)
    if cabinet_parts:
        return cabinet_parts[2]
    return text


def _build_component_cm_source_df(component_df: pd.DataFrame) -> pd.DataFrame:
    """Normalize one CM source dataset to local names while preserving row-level TYPE pairing."""
    if component_df.empty or "Name" not in component_df.columns:
        return pd.DataFrame(columns=["Name", "TYPE", "Category", "_original_order"])

    cm_source_df = component_df.copy().reset_index(drop=True)
    cm_source_df["Name"] = cm_source_df["Name"].map(_normalize_component_local_name)
    if "TYPE" not in cm_source_df.columns:
        cm_source_df["TYPE"] = ""
    cm_source_df["_original_order"] = range(len(cm_source_df))
    cm_source_df["Category"] = cm_source_df.apply(
        lambda row: _classify_component_category(row.get("Name"), row.get("TYPE")),
        axis=1,
    )
    return cm_source_df


def _component_cm_marking_rule_key(name_value: Any) -> str:
    """Return the centralized multiplicity-rule key for one CM marking Name."""
    marking_name = _stringify_cell(name_value).upper()
    if marking_name in {"Q81", "-Q81"}:
        return "Q81"
    return marking_name


def _is_component_cm_p92_name(name_value: Any) -> bool:
    """Return whether one local CM Name belongs to the P92* exception family."""
    return bool(_COMPONENT_CM_P92_NAME_PATTERN.match(_stringify_cell(name_value)))


def _is_component_cm_door_candidate_name(name_value: Any, *, cabinet_id: Any | None = None) -> bool:
    """Return whether one raw/local component Name should contribute its local name to the Door column."""
    raw_name = _stringify_cell(name_value)
    if raw_name == "":
        return False

    local_name = _normalize_component_local_name(raw_name)
    normalized_local_name = _normalize_component_name(local_name)
    if normalized_local_name == "" or _is_component_cm_p92_name(local_name):
        return False
    if normalized_local_name.startswith("-P") or normalized_local_name.startswith("-S"):
        return True
    return _is_cabinet_door_m_component(raw_name, cabinet_id=cabinet_id)


def _classify_component_cm_relay_group(type_value: Any) -> str:
    """Classify one relay TYPE into the CM-only Component-column relay display groups."""
    return _COMPONENT_CM_DISPLAY_RELAY_GROUP_BY_TYPE.get(_normalize_component_type(type_value), "")


def _infer_component_production_relay_group(name_group_df: pd.DataFrame) -> str:
    """Infer one production relay display group from the coil/timed TYPE rows within one Name group."""
    if name_group_df.empty:
        return ""

    direct_group_df = name_group_df.copy()
    direct_group_df["_production_relay_group"] = direct_group_df["TYPE"].map(_classify_component_cm_relay_group)
    direct_group_df = direct_group_df.loc[direct_group_df["_production_relay_group"].ne("")].copy()
    if direct_group_df.empty:
        return ""
    representative_row = direct_group_df.sort_values("_original_order", kind="mergesort").iloc[0]
    return _stringify_cell(representative_row["_production_relay_group"])


def _get_component_cm_marking_max_occurrences(column_key: str, name_value: Any, type_value: Any) -> int:
    """Resolve the allowed CM multiplicity for one Name/TYPE pair."""
    rule_set = _COMPONENT_CM_MARKING_MULTIPLICITY_RULES[column_key]
    normalized_type = _normalize_component_type(type_value)
    rule_key = _component_cm_marking_rule_key(name_value)
    if rule_key in rule_set["name_overrides"]:
        return int(rule_set["name_overrides"][rule_key])
    if normalized_type in rule_set["type_overrides"]:
        return int(rule_set["type_overrides"][normalized_type])
    return int(rule_set["default_max_occurrences"])


def _build_component_cm_door_source_entries(component_df: pd.DataFrame) -> list[str]:
    """Build the unique Door-name source list before the final x2 Door duplication."""
    if component_df.empty or "Name" not in component_df.columns:
        return []

    cabinet_id = component_df.attrs.get("cabinet_id")
    door_names = [
        _normalize_component_local_name(raw_name)
        for raw_name in component_df["Name"].tolist()
        if _is_component_cm_door_candidate_name(raw_name, cabinet_id=cabinet_id)
    ]
    return list(dict.fromkeys(door_names))


def _build_component_cm_door_entries(component_df: pd.DataFrame) -> list[str]:
    """Build the Door-column list as one deduplicated non-P92 block repeated twice."""
    deduplicated_door_names = _build_component_cm_door_source_entries(component_df)
    return deduplicated_door_names + deduplicated_door_names


def _deduplicate_component_cm_marking_rows(
    marking_df: pd.DataFrame,
    *,
    column_key: str,
    seen_counts: dict[str, int] | None = None,
) -> tuple[list[str], dict[str, int]]:
    """Deduplicate CM marking rows with centralized rules while preserving order."""
    if marking_df.empty:
        return [], dict(seen_counts or {})

    working_df = marking_df.copy()
    if "Name" not in working_df.columns:
        working_df["Name"] = ""
    if "TYPE" not in working_df.columns:
        working_df["TYPE"] = ""

    occurrence_counts = dict(seen_counts or {})
    grouped_rule_entries: dict[str, dict[str, Any]] = {}
    ordered_rule_keys: list[str] = []
    for row in working_df.loc[:, ["Name", "TYPE"]].itertuples(index=False):
        display_name = _stringify_cell(row.Name)
        if display_name == "":
            continue
        rule_key = _component_cm_marking_rule_key(display_name)
        max_occurrences = _get_component_cm_marking_max_occurrences(column_key, display_name, row.TYPE)
        if rule_key not in grouped_rule_entries:
            grouped_rule_entries[rule_key] = {
                "display_name": display_name,
                "max_occurrences": max_occurrences,
            }
            ordered_rule_keys.append(rule_key)
            continue
        grouped_rule_entries[rule_key]["max_occurrences"] = max(
            int(grouped_rule_entries[rule_key]["max_occurrences"]),
            max_occurrences,
        )

    deduplicated_names: list[str] = []
    for rule_key in ordered_rule_keys:
        display_name = _stringify_cell(grouped_rule_entries[rule_key]["display_name"])
        max_occurrences = int(grouped_rule_entries[rule_key]["max_occurrences"])
        remaining_occurrences = max(0, max_occurrences - occurrence_counts.get(rule_key, 0))
        if remaining_occurrences == 0:
            continue
        deduplicated_names.extend([display_name] * remaining_occurrences)
        occurrence_counts[rule_key] = occurrence_counts.get(rule_key, 0) + remaining_occurrences
    return deduplicated_names, occurrence_counts


def _deduplicate_marking_names_default(marking_names: list[Any], *, column_key: str) -> list[str]:
    """Apply the centralized default x1 CM rule to one flat marking Name list."""
    if not marking_names:
        return []
    marking_df = pd.DataFrame({"Name": marking_names, "TYPE": [""] * len(marking_names)})
    deduplicated_names, _ = _deduplicate_component_cm_marking_rows(marking_df, column_key=column_key)
    return deduplicated_names


def _deduplicate_component_cm_entries(
    ordered_groups: list[tuple[str, pd.DataFrame]],
    *,
    blocked_rule_keys: set[str] | None = None,
) -> list[str]:
    """Build grouped Component CM entries while keeping group labels outside name deduplication."""
    if not ordered_groups:
        return []

    component_entries: list[str] = []
    seen_counts: dict[str, int] = {}
    blocked_rule_keys = blocked_rule_keys or set()
    for group_label, group_df in ordered_groups:
        filtered_group_df = group_df.loc[
            group_df["Name"].map(
                lambda name_value: _component_cm_marking_rule_key(name_value) not in blocked_rule_keys
                or _is_component_cm_p92_name(name_value)
            )
        ].copy()
        group_names, seen_counts = _deduplicate_component_cm_marking_rows(
            filtered_group_df,
            column_key="component",
            seen_counts=seen_counts,
        )
        if not group_names:
            continue
        if component_entries:
            component_entries.append("")
        component_entries.append(group_label)
        component_entries.extend(group_names)
    return component_entries


def _build_component_cm_fuse_groups(cm_source_df: pd.DataFrame) -> list[tuple[str, pd.DataFrame]]:
    """Build CM fuse groups in the same voltage-group and natural-name order as the strip layout."""
    fuse_df = cm_source_df.loc[cm_source_df["Category"].eq("FUSE")].copy()
    if fuse_df.empty:
        return []

    fuse_groups: list[tuple[str, pd.DataFrame]] = []
    for voltage_group in _COMPONENT_STRIP_GROUP_ORDER:
        voltage_df = fuse_df.loc[
            fuse_df["TYPE"].map(_detect_fuse_voltage_group).eq(voltage_group)
        ].copy()
        if voltage_df.empty:
            continue

        fuse_sort_keys = voltage_df["Name"].map(_component_fuse_name_sort_key).tolist()
        voltage_df[[
            "_fuse_sort_group",
            "_fuse_sort_family",
            "_fuse_sort_variant_kind",
            "_fuse_sort_variant_number",
            "_fuse_sort_suffix",
            "_fuse_sort_text",
        ]] = pd.DataFrame(fuse_sort_keys, index=voltage_df.index)
        voltage_df = voltage_df.sort_values(
            by=[
                "_fuse_sort_group",
                "_fuse_sort_family",
                "_fuse_sort_variant_kind",
                "_fuse_sort_variant_number",
                "_fuse_sort_suffix",
                "_fuse_sort_text",
                "_original_order",
            ],
            kind="mergesort",
        ).drop(
            columns=[
                "_fuse_sort_group",
                "_fuse_sort_family",
                "_fuse_sort_variant_kind",
                "_fuse_sort_variant_number",
                "_fuse_sort_suffix",
                "_fuse_sort_text",
            ]
        ).reset_index(drop=True)
        fuse_groups.append(
            (
                _COMPONENT_CM_FUSE_GROUP_LABELS[voltage_group],
                voltage_df.reset_index(drop=True),
            )
        )
    return fuse_groups


def _build_component_cm_relay_groups(cm_source_df: pd.DataFrame) -> list[tuple[str, pd.DataFrame]]:
    """Build CM-only relay display groups without changing the strip relay logic."""
    relay_df = cm_source_df.copy()
    relay_df["_cm_relay_group"] = relay_df["TYPE"].map(_classify_component_cm_relay_group)
    relay_df = relay_df.loc[relay_df["_cm_relay_group"].ne("")].copy()
    if relay_df.empty:
        return []

    relay_groups: list[tuple[str, pd.DataFrame]] = []
    for group_label in _COMPONENT_CM_DISPLAY_RELAY_GROUP_ORDER:
        group_df = relay_df.loc[relay_df["_cm_relay_group"].eq(group_label)].copy()
        if group_df.empty:
            continue
        if group_label == "Timed relays":
            group_df = _sort_component_timed_relay_group_df(group_df)
        else:
            group_df = _sort_component_relay_group_df(group_df)
        relay_groups.append((group_label, group_df.reset_index(drop=True)))
    return relay_groups


def _build_component_cm_button_other_groups(cm_source_df: pd.DataFrame) -> list[tuple[str, pd.DataFrame]]:
    """Build CM other groups with stable local-name ordering while leaving Buttons out of Component."""
    other_df = _sort_grouped_component_rows(
        cm_source_df.loc[cm_source_df["Category"].eq("OTHER")].copy()
    )

    button_other_groups: list[tuple[str, pd.DataFrame]] = []
    if not other_df.empty:
        button_other_groups.append((_OTHER_SECTION_LABEL, other_df.reset_index(drop=True)))
    return button_other_groups


def _build_component_cm_mounting_plate_entries(component_df: pd.DataFrame) -> list[str]:
    """Build one flat Mounting plate list from local non-strip component Names only."""
    cm_source_df = _build_component_cm_source_df(component_df)
    if cm_source_df.empty:
        return []

    door_name_set = set(_build_component_cm_door_source_entries(component_df))
    mounting_plate_df = cm_source_df.loc[
        cm_source_df["Name"].map(_stringify_cell).ne("")
        & ~cm_source_df["Name"].map(_stringify_cell).isin(door_name_set)
        & cm_source_df["TYPE"].map(_detect_fuse_voltage_group).isna()
        & cm_source_df.apply(
            lambda row: _classify_component_strip_relay_group(row.get("Name"), row.get("TYPE")) == "",
            axis=1,
        )
    ].copy()
    return _deduplicate_marking_names_default(
        mounting_plate_df["Name"].map(_stringify_cell).tolist(),
        column_key="mounting_plate",
    )


def _build_component_cm_component_entries(component_df: pd.DataFrame) -> list[str]:
    """Build one CM Component-column list in strip-style group order with blank rows between groups."""
    cm_source_df = _build_component_cm_source_df(component_df)
    if cm_source_df.empty:
        return []

    ordered_groups = [
        *_build_component_cm_fuse_groups(cm_source_df),
        *_build_component_cm_relay_groups(cm_source_df),
        *_build_component_cm_button_other_groups(cm_source_df),
    ]
    blocked_rule_keys = {
        _component_cm_marking_rule_key(door_name)
        for door_name in _build_component_cm_door_entries(component_df)
        if not _is_component_cm_p92_name(door_name)
    }
    return _deduplicate_component_cm_entries(ordered_groups, blocked_rule_keys=blocked_rule_keys)


def _build_component_cm_sheet_df(component_df: pd.DataFrame | None = None) -> pd.DataFrame:
    """Build a CM sheet with grouped Component entries and Door rows from local component Names."""
    if component_df is None or component_df.empty or "Name" not in component_df.columns:
        return _ComponentCmSheetDataFrame(columns=_COMPONENT_CM_COLUMNS)

    mounting_plate_entries = _build_component_cm_mounting_plate_entries(component_df)
    door_names = _build_component_cm_door_entries(component_df)
    component_entries = _build_component_cm_component_entries(component_df)
    row_count = max(len(mounting_plate_entries), len(component_entries), len(door_names))
    return _ComponentCmSheetDataFrame(
        {
            "Mounting plate": mounting_plate_entries + [""] * (row_count - len(mounting_plate_entries)),
            "Component": component_entries + [""] * (row_count - len(component_entries)),
            "Door": door_names + [""] * (row_count - len(door_names)),
        },
        columns=_COMPONENT_CM_COLUMNS,
    )


def _extract_component_cm_component_groups(component_entries: list[Any]) -> list[tuple[str, list[str]]]:
    """Parse one flat CM Component-column list into ordered section-label/name groups."""
    ordered_groups: list[tuple[str, list[str]]] = []
    current_label = ""
    current_names: list[str] = []
    for entry_value in [*component_entries, ""]:
        text_value = _stringify_cell(entry_value)
        if text_value == "":
            if current_label != "":
                ordered_groups.append((current_label, current_names.copy()))
            current_label = ""
            current_names = []
            continue
        if current_label == "":
            current_label = text_value
            continue
        current_names.append(text_value)
    return ordered_groups


def _build_component_cm_group_entries(groups: list[tuple[str, list[str]]]) -> list[str]:
    """Flatten ordered CM section groups back into one column with blank separators."""
    valid_groups = [(group_label, group_names) for group_label, group_names in groups if _stringify_cell(group_label) != ""]
    entries: list[str] = []
    for group_index, (group_label, group_names) in enumerate(valid_groups):
        entries.append(_stringify_cell(group_label))
        entries.extend([_stringify_cell(group_name) for group_name in group_names if _stringify_cell(group_name) != ""])
        if group_index < len(valid_groups) - 1:
            entries.append("")
    return entries


def _split_final_component_cm_layout(
    component_cm_df: pd.DataFrame,
) -> tuple[_ComponentCmSheetDataFrame, pd.DataFrame, pd.DataFrame]:
    """Split the final-sheet CM Component column into main, FUSES, and RELAYS right-side blocks."""
    working_df = component_cm_df.copy().reset_index(drop=True)
    component_entries = working_df.get("Component", pd.Series(dtype=object)).tolist()
    ordered_groups = _extract_component_cm_component_groups(component_entries)

    main_component_groups: list[tuple[str, list[str]]] = []
    fuse_groups: list[tuple[str, list[str]]] = []
    relay_groups: list[tuple[str, list[str]]] = []
    for group_label, group_names in ordered_groups:
        normalized_group_label = _stringify_cell(group_label)
        if normalized_group_label in _COMPONENT_FINAL_MOVED_CM_FUSE_GROUP_LABELS:
            fuse_groups.append((normalized_group_label, group_names))
            continue
        relay_display_label = _COMPONENT_FINAL_MOVED_CM_RELAY_GROUP_LABEL_ALIASES.get(normalized_group_label, "")
        if relay_display_label:
            relay_groups.append((relay_display_label, group_names))
            continue
        main_component_groups.append((normalized_group_label, group_names))

    main_component_entries = _build_component_cm_group_entries(main_component_groups)
    row_count = len(working_df.index)
    working_df["Component"] = main_component_entries + [""] * max(0, row_count - len(main_component_entries))
    main_component_df = _ComponentCmSheetDataFrame(working_df, columns=_COMPONENT_CM_COLUMNS)

    fuse_block_entries = _build_component_cm_group_entries(fuse_groups)
    relay_block_entries = _build_component_cm_group_entries(relay_groups)
    fuse_block_df = pd.DataFrame({_COMPONENT_FINAL_FUSES_BLOCK_TITLE: fuse_block_entries})
    relay_block_df = pd.DataFrame({_COMPONENT_FINAL_RELAYS_BLOCK_TITLE: relay_block_entries})
    return main_component_df, fuse_block_df, relay_block_df


def _extract_component_cabinet_parts(name_value: Any) -> tuple[str, str, str] | None:
    """Extract one raw cabinet ID, sanitized sheet-safe cabinet ID, and normalized cabinet-local Name."""
    text = _stringify_cell(name_value)
    if not text:
        return None
    match = _COMPONENT_CABINET_NAME_PATTERN.match(text)
    if not match:
        return None

    raw_cabinet_id = _stringify_cell(match.group("cabinet_id")).upper()
    cabinet_id = _sanitize_component_cabinet_sheet_name(raw_cabinet_id)
    normalized_name = _stringify_cell(match.group("normalized_name"))
    if not cabinet_id or not normalized_name:
        return None
    return raw_cabinet_id, cabinet_id, normalized_name


def _build_component_cabinet_map(
    component_marking_df: pd.DataFrame,
) -> tuple[dict[str, pd.DataFrame], dict[str, Any]]:
    """Build per-cabinet component datasets from kept +A*-prefixed rows without altering the main dataset."""
    cabinet_stats: dict[str, Any] = {
        "cabinet_ids": [],
        "raw_cabinet_ids": [],
        "sanitized_cabinet_ids": [],
        "sheet_name_sanitizations": [],
        "row_counts": {},
        "example_transforms": [],
    }
    if component_marking_df.empty or "Name" not in component_marking_df.columns:
        return {}, cabinet_stats

    cabinet_source_df = component_marking_df.copy().reset_index(drop=True)
    cabinet_parts = cabinet_source_df["Name"].map(_extract_component_cabinet_parts)
    cabinet_source_df["_cabinet_raw_id"] = cabinet_parts.map(lambda parts: parts[0] if parts else "")
    cabinet_source_df["_cabinet_id"] = cabinet_parts.map(lambda parts: parts[1] if parts else "")
    cabinet_source_df["_cabinet_normalized_name"] = cabinet_parts.map(lambda parts: parts[2] if parts else "")
    cabinet_source_df = cabinet_source_df.loc[cabinet_source_df["_cabinet_id"].ne("")].copy()
    if cabinet_source_df.empty:
        return {}, cabinet_stats

    cabinet_stats["raw_cabinet_ids"] = list(
        dict.fromkeys(
            cabinet_source_df["_cabinet_raw_id"].map(_stringify_cell).tolist()
        )
    )
    cabinet_stats["sanitized_cabinet_ids"] = list(
        dict.fromkeys(
            cabinet_source_df["_cabinet_id"].map(_stringify_cell).tolist()
        )
    )
    cabinet_stats["cabinet_ids"] = list(cabinet_stats["sanitized_cabinet_ids"])
    for raw_cabinet_id, sanitized_cabinet_id in (
        cabinet_source_df.loc[:, ["_cabinet_raw_id", "_cabinet_id"]]
        .drop_duplicates()
        .itertuples(index=False, name=None)
    ):
        if _stringify_cell(raw_cabinet_id) != _stringify_cell(sanitized_cabinet_id):
            cabinet_stats["sheet_name_sanitizations"].append(
                f"{_stringify_cell(raw_cabinet_id)} -> {_stringify_cell(sanitized_cabinet_id)}"
            )

    cabinet_map: dict[str, pd.DataFrame] = {}
    for cabinet_id, cabinet_df in cabinet_source_df.groupby("_cabinet_id", sort=False, dropna=False):
        normalized_df = cabinet_df.copy()
        normalized_df["Name"] = normalized_df["_cabinet_normalized_name"]
        if len(cabinet_stats["example_transforms"]) < 6:
            example_rows = normalized_df.head(3)
            cabinet_stats["example_transforms"].extend(
                [
                    f"{cabinet_id}: {original_name} -> {normalized_name}"
                    for original_name, normalized_name in zip(
                        cabinet_df["Name"].head(3).tolist(),
                        example_rows["Name"].tolist(),
                    )
                ]
            )
        normalized_df = normalized_df.drop(
            columns=["_cabinet_raw_id", "_cabinet_id", "_cabinet_normalized_name"]
        )
        normalized_df = normalized_df.reset_index(drop=True)
        normalized_df.attrs["cabinet_id"] = raw_cabinet_id
        cabinet_map[cabinet_id] = normalized_df
        cabinet_stats["row_counts"][cabinet_id] = len(normalized_df)

    return cabinet_map, cabinet_stats


def _classify_component_category(name_value: Any, type_value: Any) -> str:
    """Classify component rows into grouped production/output categories."""
    normalized_name = _normalize_component_name(name_value)
    normalized_type = _normalize_component_type(type_value)

    if normalized_name.startswith("-S"):
        return "BUTTON"
    if normalized_name.startswith("-P") and normalized_type in _BUTTON_P_TYPES:
        return "BUTTON"
    if normalized_type in _FUSE_TYPES:
        return "FUSE"
    if normalized_type in _RELAY_1P_TYPES:
        return "RELAY_1P"
    if normalized_type in _RELAY_4P_TYPES:
        return "RELAY_4P"
    if normalized_type in _RELAY_2P_TYPES:
        return "RELAY_2P"
    return "OTHER"


def _component_name_sort_key(value: Any) -> str:
    """Build a stable, case-insensitive Name sort key for grouped sections."""
    return _stringify_cell(value).casefold()


def _component_fuse_name_sort_key(name: Any) -> tuple[int, int, int, int, str, str]:
    """Build a natural sort key where A-suffix fuses follow normal rows of the same base."""
    fuse_name = _stringify_cell(name)
    normalized_name = fuse_name.casefold()
    a_suffix_match = _FUSE_A_SUFFIX_SORT_PATTERN.match(fuse_name)
    if a_suffix_match:
        suffix_number_text = _stringify_cell(a_suffix_match.group("suffix_number"))
        return (
            0,
            int(a_suffix_match.group("family")),
            1,
            int(suffix_number_text) if suffix_number_text else 0,
            _stringify_cell(a_suffix_match.group("suffix_text")).casefold(),
            normalized_name,
        )

    match = _FUSE_NAME_SORT_PATTERN.match(fuse_name)
    if not match:
        return (1, 0, 0, 0, "", normalized_name)

    numeric_value = int(match.group("number"))
    suffix_text = _stringify_cell(match.group("suffix")).casefold()
    return (
        0,
        numeric_value // 10,
        0 if not suffix_text else 2,
        numeric_value % 10 if not suffix_text else 0,
        suffix_text,
        normalized_name,
    )


def _is_component_timed_relay_name(name_value: Any) -> bool:
    """Reuse the Component Correction timed family detection for K192* and K192A* Names."""
    relay_name = _stringify_cell(name_value)
    return bool(_TIMED_RELAY_PATTERN.match(relay_name) or _TIMED_RELAY_A_PATTERN.match(relay_name))


def _classify_component_strip_relay_group(name_value: Any, type_value: Any) -> str:
    """Classify one relay strip source row using base TYPE plus Component Correction timed-name rules."""
    normalized_type = _normalize_component_type(type_value)
    base_group = _RELAY_STRIP_GROUP_BY_TYPE.get(normalized_type, "")
    if not base_group:
        return ""
    if _is_component_timed_relay_name(name_value):
        return "timed"
    return base_group


def _component_relay_strip_type_priority(type_value: Any) -> tuple[int, str]:
    """Prefer base relay TYPE rows over socket rows when deduplicating one physical relay stack by Name."""
    normalized_type = _normalize_component_type(type_value)
    return (_RELAY_STRIP_TYPE_PRIORITY.get(normalized_type, 99), normalized_type)


def _deduplicate_component_relay_strip_source(relay_df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    """Collapse one physical relay stack to one strip source row per Name."""
    if relay_df.empty:
        return relay_df.copy().reset_index(drop=True), 0

    working_df = relay_df.copy()
    working_df["_relay_name"] = working_df["Name"].map(_stringify_cell)
    working_df = working_df.loc[working_df["_relay_name"].ne("")].copy()
    if working_df.empty:
        return working_df.reset_index(drop=True), 0

    deduplicated_rows: list[dict[str, Any]] = []
    duplicates_removed = 0
    for _, same_name_df in working_df.groupby("_relay_name", sort=False, dropna=False):
        priority_values = same_name_df["TYPE"].map(_component_relay_strip_type_priority).tolist()
        prioritized_df = same_name_df.copy()
        prioritized_df[["_relay_type_priority", "_relay_type_priority_text"]] = pd.DataFrame(
            priority_values,
            index=prioritized_df.index,
        )
        representative_row = prioritized_df.sort_values(
            by=["_relay_type_priority", "_relay_type_priority_text", "_original_order"],
            kind="mergesort",
        ).iloc[0]
        deduplicated_rows.append(representative_row.to_dict())
        duplicates_removed += len(same_name_df) - 1

    deduplicated_df = pd.DataFrame(deduplicated_rows)
    deduplicated_df = deduplicated_df.drop(
        columns=["_relay_name", "_relay_type_priority", "_relay_type_priority_text"],
        errors="ignore",
    )
    return deduplicated_df.reset_index(drop=True), duplicates_removed


def _sort_component_relay_group_df(relay_group_df: pd.DataFrame) -> pd.DataFrame:
    """Sort one relay strip group by natural K-name order while preserving input order as a tiebreaker."""
    if relay_group_df.empty:
        return relay_group_df.reset_index(drop=True)

    sorted_df = relay_group_df.copy()
    relay_sort_keys = sorted_df["Name"].map(_component_relay_name_sort_key).tolist()
    sorted_df[[
        "_relay_sort_group",
        "_relay_sort_family",
        "_relay_sort_variant_kind",
        "_relay_sort_variant_number",
        "_relay_sort_suffix",
        "_relay_sort_text",
    ]] = pd.DataFrame(relay_sort_keys, index=sorted_df.index)
    sorted_df = sorted_df.sort_values(
        by=[
            "_relay_sort_group",
            "_relay_sort_family",
            "_relay_sort_variant_kind",
            "_relay_sort_variant_number",
            "_relay_sort_suffix",
            "_relay_sort_text",
            "_original_order",
        ],
        kind="mergesort",
    ).drop(
        columns=[
            "_relay_sort_group",
            "_relay_sort_family",
            "_relay_sort_variant_kind",
            "_relay_sort_variant_number",
            "_relay_sort_suffix",
            "_relay_sort_text",
        ]
    )
    return sorted_df.reset_index(drop=True)


def _component_timed_relay_name_sort_key(name: Any) -> tuple[int, int, str]:
    """Reuse the Component Correction timed ordering: K192* first, then K192A*, both numeric."""
    relay_name = _stringify_cell(name)
    normalized_name = relay_name.casefold()

    timed_match = _TIMED_RELAY_PATTERN.match(relay_name)
    if timed_match:
        return (0, int(timed_match.group("suffix_number")), normalized_name)

    timed_a_match = _TIMED_RELAY_A_PATTERN.match(relay_name)
    if timed_a_match:
        return (1, int(timed_a_match.group("suffix_number")), normalized_name)

    return (2, 10**9, normalized_name)


def _sort_component_timed_relay_group_df(relay_group_df: pd.DataFrame) -> pd.DataFrame:
    """Sort timed relay rows using the Component Correction K192/K192A ordering rules."""
    if relay_group_df.empty:
        return relay_group_df.reset_index(drop=True)

    sorted_df = relay_group_df.copy()
    timed_sort_keys = sorted_df["Name"].map(_component_timed_relay_name_sort_key).tolist()
    sorted_df[["_timed_sort_group", "_timed_sort_number", "_timed_sort_text"]] = pd.DataFrame(
        timed_sort_keys,
        index=sorted_df.index,
    )
    sorted_df = sorted_df.sort_values(
        by=["_timed_sort_group", "_timed_sort_number", "_timed_sort_text", "_original_order"],
        kind="mergesort",
    ).drop(columns=["_timed_sort_group", "_timed_sort_number", "_timed_sort_text"])
    return sorted_df.reset_index(drop=True)


def _detect_component_relay_strip_space(type_value: Any, relay_group: str) -> float:
    """Return the printed strip width for one relay strip row based on relay strip group and base TYPE."""
    normalized_type = _normalize_component_type(type_value)
    if relay_group == "4_pole":
        return _RELAY_STRIP_4POLE_WIDTH
    if relay_group == "1_pole":
        return _RELAY_STRIP_1POLE_WIDTH
    if normalized_type in _RELAY_RE22_TYPES:
        return _RELAY_STRIP_RE22_WIDTH
    return _RELAY_STRIP_2POLE_WIDTH


def _component_relay_name_sort_key(name: Any) -> tuple[int, int, int, int, str, str]:
    """Build a natural relay Name sort key where K*A* names follow normal K names."""
    relay_name = _stringify_cell(name)
    normalized_name = relay_name.casefold()

    a_suffix_match = _RELAY_NAME_A_SUFFIX_SORT_PATTERN.match(relay_name)
    if a_suffix_match:
        suffix_number_text = _stringify_cell(a_suffix_match.group("suffix_number"))
        return (
            0,
            int(a_suffix_match.group("family")),
            1,
            int(suffix_number_text) if suffix_number_text else 0,
            _stringify_cell(a_suffix_match.group("suffix_text")).casefold(),
            normalized_name,
        )

    match = _RELAY_NAME_SORT_PATTERN.match(relay_name)
    if not match:
        return (1, 10**9, 10**9, 10**9, "", normalized_name)

    numeric_value = int(match.group("number"))
    suffix_text = _stringify_cell(match.group("suffix")).casefold()
    return (
        0,
        numeric_value // 10,
        0 if not suffix_text else 2,
        numeric_value % 10 if not suffix_text else 0,
        suffix_text,
        normalized_name,
    )


def _component_strip_relay_preview_names(relay_group_df: pd.DataFrame, limit: int = 3) -> list[str]:
    """Return the first few relay Names detected for developer debugging."""
    if relay_group_df.empty or "Name" not in relay_group_df.columns:
        return []
    preview_names = [
        _stringify_cell(name_value)
        for name_value in relay_group_df["Name"].tolist()
        if _stringify_cell(name_value)
    ]
    return preview_names[:limit]


def _validate_component_export_df(sheet_name: str, sheet_df: pd.DataFrame) -> None:
    """Validate one flat Component export DataFrame for shared workbook compatibility."""
    duplicate_columns = sheet_df.columns[sheet_df.columns.duplicated()].tolist()
    if duplicate_columns:
        duplicate_list = ", ".join(repr(column_name) for column_name in duplicate_columns)
        raise ValueError(f"{sheet_name} contains duplicate export columns: {duplicate_list}")

    invalid_cell_types = (pd.Series, pd.DataFrame, list, dict)
    for row_number, row_values in enumerate(sheet_df.itertuples(index=False, name=None), start=2):
        for column_name, cell_value in zip(sheet_df.columns, row_values):
            if isinstance(cell_value, invalid_cell_types):
                raise ValueError(
                    f"{sheet_name} contains non-scalar value in column {column_name!r}, row {row_number}"
                )


def _validate_component_export_sheet(sheet_name: str, sheet_content: Any) -> None:
    """Keep Component sheets compatible with the shared workbook exporter contract."""
    if isinstance(sheet_content, pd.DataFrame):
        _validate_component_export_df(sheet_name, sheet_content)
        return

    if isinstance(sheet_content, dict) and sheet_content.get("layout") == "component_strip":
        _validate_component_export_df(
            f"{sheet_name} (Fuse strip)",
            sheet_content.get("fuse_strip_df", pd.DataFrame(columns=_COMPONENT_STRIP_SIDE_COLUMNS)),
        )
        _validate_component_export_df(
            f"{sheet_name} (Relay strip)",
            sheet_content.get("relay_strip_df", pd.DataFrame(columns=_COMPONENT_STRIP_SIDE_COLUMNS)),
        )
        return

    raise TypeError(f"{sheet_name} must be a pandas DataFrame or approved layout object for export")


def _coerce_excel_number(value: Any) -> int | float | None:
    """Return a numeric value when Excel should store the cell as a number."""
    if value is None or pd.isna(value):
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        numeric_value = float(value)
        return int(numeric_value) if numeric_value.is_integer() else numeric_value

    text = _stringify_cell(value).replace(" ", "")
    if not text:
        return None

    numeric_value = pd.to_numeric(text, errors="coerce")
    if pd.isna(numeric_value) and "," in text and "." not in text:
        numeric_value = pd.to_numeric(text.replace(",", "."), errors="coerce")
    if pd.isna(numeric_value):
        return None

    numeric_value = float(numeric_value)
    return int(numeric_value) if numeric_value.is_integer() else numeric_value


def _sort_grouped_component_rows(component_df: pd.DataFrame) -> pd.DataFrame:
    """Sort grouped component rows by Name while keeping equal names adjacent and stable."""
    if component_df.empty:
        return component_df.copy().reset_index(drop=True)

    sorted_df = component_df.copy()
    sorted_df["_name_sort_key"] = sorted_df["Name"].map(_component_name_sort_key)
    sorted_df = sorted_df.sort_values(
        by=["_name_sort_key", "_original_order"],
        kind="mergesort",
    ).drop(columns=["_name_sort_key"])
    return sorted_df.reset_index(drop=True)


def _component_group_label_from_category(category_value: Any) -> str:
    """Map internal production categories to flat Component Marking group labels."""
    normalized_category = _stringify_cell(category_value).upper()
    if normalized_category == "FUSE":
        return "FUSES"
    if normalized_category in {"RELAY_1P", "RELAY_2P", "RELAY_4P"}:
        return normalized_category
    if normalized_category == "BUTTON":
        return "BUTTONS"
    return "OTHER"


def _drop_production_only_component_columns(component_df: pd.DataFrame) -> pd.DataFrame:
    """Keep production-only source fields out of non-production workbook sheets."""
    return component_df.drop(columns=list(_PRODUCTION_ONLY_COMPONENT_COLUMNS), errors="ignore").copy()


def _filter_component_production_fuse_rows(component_df: pd.DataFrame) -> pd.DataFrame:
    """Remove fuse rows from the production-only flow without changing row shape."""
    if component_df.empty:
        return component_df.copy().reset_index(drop=True)

    group_values = component_df.get("Group", pd.Series("", index=component_df.index, dtype=object)).map(
        _stringify_cell
    )
    category_values = component_df.get(
        "Category", pd.Series("", index=component_df.index, dtype=object)
    ).map(_stringify_cell)
    type_values = component_df.get("TYPE", pd.Series("", index=component_df.index, dtype=object)).map(
        _normalize_component_type
    )

    include_mask = ~(
        group_values.str.upper().eq("FUSES")
        | category_values.str.upper().eq("FUSE")
        | type_values.isin(_FUSE_TYPES)
    )
    return component_df.loc[include_mask].copy().reset_index(drop=True)


def _split_component_groups(
    component_marking_df: pd.DataFrame,
) -> tuple[list[tuple[str, pd.DataFrame, str]], pd.DataFrame, dict[str, int]]:
    """Split component rows into ordered grouped sections plus remaining rows."""
    group_counts = {"relay_rows": 0, "fuse_rows": 0, "button_rows": 0}
    if component_marking_df.empty:
        return [], component_marking_df.copy().reset_index(drop=True), group_counts

    working_df = component_marking_df.copy().reset_index(drop=True)
    for column_name in ("Name", "TYPE", "Quantity", "Description", "Category"):
        if column_name not in working_df.columns:
            working_df[column_name] = ""
    working_df["_original_order"] = range(len(working_df))

    grouped_sections: list[tuple[str, pd.DataFrame, str]] = []
    grouped_categories: set[str] = set()
    for section_label, section_categories, count_key in _GROUPED_COMPONENT_SECTIONS:
        section_df = working_df.loc[working_df["Category"].isin(section_categories)].copy()
        sorted_section_df = _sort_grouped_component_rows(section_df)
        grouped_sections.append((section_label, sorted_section_df, count_key))
        grouped_categories.update(section_categories)
        group_counts[count_key] = len(sorted_section_df)

    other_df = working_df.loc[
        ~working_df["Category"].isin(grouped_categories)
    ].copy().sort_values("_original_order", kind="mergesort").reset_index(drop=True)
    return grouped_sections, other_df, group_counts


def _build_component_production_source_df(component_marking_df: pd.DataFrame) -> pd.DataFrame:
    """Prepare one production-order source frame without changing any row-level content."""
    if component_marking_df.empty:
        return pd.DataFrame(
            columns=["Name", "Article No.", "TYPE", "Quantity", "Description", "Category", "_original_order"]
        )

    working_df = component_marking_df.copy().reset_index(drop=True)
    for column_name in ("Name", "Article No.", "TYPE", "Quantity", "Description", "Category"):
        if column_name not in working_df.columns:
            working_df[column_name] = ""
    category_is_blank = working_df["Category"].map(_stringify_cell).eq("")
    if bool(category_is_blank.any()):
        working_df.loc[category_is_blank, "Category"] = working_df.loc[category_is_blank].apply(
            lambda row: _classify_component_category(row.get("Name"), row.get("TYPE")),
            axis=1,
        )
    working_df["_original_order"] = range(len(working_df))
    return working_df


def _build_component_production_fuse_sections(
    working_df: pd.DataFrame,
) -> tuple[list[tuple[str, pd.DataFrame]], int]:
    """Build production fuse sections in the same voltage and natural-name order as strip/marking."""
    fuse_df = working_df.loc[working_df["Category"].eq("FUSE")].copy()
    if fuse_df.empty:
        return [], 0

    ordered_sections: list[tuple[str, pd.DataFrame]] = []
    fuse_row_count = 0
    for voltage_group in _COMPONENT_STRIP_GROUP_ORDER:
        voltage_df = fuse_df.loc[
            fuse_df["TYPE"].map(_detect_fuse_voltage_group).eq(voltage_group)
        ].copy()
        if voltage_df.empty:
            continue

        fuse_sort_keys = voltage_df["Name"].map(_component_fuse_name_sort_key).tolist()
        voltage_df[[
            "_fuse_sort_group",
            "_fuse_sort_family",
            "_fuse_sort_variant_kind",
            "_fuse_sort_variant_number",
            "_fuse_sort_suffix",
            "_fuse_sort_text",
        ]] = pd.DataFrame(fuse_sort_keys, index=voltage_df.index)
        voltage_df = voltage_df.sort_values(
            by=[
                "_fuse_sort_group",
                "_fuse_sort_family",
                "_fuse_sort_variant_kind",
                "_fuse_sort_variant_number",
                "_fuse_sort_suffix",
                "_fuse_sort_text",
                "_original_order",
            ],
            kind="mergesort",
        ).drop(
            columns=[
                "_fuse_sort_group",
                "_fuse_sort_family",
                "_fuse_sort_variant_kind",
                "_fuse_sort_variant_number",
                "_fuse_sort_suffix",
                "_fuse_sort_text",
            ]
        ).reset_index(drop=True)
        ordered_sections.append((_COMPONENT_CM_FUSE_GROUP_LABELS[voltage_group], voltage_df))
        fuse_row_count += len(voltage_df)
    return ordered_sections, fuse_row_count


def _build_component_production_relay_sections(
    working_df: pd.DataFrame,
) -> tuple[list[tuple[str, pd.DataFrame]], int]:
    """Build production relay sections in CM display-group order without deduplicating rows."""
    relay_df = working_df.loc[working_df["Category"].isin({"RELAY_1P", "RELAY_2P", "RELAY_4P"})].copy()
    if relay_df.empty:
        return [], 0

    production_relay_rows: list[pd.DataFrame] = []
    for _, same_name_df in relay_df.groupby("Name", sort=False, dropna=False):
        production_relay_group = _infer_component_production_relay_group(same_name_df)
        if production_relay_group == "":
            continue
        grouped_name_df = same_name_df.copy()
        grouped_name_df["_production_relay_group"] = production_relay_group
        production_relay_rows.append(grouped_name_df)
    if not production_relay_rows:
        return [], 0

    production_relay_df = pd.concat(production_relay_rows, ignore_index=True)
    ordered_sections: list[tuple[str, pd.DataFrame]] = []
    relay_row_count = 0
    for group_label in _COMPONENT_CM_DISPLAY_RELAY_GROUP_ORDER:
        group_df = production_relay_df.loc[
            production_relay_df["_production_relay_group"].eq(group_label)
        ].copy()
        if group_df.empty:
            continue
        if group_label == "Timed relays":
            group_df = _sort_component_timed_relay_group_df(group_df)
        else:
            group_df = _sort_component_relay_group_df(group_df)
        ordered_sections.append((group_label, group_df.reset_index(drop=True)))
        relay_row_count += len(group_df)
    return ordered_sections, relay_row_count


def _build_component_production_button_other_sections(
    working_df: pd.DataFrame,
) -> tuple[list[tuple[str, pd.DataFrame]], int]:
    """Build production button/other sections after fuse and relay groups."""
    button_df = _sort_grouped_component_rows(
        working_df.loc[working_df["Category"].eq("BUTTON")].copy()
    )
    other_df = _sort_grouped_component_rows(
        working_df.loc[working_df["Category"].eq("OTHER")].copy()
    )

    ordered_sections: list[tuple[str, pd.DataFrame]] = []
    if not button_df.empty:
        ordered_sections.append((_COMPONENT_CM_BUTTONS_LABEL, button_df))
    if not other_df.empty:
        ordered_sections.append((_OTHER_SECTION_LABEL, other_df))
    return ordered_sections, len(button_df)


def _build_component_production_ordered_sections(
    component_marking_df: pd.DataFrame,
) -> tuple[list[tuple[str, pd.DataFrame]], dict[str, int]]:
    """Build production workbook sections in the same visible order as Marking."""
    group_counts = {"relay_rows": 0, "fuse_rows": 0, "button_rows": 0}
    working_df = _filter_component_production_fuse_rows(
        _build_component_production_source_df(component_marking_df)
    )
    if working_df.empty:
        return [], group_counts

    fuse_sections, group_counts["fuse_rows"] = _build_component_production_fuse_sections(working_df)
    relay_sections, group_counts["relay_rows"] = _build_component_production_relay_sections(working_df)
    button_other_sections, group_counts["button_rows"] = _build_component_production_button_other_sections(
        working_df
    )
    return [*fuse_sections, *relay_sections, *button_other_sections], group_counts


def _build_production_section_row(label: str) -> dict[str, Any]:
    """Create a visual section row for grouped component entries."""
    return {
        "Name": label,
        "Article No.": "",
        "TYPE": "",
        "Quantity": "",
        "Marked": "",
        "Description": "",
        "Comments": "",
        "_is_section": True,
        "_is_separator": False,
        _PRODUCTION_TECHNICAL_FLAG_COLUMN: 0,
    }


def _build_production_separator_row() -> dict[str, Any]:
    """Create an empty visual separator row after a grouped section."""
    return {
        "Name": "",
        "Article No.": "",
        "TYPE": "",
        "Quantity": "",
        "Marked": "",
        "Description": "",
        "Comments": "",
        "_is_section": False,
        "_is_separator": True,
        _PRODUCTION_TECHNICAL_FLAG_COLUMN: 0,
    }


def _component_rows_to_production_records(component_df: pd.DataFrame) -> list[dict[str, Any]]:
    """Convert actual component rows into production-sheet records."""
    filtered_component_df = _filter_component_production_fuse_rows(component_df)
    if filtered_component_df.empty:
        return []

    production_rows = pd.DataFrame(index=filtered_component_df.index)
    for column_name in ("Name", "Article No.", "TYPE", "Quantity", "Description"):
        if column_name in filtered_component_df.columns:
            production_rows[column_name] = filtered_component_df[column_name]
        else:
            production_rows[column_name] = ""
    production_rows["Marked"] = ""
    production_rows["Comments"] = ""
    production_rows["_is_section"] = False
    production_rows["_is_separator"] = False
    production_rows[_PRODUCTION_TECHNICAL_FLAG_COLUMN] = 1
    return production_rows.loc[:, [*_PRODUCTION_COLUMNS, "_is_section", "_is_separator", _PRODUCTION_TECHNICAL_FLAG_COLUMN]].to_dict("records")


def _build_component_marking_sheet_df(
    component_marking_df: pd.DataFrame,
) -> pd.DataFrame:
    """Build a flat Component Marking data sheet with a stable Group column."""
    if component_marking_df.empty:
        output_df = _drop_production_only_component_columns(component_marking_df).reset_index(drop=True)
        if "Group" not in output_df.columns:
            output_df["Group"] = pd.Series(dtype=object)
        return output_df

    output_df = _drop_production_only_component_columns(component_marking_df).reset_index(drop=True)
    output_df["Group"] = output_df.get("Category", pd.Series(index=output_df.index, dtype=object)).map(
        _component_group_label_from_category
    ).fillna("OTHER")

    if "Category" in output_df.columns:
        category_column = output_df.pop("Category")
        output_df["Category"] = category_column
        output_df = output_df.drop(columns=["Category"])
    return output_df


def _detect_fuse_voltage_group(type_value: Any) -> str | None:
    """Map fuse TYPE values to the required strip voltage groups."""
    normalized_type = _normalize_component_type(type_value)
    return _FUSE_TYPE_TO_VOLTAGE_GROUP.get(normalized_type)


def _build_component_fuse_strip_rows(working_df: pd.DataFrame) -> tuple[list[tuple[Any, Any]], dict[str, Any]]:
    """Build left-side fuse strip rows while keeping the existing fuse logic intact."""
    strip_stats: dict[str, Any] = {
        "24vdc_rows": 0,
        "230vac_rows": 0,
        "f92_wide_name": "",
    }
    fuse_df = working_df.loc[working_df["Group"].map(_stringify_cell).eq("FUSES")].copy()
    if fuse_df.empty:
        return [], strip_stats

    sorted_group_dfs: dict[str, pd.DataFrame] = {}
    last_f92_marker: tuple[str, int] | None = None
    for voltage_group in _COMPONENT_STRIP_GROUP_ORDER:
        voltage_df = fuse_df.loc[
            fuse_df["TYPE"].map(_detect_fuse_voltage_group).eq(voltage_group)
        ].copy()
        if not voltage_df.empty:
            fuse_sort_keys = voltage_df["Name"].map(_component_fuse_name_sort_key).tolist()
            voltage_df[[
                "_fuse_sort_group",
                "_fuse_sort_family",
                "_fuse_sort_variant_kind",
                "_fuse_sort_variant_number",
                "_fuse_sort_suffix",
                "_fuse_sort_text",
            ]] = pd.DataFrame(fuse_sort_keys, index=voltage_df.index)
            voltage_df = voltage_df.sort_values(
                by=[
                    "_fuse_sort_group",
                    "_fuse_sort_family",
                    "_fuse_sort_variant_kind",
                    "_fuse_sort_variant_number",
                    "_fuse_sort_suffix",
                    "_fuse_sort_text",
                    "_original_order",
                ],
                kind="mergesort",
            ).drop(
                columns=[
                    "_fuse_sort_group",
                    "_fuse_sort_family",
                    "_fuse_sort_variant_kind",
                    "_fuse_sort_variant_number",
                    "_fuse_sort_suffix",
                    "_fuse_sort_text",
                ]
            ).reset_index(drop=True)
        else:
            voltage_df = voltage_df.reset_index(drop=True)
        sorted_group_dfs[voltage_group] = voltage_df
        strip_stats["24vdc_rows" if voltage_group == "24VDC" else "230vac_rows"] = len(voltage_df)

    strip_rows: list[tuple[Any, Any]] = []
    for voltage_group in _COMPONENT_STRIP_GROUP_ORDER:
        voltage_df = sorted_group_dfs[voltage_group]
        if voltage_df.empty:
            continue
        for row_index, row in enumerate(voltage_df.to_dict("records")):
            if _F92_FUSE_PATTERN.match(_stringify_cell(row.get("Name"))):
                last_f92_marker = (voltage_group, row_index)

    for voltage_group in _COMPONENT_STRIP_GROUP_ORDER:
        voltage_df = sorted_group_dfs[voltage_group]
        if voltage_df.empty:
            continue

        if voltage_group == "24VDC":
            strip_rows.append((_FUSE_STRIP_WIDTH, voltage_group))
        else:
            strip_rows.append((_FUSE_STRIP_230VAC_SEPARATOR_SPACE, ""))
        for row_index, row in enumerate(voltage_df.to_dict("records")):
            row_name = _stringify_cell(row.get("Name"))
            row_space = (
                _FUSE_STRIP_COVERED_WIDTH
                if last_f92_marker == (voltage_group, row_index)
                else _FUSE_STRIP_WIDTH
            )
            if row_space == _FUSE_STRIP_COVERED_WIDTH:
                strip_stats["f92_wide_name"] = row_name
            strip_rows.append((row_space, row_name))

    return strip_rows, strip_stats


def _build_component_relay_strip_rows(working_df: pd.DataFrame) -> tuple[list[tuple[Any, Any]], dict[str, Any]]:
    """Build right-side relay strip rows in 2-pole, 4-pole, timed, 1-pole order."""
    strip_stats = {
        "2pole_rows": 0,
        "4pole_rows": 0,
        "timed_rows": 0,
        "1pole_rows": 0,
        "duplicate_rows_removed": 0,
        "clipfix_rows": 0,
        "2pole_preview_names": [],
        "4pole_preview_names": [],
        "timed_preview_names": [],
        "1pole_preview_names": [],
        "start_rows": 0,
        "stop_rows": 0,
        "width_15_8_rows": 0,
        "width_22_5_rows": 0,
        "width_27_rows": 0,
        "width_6_2_rows": 0,
    }

    relay_df = working_df.copy()
    relay_df["_relay_group"] = relay_df.apply(
        lambda row: _classify_component_strip_relay_group(row.get("Name"), row.get("TYPE")),
        axis=1,
    )
    relay_df = relay_df.loc[relay_df["_relay_group"].ne("")].copy()
    relay_df, strip_stats["duplicate_rows_removed"] = _deduplicate_component_relay_strip_source(relay_df)

    two_pole_df = _sort_component_relay_group_df(
        relay_df.loc[relay_df["_relay_group"].eq("2_pole")].copy()
    )
    four_pole_df = _sort_component_relay_group_df(
        relay_df.loc[relay_df["_relay_group"].eq("4_pole")].copy()
    )
    timed_df = _sort_component_timed_relay_group_df(
        relay_df.loc[relay_df["_relay_group"].eq("timed")].copy()
    )
    one_pole_df = _sort_component_relay_group_df(
        relay_df.loc[relay_df["_relay_group"].eq("1_pole")].copy()
    )

    strip_stats["2pole_rows"] = len(two_pole_df)
    strip_stats["4pole_rows"] = len(four_pole_df)
    strip_stats["timed_rows"] = len(timed_df)
    strip_stats["1pole_rows"] = len(one_pole_df)
    strip_stats["2pole_preview_names"] = _component_strip_relay_preview_names(two_pole_df)
    strip_stats["4pole_preview_names"] = _component_strip_relay_preview_names(four_pole_df)
    strip_stats["timed_preview_names"] = _component_strip_relay_preview_names(timed_df)
    strip_stats["1pole_preview_names"] = _component_strip_relay_preview_names(one_pole_df)

    relay_groups = [
        ("2_pole", two_pole_df),
        ("4_pole", four_pole_df),
        ("timed", timed_df),
        ("1_pole", one_pole_df),
    ]
    non_empty_relay_groups = [
        (group_label, group_df)
        for group_label, group_df in relay_groups
        if not group_df.empty
    ]
    relay_strip_rows: list[tuple[Any, Any]] = []
    if non_empty_relay_groups:
        relay_strip_rows.append((_RELAY_STRIP_START_STOP_SPACE, _RELAY_STRIP_START_TEXT))
        strip_stats["start_rows"] = 1

    for group_index, (group_label, group_df) in enumerate(non_empty_relay_groups):
        for row in group_df.to_dict("records"):
            row_name = _stringify_cell(row.get("Name"))
            row_type = row.get("TYPE")
            row_space = _detect_component_relay_strip_space(row_type, group_label)
            relay_strip_rows.append((row_space, row_name))

            if row_space == _RELAY_STRIP_2POLE_WIDTH:
                strip_stats["width_15_8_rows"] += 1
            elif row_space == _RELAY_STRIP_RE22_WIDTH:
                strip_stats["width_22_5_rows"] += 1
            elif row_space == _RELAY_STRIP_4POLE_WIDTH:
                strip_stats["width_27_rows"] += 1
            elif row_space == _RELAY_STRIP_1POLE_WIDTH:
                strip_stats["width_6_2_rows"] += 1

        if group_index < len(non_empty_relay_groups) - 1:
            relay_strip_rows.append((_RELAY_STRIP_CLIPFIX_SPACE, ""))
            strip_stats["clipfix_rows"] += 1

    if non_empty_relay_groups:
        relay_strip_rows.append((_RELAY_STRIP_START_STOP_SPACE, _RELAY_STRIP_STOP_TEXT))
        strip_stats["stop_rows"] = 1

    return relay_strip_rows, strip_stats


def _build_component_strip_side_df(strip_rows: list[tuple[Any, Any]]) -> pd.DataFrame:
    """Convert strip rows into one side of the exported Component Strip layout."""
    return pd.DataFrame(strip_rows, columns=_COMPONENT_STRIP_SIDE_COLUMNS)


def _append_component_strip_stop_row(strip_rows: list[tuple[Any, Any]]) -> list[tuple[Any, Any]]:
    """Append STOP immediately after the last real row for one strip side."""
    if not strip_rows:
        return []
    return [
        *list(strip_rows),
        (_RELAY_STRIP_START_STOP_SPACE, _RELAY_STRIP_STOP_TEXT),
    ]


def _build_component_strip_layout(
    fuse_strip_rows: list[tuple[Any, Any]],
    relay_strip_rows: list[tuple[Any, Any]],
) -> dict[str, Any]:
    """Build a structured side-by-side Component Strip layout for the shared exporter."""
    wrapped_fuse_strip_rows = _append_component_strip_stop_row(fuse_strip_rows)
    return {
        "layout": "component_strip",
        "fuse_strip_df": _build_component_strip_side_df(wrapped_fuse_strip_rows),
        "relay_strip_df": _build_component_strip_side_df(relay_strip_rows),
    }


def _build_component_strip_df(component_marking_sheet_df: pd.DataFrame) -> tuple[dict[str, Any], dict[str, Any]]:
    """Build the Component Strip layout with fuse strip on the left and relay strip on the right."""
    empty_strip_layout = _build_component_strip_layout([], [])
    strip_stats: dict[str, Any] = {
        "24vdc_rows": 0,
        "230vac_rows": 0,
        "f92_wide_name": "",
        "relay_2pole_rows": 0,
        "relay_4pole_rows": 0,
        "relay_timed_rows": 0,
        "relay_1pole_rows": 0,
        "relay_duplicate_rows_removed": 0,
        "relay_clipfix_rows": 0,
        "relay_2pole_preview_names": [],
        "relay_4pole_preview_names": [],
        "relay_timed_preview_names": [],
        "relay_1pole_preview_names": [],
        "relay_start_rows": 0,
        "relay_stop_rows": 0,
        "relay_width_15_8_rows": 0,
        "relay_width_22_5_rows": 0,
        "relay_width_27_rows": 0,
        "relay_width_6_2_rows": 0,
        "layout_rows": 0,
    }
    if component_marking_sheet_df.empty:
        return empty_strip_layout, strip_stats

    working_df = component_marking_sheet_df.copy().reset_index(drop=True)
    for column_name in ("Name", "TYPE", "Group", "Category"):
        if column_name not in working_df.columns:
            working_df[column_name] = ""
    working_df["_original_order"] = range(len(working_df))

    fuse_strip_rows, fuse_stats = _build_component_fuse_strip_rows(working_df)
    relay_strip_rows, relay_stats = _build_component_relay_strip_rows(working_df)
    strip_stats.update(fuse_stats)
    strip_stats["relay_2pole_rows"] = relay_stats["2pole_rows"]
    strip_stats["relay_4pole_rows"] = relay_stats["4pole_rows"]
    strip_stats["relay_timed_rows"] = relay_stats["timed_rows"]
    strip_stats["relay_1pole_rows"] = relay_stats["1pole_rows"]
    strip_stats["relay_duplicate_rows_removed"] = relay_stats["duplicate_rows_removed"]
    strip_stats["relay_clipfix_rows"] = relay_stats["clipfix_rows"]
    strip_stats["relay_2pole_preview_names"] = relay_stats["2pole_preview_names"]
    strip_stats["relay_4pole_preview_names"] = relay_stats["4pole_preview_names"]
    strip_stats["relay_timed_preview_names"] = relay_stats["timed_preview_names"]
    strip_stats["relay_1pole_preview_names"] = relay_stats["1pole_preview_names"]
    strip_stats["relay_start_rows"] = relay_stats["start_rows"]
    strip_stats["relay_stop_rows"] = relay_stats["stop_rows"]
    strip_stats["relay_width_15_8_rows"] = relay_stats["width_15_8_rows"]
    strip_stats["relay_width_22_5_rows"] = relay_stats["width_22_5_rows"]
    strip_stats["relay_width_27_rows"] = relay_stats["width_27_rows"]
    strip_stats["relay_width_6_2_rows"] = relay_stats["width_6_2_rows"]
    strip_layout = _build_component_strip_layout(fuse_strip_rows, relay_strip_rows)
    strip_stats["layout_rows"] = max(
        len(strip_layout["fuse_strip_df"]),
        len(strip_layout["relay_strip_df"]),
    )

    return strip_layout, strip_stats


def _build_component_final_marking_sheet(
    component_strip_layout: dict[str, Any],
    component_cm_df: pd.DataFrame,
) -> pd.DataFrame:
    """Build one final main-workbook component sheet with Strip and CM blocks combined."""
    return _ComponentFinalMarkingSheetDataFrame(
        pd.DataFrame(),
        component_strip_layout={
            "layout": "component_strip",
            "fuse_strip_df": component_strip_layout.get(
                "fuse_strip_df",
                pd.DataFrame(columns=_COMPONENT_STRIP_SIDE_COLUMNS),
            ).copy().reset_index(drop=True),
            "relay_strip_df": component_strip_layout.get(
                "relay_strip_df",
                pd.DataFrame(columns=_COMPONENT_STRIP_SIDE_COLUMNS),
            ).copy().reset_index(drop=True),
        },
        component_cm_df=component_cm_df.copy().reset_index(drop=True),
    )


def _build_component_production_df(
    component_marking_df: pd.DataFrame,
) -> tuple[pd.DataFrame, dict[str, int]]:
    """Build the production-check sheet rows without merging relay TYPE values."""
    if component_marking_df.empty:
        empty_df = pd.DataFrame(
            columns=[*_PRODUCTION_COLUMNS, "_is_section", "_is_separator", _PRODUCTION_TECHNICAL_FLAG_COLUMN]
        )
        return empty_df, {"relay_rows": 0, "fuse_rows": 0, "button_rows": 0}

    ordered_sections, group_counts = _build_component_production_ordered_sections(component_marking_df)
    ordered_records: list[dict[str, Any]] = []
    for section_label, section_df in ordered_sections:
        section_records = _component_rows_to_production_records(section_df)
        if not section_records:
            continue
        ordered_records.append(_build_production_section_row(section_label))
        ordered_records.extend(section_records)
        ordered_records.append(_build_production_separator_row())

    production_df = pd.DataFrame(
        ordered_records,
        columns=[*_PRODUCTION_COLUMNS, "_is_section", "_is_separator", _PRODUCTION_TECHNICAL_FLAG_COLUMN],
    )
    return production_df.reset_index(drop=True), group_counts


def _build_component_production_filename(file_name: str) -> str:
    """Build a stable filename for the separate production workbook."""
    base_name = Path(file_name or "component_marking").stem or "component_marking"
    return f"{base_name}_production_check.xlsx"


def _component_cabinet_sort_key(cabinet_id: Any) -> tuple[str, int, str]:
    """Sort cabinet ids naturally so A2 comes before A10."""
    text = _stringify_cell(cabinet_id).upper()
    match = re.match(r"^(?P<prefix>[A-Z]+)(?P<number>\d+)(?P<suffix>.*)$", text)
    if not match:
        return (text, -1, "")
    return (
        _stringify_cell(match.group("prefix")),
        int(match.group("number")),
        _stringify_cell(match.group("suffix")),
    )


def _write_component_production_sheet(
    worksheet: Any,
    production_df: pd.DataFrame,
    *,
    columns: list[str],
    column_widths: dict[str, int],
    column_indexes: dict[str, int],
    marked_col_index: int,
    technical_flag_col_index: int,
    xl_col_to_name: Any,
    header_format: Any,
    text_format: Any,
    quantity_format: Any,
    marked_format: Any,
    green_row_format: Any,
    red_row_format: Any,
    section_format: Any,
    technical_flag_format: Any,
) -> int:
    """Write one production-style sheet and return its last data row index."""
    for col_index, column_name in enumerate(columns):
        worksheet.write(0, col_index, column_name, header_format)
        worksheet.set_column(col_index, col_index, column_widths[column_name])
    worksheet.write_comment(0, marked_col_index, "1 = Marked\n0 = Missing")
    worksheet.write(0, technical_flag_col_index, _PRODUCTION_TECHNICAL_FLAG_COLUMN, header_format)
    worksheet.set_column(technical_flag_col_index, technical_flag_col_index, None, None, {"hidden": True})
    worksheet.freeze_panes(1, 0)

    for row_offset, row_data in enumerate(production_df.to_dict("records"), start=1):
        is_section_row = bool(row_data.get("_is_section"))
        is_separator_row = bool(row_data.get("_is_separator"))
        worksheet.set_row(row_offset, 12 if is_separator_row else (24 if is_section_row else 20))

        include_flag = int(row_data.get(_PRODUCTION_TECHNICAL_FLAG_COLUMN, 0) or 0)
        worksheet.write_number(row_offset, technical_flag_col_index, include_flag, technical_flag_format)

        if is_section_row:
            worksheet.merge_range(
                row_offset,
                0,
                row_offset,
                len(columns) - 1,
                _stringify_cell(row_data.get("Name")),
                section_format,
            )
            continue
        if is_separator_row:
            continue

        for text_column in ("Name", "Article No.", "TYPE", "Description", "Comments"):
            value = _stringify_cell(row_data.get(text_column))
            column_index = column_indexes[text_column]
            if value:
                worksheet.write(row_offset, column_index, value, text_format)
            else:
                worksheet.write_blank(row_offset, column_index, None, text_format)

        quantity_value = row_data.get("Quantity")
        quantity_column_index = column_indexes["Quantity"]
        numeric_quantity = _coerce_excel_number(quantity_value)
        if numeric_quantity is not None:
            worksheet.write_number(row_offset, quantity_column_index, numeric_quantity, quantity_format)
        elif _stringify_cell(quantity_value):
            worksheet.write(row_offset, quantity_column_index, _stringify_cell(quantity_value), quantity_format)
        else:
            worksheet.write_blank(row_offset, quantity_column_index, None, quantity_format)

        marked_value = _stringify_cell(row_data.get("Marked"))
        if marked_value:
            numeric_marked = _coerce_excel_number(marked_value)
            if numeric_marked is not None:
                worksheet.write_number(row_offset, marked_col_index, numeric_marked, marked_format)
            else:
                worksheet.write(row_offset, marked_col_index, marked_value, marked_format)
        else:
            worksheet.write_blank(row_offset, marked_col_index, None, marked_format)

    last_data_row = len(production_df)
    if last_data_row >= 1:
        marked_col_letter = xl_col_to_name(marked_col_index)
        technical_flag_col_letter = xl_col_to_name(technical_flag_col_index)
        worksheet.conditional_format(
            1,
            0,
            last_data_row,
            len(columns) - 1,
            {
                "type": "formula",
                "criteria": f'=AND(${technical_flag_col_letter}2=1,${marked_col_letter}2<>"",${marked_col_letter}2=1)',
                "format": green_row_format,
            },
        )
        worksheet.conditional_format(
            1,
            0,
            last_data_row,
            len(columns) - 1,
            {
                "type": "formula",
                "criteria": f'=AND(${technical_flag_col_letter}2=1,${marked_col_letter}2<>"",${marked_col_letter}2=0)',
                "format": red_row_format,
            },
        )
    return last_data_row


def _list_component_calculation_types(production_df: pd.DataFrame) -> list[str]:
    """Return the stable ordered TYPE list used by the Calculation sheet."""
    include_mask = pd.to_numeric(
        production_df.get(
            _PRODUCTION_TECHNICAL_FLAG_COLUMN,
            pd.Series(0, index=production_df.index),
        ),
        errors="coerce",
    ).fillna(0).astype(int) == 1
    actual_rows_df = production_df.loc[include_mask].copy()
    type_values = [
        type_value
        for type_value in actual_rows_df.get("TYPE", pd.Series(dtype=object)).map(_stringify_cell).tolist()
        if type_value != ""
    ]
    return list(dict.fromkeys(type_values))


def _count_component_production_actual_rows(production_df: pd.DataFrame) -> int:
    """Count actual exported component rows, excluding section/separator helper rows."""
    include_values = pd.to_numeric(
        production_df.get(
            _PRODUCTION_TECHNICAL_FLAG_COLUMN,
            pd.Series(0, index=production_df.index),
        ),
        errors="coerce",
    ).fillna(0)
    return int((include_values.astype(int) == 1).sum())


def _write_component_calculation_block(
    calculation_sheet: Any,
    *,
    start_row: int,
    start_col: int,
    title: str | None,
    source_sheet_name: str,
    source_production_df: pd.DataFrame,
    calculation_columns: list[str],
    calculation_widths: dict[str, int],
    production_column_indexes: dict[str, int],
    technical_flag_col_index: int,
    xl_col_to_name: Any,
    xl_rowcol_to_cell: Any,
    header_format: Any,
    section_format: Any,
    calculation_text_format: Any,
    calculation_number_format: Any,
    marked_header_format: Any,
    missing_header_format: Any,
    marked_calculation_number_format: Any,
    missing_calculation_number_format: Any,
) -> None:
    """Write one calculation summary block that references a single production sheet."""
    header_row = start_row
    if title:
        calculation_sheet.merge_range(
            start_row,
            start_col,
            start_row,
            start_col + len(calculation_columns) - 1,
            title,
            section_format,
        )
        header_row = start_row + 1

    for col_offset, column_name in enumerate(calculation_columns):
        header_cell_format = header_format
        if column_name == "Marked Quantity":
            header_cell_format = marked_header_format
        elif column_name == "Missing Quantity":
            header_cell_format = missing_header_format
        calculation_sheet.write(header_row, start_col + col_offset, column_name, header_cell_format)
        calculation_sheet.set_column(
            start_col + col_offset,
            start_col + col_offset,
            calculation_widths[column_name],
        )

    calculation_types = _list_component_calculation_types(source_production_df)
    if not calculation_types:
        return

    last_excel_row = len(source_production_df) + 1
    sheet_reference = "'" + source_sheet_name.replace("'", "''") + "'"
    type_col_letter = xl_col_to_name(production_column_indexes["TYPE"])
    quantity_col_letter = xl_col_to_name(production_column_indexes["Quantity"])
    marked_col_letter = xl_col_to_name(production_column_indexes["Marked"])
    type_range = f"{sheet_reference}!${type_col_letter}$2:${type_col_letter}${last_excel_row}"
    quantity_range = f"{sheet_reference}!${quantity_col_letter}$2:${quantity_col_letter}${last_excel_row}"
    marked_range = f"{sheet_reference}!${marked_col_letter}$2:${marked_col_letter}${last_excel_row}"
    include_range = (
        f"{sheet_reference}!${xl_col_to_name(technical_flag_col_index)}$2:"
        f"${xl_col_to_name(technical_flag_col_index)}${last_excel_row}"
    )

    for row_offset, type_value in enumerate(calculation_types, start=1):
        row_index = header_row + row_offset
        calculation_sheet.write(row_index, start_col, type_value, calculation_text_format)

        type_cell = xl_rowcol_to_cell(row_index, start_col, col_abs=True)
        total_cell = xl_rowcol_to_cell(row_index, start_col + 1)
        marked_cell = xl_rowcol_to_cell(row_index, start_col + 3)
        missing_cell = xl_rowcol_to_cell(row_index, start_col + 4)

        calculation_sheet.write_formula(
            row_index,
            start_col + 1,
            f"=SUMIFS({quantity_range},{type_range},{type_cell},{include_range},1)",
            calculation_number_format,
        )
        calculation_sheet.write_formula(
            row_index,
            start_col + 2,
            f"={total_cell}-{marked_cell}-{missing_cell}",
            calculation_number_format,
        )
        calculation_sheet.write_formula(
            row_index,
            start_col + 3,
            f"=SUMIFS({quantity_range},{type_range},{type_cell},{include_range},1,{marked_range},1)",
            marked_calculation_number_format,
        )
        calculation_sheet.write_formula(
            row_index,
            start_col + 4,
            f"=SUMIFS({quantity_range},{type_range},{type_cell},{include_range},1,{marked_range},0)",
            missing_calculation_number_format,
        )


def _export_component_production_workbook(
    production_df: pd.DataFrame,
    cabinet_production_dfs: dict[str, pd.DataFrame] | None = None,
) -> bytes:
    """Export a separate production workbook with manual 1/0 marking cells."""
    try:
        import xlsxwriter
        from xlsxwriter.utility import xl_col_to_name, xl_rowcol_to_cell
    except ModuleNotFoundError as exc:
        raise RuntimeError("xlsxwriter is required for component production workbook export") from exc

    output = BytesIO()
    workbook = xlsxwriter.Workbook(output, {"in_memory": True})
    cabinet_production_dfs = cabinet_production_dfs or {}
    sorted_cabinet_ids = sorted(cabinet_production_dfs, key=_component_cabinet_sort_key)
    cabinet_mode = bool(sorted_cabinet_ids)
    cabinet_sheet_names = {
        cabinet_id: _sanitize_component_cabinet_sheet_name(cabinet_id) or "CABINET"
        for cabinet_id in sorted_cabinet_ids
    }
    production_sheet = None
    cabinet_sheets: dict[str, Any] = {}
    if cabinet_mode:
        cabinet_sheets = {
            cabinet_id: workbook.add_worksheet(cabinet_sheet_names[cabinet_id])
            for cabinet_id in sorted_cabinet_ids
        }
    else:
        production_sheet = workbook.add_worksheet("Production check")
    calculation_sheet = workbook.add_worksheet("Calculation")

    header_format = workbook.add_format(
        {
            "bold": True,
            "border": 1,
            "align": "center",
            "valign": "vcenter",
        }
    )
    text_format = workbook.add_format({"border": 1, "valign": "vcenter"})
    quantity_format = workbook.add_format(
        {"border": 1, "align": "center", "valign": "vcenter", "num_format": "0.############"}
    )
    marked_format = workbook.add_format({"border": 1, "align": "center", "valign": "vcenter"})
    green_row_format = workbook.add_format({"bg_color": "#C6EFCE", "border": 1})
    red_row_format = workbook.add_format({"bg_color": "#F4CCCC", "border": 1})
    section_format = workbook.add_format(
        {"bold": True, "font_size": 13, "border": 1, "align": "left", "valign": "vcenter"}
    )
    calculation_text_format = workbook.add_format({"border": 1, "valign": "vcenter"})
    calculation_number_format = workbook.add_format(
        {"border": 1, "align": "center", "valign": "vcenter", "num_format": "0.############"}
    )
    marked_header_format = workbook.add_format(
        {"bold": True, "border": 1, "align": "center", "valign": "vcenter", "bg_color": "#C6EFCE"}
    )
    missing_header_format = workbook.add_format(
        {"bold": True, "border": 1, "align": "center", "valign": "vcenter", "bg_color": "#F4CCCC"}
    )
    marked_calculation_number_format = workbook.add_format(
        {"border": 1, "align": "center", "valign": "vcenter", "num_format": "0.############", "bg_color": "#C6EFCE"}
    )
    missing_calculation_number_format = workbook.add_format(
        {"border": 1, "align": "center", "valign": "vcenter", "num_format": "0.############", "bg_color": "#F4CCCC"}
    )
    technical_flag_format = workbook.add_format({"num_format": "0"})

    columns = list(_PRODUCTION_COLUMNS)
    column_widths = {
        "Name": 13.5,
        "Article No.": 20,
        "TYPE": 24,
        "Quantity": 12,
        "Marked": 10,
        "Description": 85,
        "Comments": 28,
    }
    column_indexes = {column_name: column_index for column_index, column_name in enumerate(columns)}
    marked_col_index = column_indexes["Marked"]
    technical_flag_col_index = len(columns)
    if cabinet_mode:
        for cabinet_id in sorted_cabinet_ids:
            _write_component_production_sheet(
                cabinet_sheets[cabinet_id],
                cabinet_production_dfs[cabinet_id],
                columns=columns,
                column_widths=column_widths,
                column_indexes=column_indexes,
                marked_col_index=marked_col_index,
                technical_flag_col_index=technical_flag_col_index,
                xl_col_to_name=xl_col_to_name,
                header_format=header_format,
                text_format=text_format,
                quantity_format=quantity_format,
                marked_format=marked_format,
                green_row_format=green_row_format,
                red_row_format=red_row_format,
                section_format=section_format,
                technical_flag_format=technical_flag_format,
            )
    else:
        _write_component_production_sheet(
            production_sheet,
            production_df,
            columns=columns,
            column_widths=column_widths,
            column_indexes=column_indexes,
            marked_col_index=marked_col_index,
            technical_flag_col_index=technical_flag_col_index,
            xl_col_to_name=xl_col_to_name,
            header_format=header_format,
            text_format=text_format,
            quantity_format=quantity_format,
            marked_format=marked_format,
            green_row_format=green_row_format,
            red_row_format=red_row_format,
            section_format=section_format,
            technical_flag_format=technical_flag_format,
        )

    calculation_columns = [
        "Type numeriai",
        "Total quantity",
        "Not Marked Quantity",
        "Marked Quantity",
        "Missing Quantity",
    ]
    calculation_widths = {
        "Type numeriai": 24,
        "Total quantity": 16,
        "Not Marked Quantity": 20,
        "Marked Quantity": 18,
        "Missing Quantity": 18,
    }
    calculation_sheet.freeze_panes(2 if cabinet_mode else 1, 0)
    if cabinet_mode:
        cabinet_block_width = len(calculation_columns) + 2
        for cabinet_index, cabinet_id in enumerate(sorted_cabinet_ids):
            _write_component_calculation_block(
                calculation_sheet,
                start_row=0,
                start_col=cabinet_index * cabinet_block_width,
                title=cabinet_id,
                source_sheet_name=cabinet_sheet_names[cabinet_id],
                source_production_df=cabinet_production_dfs[cabinet_id],
                calculation_columns=calculation_columns,
                calculation_widths=calculation_widths,
                production_column_indexes=column_indexes,
                technical_flag_col_index=technical_flag_col_index,
                xl_col_to_name=xl_col_to_name,
                xl_rowcol_to_cell=xl_rowcol_to_cell,
                header_format=header_format,
                section_format=section_format,
                calculation_text_format=calculation_text_format,
                calculation_number_format=calculation_number_format,
                marked_header_format=marked_header_format,
                missing_header_format=missing_header_format,
                marked_calculation_number_format=marked_calculation_number_format,
                missing_calculation_number_format=missing_calculation_number_format,
            )
    else:
        _write_component_calculation_block(
            calculation_sheet,
            start_row=0,
            start_col=0,
            title=None,
            source_sheet_name="Production check",
            source_production_df=production_df,
            calculation_columns=calculation_columns,
            calculation_widths=calculation_widths,
            production_column_indexes=column_indexes,
            technical_flag_col_index=technical_flag_col_index,
            xl_col_to_name=xl_col_to_name,
            xl_rowcol_to_cell=xl_rowcol_to_cell,
            header_format=header_format,
            section_format=section_format,
            calculation_text_format=calculation_text_format,
            calculation_number_format=calculation_number_format,
            marked_header_format=marked_header_format,
            missing_header_format=missing_header_format,
            marked_calculation_number_format=marked_calculation_number_format,
            missing_calculation_number_format=missing_calculation_number_format,
        )

    workbook.close()
    output.seek(0)
    return output.getvalue()


def process_component_result(file_bytes: bytes, file_name: str) -> dict[str, Any]:
    """Parse the component workbook and split rows into Component Marking and Unused."""
    component_df, _, developer_debug_messages = _load_component_input(file_bytes)

    if "Name" in component_df.columns:
        unused_mask = component_df["Name"].map(_is_unused_component_name)
        cabinet_prefixed_mask = component_df["Name"].map(
            lambda value: _stringify_cell(value).startswith("+A")
        )
    else:
        unused_mask = pd.Series(True, index=component_df.index)
        cabinet_prefixed_mask = pd.Series(False, index=component_df.index)

    unused_df = component_df.loc[unused_mask].reset_index(drop=True)
    component_marking_df = component_df.loc[~unused_mask].reset_index(drop=True)

    component_marking_df["Category"] = component_marking_df.apply(
        lambda row: _classify_component_category(row.get("Name"), row.get("TYPE")),
        axis=1,
    )
    cabinet_map, cabinet_stats = _build_component_cabinet_map(component_marking_df)
    sorted_cabinet_ids = sorted(cabinet_map, key=_component_cabinet_sort_key)
    has_detected_cabinet_ids = bool(sorted_cabinet_ids)
    use_single_cabinet_local_dataset = (not has_detected_cabinet_ids) and not component_marking_df.empty
    cabinet_production_dfs: dict[str, pd.DataFrame] = {}
    cabinet_source_row_counts: dict[str, int] = {}
    cabinet_production_row_counts: dict[str, int] = {}
    for cabinet_id in sorted_cabinet_ids:
        cabinet_production_source_df = cabinet_map[cabinet_id].copy().reset_index(drop=True)
        cabinet_source_row_counts[cabinet_id] = len(cabinet_production_source_df)
        cabinet_production_source_df["Category"] = cabinet_production_source_df.apply(
            lambda row: _classify_component_category(row.get("Name"), row.get("TYPE")),
            axis=1,
        )
        cabinet_production_df, _ = _build_component_production_df(cabinet_production_source_df)
        cabinet_production_df = cabinet_production_df.reset_index(drop=True).copy()
        cabinet_production_row_counts[cabinet_id] = _count_component_production_actual_rows(cabinet_production_df)
        if cabinet_source_row_counts[cabinet_id] > 0 and cabinet_production_row_counts[cabinet_id] == 0:
            raise ValueError(
                "component production workbook: cabinet "
                f"{cabinet_id} produced no export rows from {cabinet_source_row_counts[cabinet_id]} source rows"
            )
        cabinet_production_dfs[cabinet_id] = cabinet_production_df

    production_df, grouped_row_counts = _build_component_production_df(component_marking_df)
    component_marking_sheet_df = _build_component_marking_sheet_df(component_marking_df)
    component_strip_sheet, component_strip_stats = _build_component_strip_df(component_marking_sheet_df)
    unused_export_df = _drop_production_only_component_columns(unused_df)
    production_workbook_bytes = _export_component_production_workbook(production_df, cabinet_production_dfs)
    category_counts = component_marking_df["Category"].value_counts(dropna=False)
    group_counts = component_marking_sheet_df.get("Group", pd.Series(dtype=object)).value_counts(dropna=False)

    developer_debug_messages.append(
        f"component parser: kept +A* cabinet-prefixed rows in main Component flow -> {int((cabinet_prefixed_mask & ~unused_mask).sum())}"
    )
    developer_debug_messages.append(
        f"component parser: +A* cabinet-prefixed rows still moved to Unused by normal rules -> {int((cabinet_prefixed_mask & unused_mask).sum())}"
    )
    developer_debug_messages.append(
        "component parser: raw cabinet ids detected -> "
        + (", ".join(cabinet_stats["raw_cabinet_ids"]) if cabinet_stats["raw_cabinet_ids"] else "none")
    )
    developer_debug_messages.append(
        "component parser: sanitized cabinet ids used for sheets -> "
        + (
            ", ".join(cabinet_stats["sanitized_cabinet_ids"])
            if cabinet_stats["sanitized_cabinet_ids"]
            else "none"
        )
    )
    if len(sorted_cabinet_ids) > 1:
        developer_debug_messages.append(
            "component cabinet detection: multi cabinet ids -> " + ", ".join(sorted_cabinet_ids)
        )
    elif len(sorted_cabinet_ids) == 1:
        developer_debug_messages.append(
            "component cabinet detection: single detected cabinet -> " + sorted_cabinet_ids[0]
        )
    elif use_single_cabinet_local_dataset:
        developer_debug_messages.append(
            "component cabinet detection: no +A* found, using single-cabinet local dataset"
        )
    for sheet_name_change in cabinet_stats["sheet_name_sanitizations"]:
        developer_debug_messages.append(
            f"component parser: cabinet sheet id sanitized -> {sheet_name_change}"
        )
    developer_debug_messages.append(
        "component parser: detected cabinets -> "
        + (", ".join(cabinet_stats["cabinet_ids"]) if cabinet_stats["cabinet_ids"] else "none")
    )
    for cabinet_id in cabinet_stats["cabinet_ids"]:
        developer_debug_messages.append(
            f"component parser: cabinet {cabinet_id} rows -> {cabinet_stats['row_counts'][cabinet_id]}"
        )
        cabinet_sample_names = ", ".join(
            [
                _stringify_cell(name_value)
                for name_value in cabinet_map[cabinet_id].get("Name", pd.Series(dtype=object)).head(3).tolist()
                if _stringify_cell(name_value)
            ]
        )
        developer_debug_messages.append(
            f"component parser: cabinet {cabinet_id} sample names -> {cabinet_sample_names or 'none'}"
        )
    developer_debug_messages.append(f"component parser: moved {len(unused_df)} rows to Unused")
    developer_debug_messages.append(f"component parser: FUSE rows -> {int(category_counts.get('FUSE', 0))}")
    developer_debug_messages.append(f"component parser: RELAY_1P rows -> {int(category_counts.get('RELAY_1P', 0))}")
    developer_debug_messages.append(f"component parser: RELAY_4P rows -> {int(category_counts.get('RELAY_4P', 0))}")
    developer_debug_messages.append(f"component parser: RELAY_2P rows -> {int(category_counts.get('RELAY_2P', 0))}")
    developer_debug_messages.append(f"component parser: BUTTON rows -> {int(category_counts.get('BUTTON', 0))}")
    developer_debug_messages.append(f"component parser: OTHER rows -> {int(category_counts.get('OTHER', 0))}")
    developer_debug_messages.append(f"component parser: final component rows -> {len(component_marking_df)}")
    developer_debug_messages.append(f"component parser: final unused rows -> {len(unused_df)}")
    developer_debug_messages.append("relay TYPE merge removed from component production workbook")
    developer_debug_messages.append(f"grouped relay rows count: {grouped_row_counts['relay_rows']}")
    developer_debug_messages.append(f"grouped fuse rows count: {grouped_row_counts['fuse_rows']}")
    developer_debug_messages.append(f"grouped button rows count: {grouped_row_counts['button_rows']}")
    developer_debug_messages.append("Component Marking sheet kept flat with original row order")
    developer_debug_messages.append("Component Marking sheet uses Group classification column")
    if use_single_cabinet_local_dataset:
        developer_debug_messages.append(
            f"component single cabinet: CM source rows -> {len(component_marking_df)}"
        )
        developer_debug_messages.append(
            f"component single cabinet: production source rows -> {len(component_marking_df)}"
        )
    developer_debug_messages.append(
        "Component Marking group counts -> "
        f"FUSES={int(group_counts.get('FUSES', 0))}, "
        f"RELAY_1P={int(group_counts.get('RELAY_1P', 0))}, "
        f"RELAY_2P={int(group_counts.get('RELAY_2P', 0))}, "
        f"RELAY_4P={int(group_counts.get('RELAY_4P', 0))}, "
        f"BUTTONS={int(group_counts.get('BUTTONS', 0))}, "
        f"OTHER={int(group_counts.get('OTHER', 0))}"
    )
    developer_debug_messages.append("Buttons grouping applied to component production workbook")
    developer_debug_messages.append("component production: Article No. column enabled")
    developer_debug_messages.append("production workbook header note added to Marked")
    developer_debug_messages.append("calculation sheet created")
    if sorted_cabinet_ids:
        developer_debug_messages.append(
            "component production workbook: cabinet mode active -> "
            + ", ".join(sorted_cabinet_ids)
        )
        developer_debug_messages.append(
            "component production workbook: global Production check skipped because cabinet sheets exist"
        )
        developer_debug_messages.append(
            "component calculation sheet: cabinet summaries added -> "
            + ", ".join(sorted_cabinet_ids)
        )
        for cabinet_id in sorted_cabinet_ids:
            developer_debug_messages.append(
                f"component production workbook: cabinet {cabinet_id} source rows -> {cabinet_source_row_counts[cabinet_id]}"
            )
            developer_debug_messages.append(
                f"component production workbook: cabinet {cabinet_id} production rows -> {cabinet_production_row_counts[cabinet_id]}"
            )
    else:
        developer_debug_messages.append(
            "component production workbook: no cabinets detected, using Production check fallback"
        )
    developer_debug_messages.append(
        "component strip voltage split -> "
        f"24VDC={component_strip_stats['24vdc_rows']}, 230VAC={component_strip_stats['230vac_rows']}"
    )
    developer_debug_messages.append(
        "component strip last -F92* width 8.3 -> "
        + (
            f"applied to `{component_strip_stats['f92_wide_name']}`"
            if component_strip_stats["f92_wide_name"]
            else "not applied"
        )
    )
    developer_debug_messages.append("component strip: fuse rows sorted by numeric Name key")
    developer_debug_messages.append("component strip: fuse A-suffix sort applied after normal fuse numbers")
    developer_debug_messages.append("fuse strip: replaced 230VAC label with 13.45 spacing")
    developer_debug_messages.append(
        "component strip supported widths -> "
        f"fuse={_FUSE_STRIP_WIDTH}, "
        f"fuse_with_cover={_FUSE_STRIP_COVERED_WIDTH}, "
        f"separate_cover={_FUSE_STRIP_SEPARATE_COVER_WIDTH}, "
        f"relay_start_stop={_RELAY_STRIP_START_STOP_SPACE}, "
        f"clipfix={_RELAY_STRIP_CLIPFIX_SPACE}, "
        f"relay_2pole={_RELAY_STRIP_2POLE_WIDTH}, "
        f"relay_4pole={_RELAY_STRIP_4POLE_WIDTH}, "
        f"relay_timed_re22={_RELAY_STRIP_RE22_WIDTH}, "
        f"relay_1pole={_RELAY_STRIP_1POLE_WIDTH}, "
        f"relay_start={_RELAY_STRIP_START_TEXT}, "
        f"relay_stop={_RELAY_STRIP_STOP_TEXT}"
    )
    developer_debug_messages.append(
        "component strip relay groups -> "
        f"2_pole={component_strip_stats['relay_2pole_rows']}, "
        f"4_pole={component_strip_stats['relay_4pole_rows']}, "
        f"timed={component_strip_stats['relay_timed_rows']}, "
        f"1_pole={component_strip_stats['relay_1pole_rows']}"
    )
    developer_debug_messages.append(
        "component strip relay group order -> 2_pole -> 4_pole -> timed -> 1_pole"
    )
    developer_debug_messages.append(
        f"component strip duplicate relay rows removed -> {component_strip_stats['relay_duplicate_rows_removed']}"
    )
    developer_debug_messages.append(
        f"component strip relay clipfix separators -> {component_strip_stats['relay_clipfix_rows']}"
    )
    developer_debug_messages.append("component strip: START/STOP rows added")
    developer_debug_messages.append(
        "component strip timed ordering sourced from Component Correction processor K192*/K192A* logic"
    )
    developer_debug_messages.append(
        "component strip relay preview 2_pole -> "
        + (
            ", ".join(component_strip_stats["relay_2pole_preview_names"])
            if component_strip_stats["relay_2pole_preview_names"]
            else "none"
        )
    )
    developer_debug_messages.append(
        "component strip relay preview 4_pole -> "
        + (
            ", ".join(component_strip_stats["relay_4pole_preview_names"])
            if component_strip_stats["relay_4pole_preview_names"]
            else "none"
        )
    )
    developer_debug_messages.append(
        "component strip relay preview timed -> "
        + (
            ", ".join(component_strip_stats["relay_timed_preview_names"])
            if component_strip_stats["relay_timed_preview_names"]
            else "none"
        )
    )
    developer_debug_messages.append(
        "component strip relay preview 1_pole -> "
        + (
            ", ".join(component_strip_stats["relay_1pole_preview_names"])
            if component_strip_stats["relay_1pole_preview_names"]
            else "none"
        )
    )
    developer_debug_messages.append(
        "component strip relay width counts -> "
        f"15.8={component_strip_stats['relay_width_15_8_rows']}, "
        f"27={component_strip_stats['relay_width_27_rows']}, "
        f"22.5={component_strip_stats['relay_width_22_5_rows']}, "
        f"6.2={component_strip_stats['relay_width_6_2_rows']}"
    )
    developer_debug_messages.append(
        f"component strip rows exported -> {component_strip_stats['layout_rows']}"
    )
    cabinet_final_component_sheets: dict[str, Any] = {}
    multi_cabinet_cm_mode = len(sorted_cabinet_ids) > 1
    if multi_cabinet_cm_mode:
        developer_debug_messages.append(
            "component markings workbook: final cabinet sheets added -> "
            + ", ".join(sorted_cabinet_ids)
        )
        for cabinet_id in sorted_cabinet_ids:
            cabinet_component_df = cabinet_map[cabinet_id].copy().reset_index(drop=True)
            cabinet_final_sheet_name = _build_component_markings_workbook_sheet_name(
                cabinet_id, "Component Marking"
            )
            cabinet_marking_sheet_df = _build_component_marking_sheet_df(cabinet_component_df)
            cabinet_strip_sheet, cabinet_strip_sheet_stats = _build_component_strip_df(cabinet_marking_sheet_df)
            cabinet_cm_sheet_df = _build_component_cm_sheet_df(cabinet_component_df)
            cabinet_final_component_sheets[cabinet_final_sheet_name] = _build_component_final_marking_sheet(
                cabinet_strip_sheet,
                cabinet_cm_sheet_df,
            )

            developer_debug_messages.append(
                f"component markings workbook: added final combined sheet {cabinet_final_sheet_name}"
            )
            developer_debug_messages.append(
                f"component markings workbook: {cabinet_final_sheet_name} strip rows -> {cabinet_strip_sheet_stats['layout_rows']}"
            )
            developer_debug_messages.append(
                f"component workbook routing: removed standalone intermediate sheets for {cabinet_final_sheet_name}"
            )

    localized_component_source_df = _build_component_local_name_source_df(component_marking_df)
    localized_component_marking_sheet_df = _build_component_marking_sheet_df(localized_component_source_df)
    localized_component_strip_sheet, _ = _build_component_strip_df(localized_component_marking_sheet_df)
    component_cm_sheet = _build_component_cm_sheet_df(component_marking_df)
    single_final_component_sheet = _build_component_final_marking_sheet(
        localized_component_strip_sheet,
        component_cm_sheet,
    )
    routing_mode = _determine_component_routing_mode(
        sorted_cabinet_ids,
        use_single_cabinet_local_dataset,
    )

    developer_debug_messages.append("component production workbook created")
    developer_debug_messages.append(f"production rows exported: {len(production_df)}")
    developer_debug_messages.append("production workbook uses filtered Component Marking rows only")

    user_info_messages = [
        "component input processed successfully",
        f"component rows exported: {len(component_marking_df)}",
        f"unused component rows exported: {len(unused_df)}",
        "component marking sheets created",
        "component production workbook created",
        f"production rows exported: {len(production_df)}",
        "production workbook uses filtered Component Marking rows only",
    ]

    main_component_sheets: dict[str, Any]
    debug_component_sheets: dict[str, Any]
    final_debug_sheet_names: list[str]
    expected_final_component_marking_sheet_names: list[str] = []
    developer_debug_messages.append(f"component debug workbook: mode -> {routing_mode}")
    if routing_mode == "no_cabinet":
        main_component_sheets = {
            "Component Marking": component_marking_sheet_df,
            "Component Strip": component_strip_sheet,
            "Unused": unused_export_df,
            "CM": component_cm_sheet,
        }
        debug_component_sheets = {}
        final_debug_sheet_names = ["Developer Debug"]
        developer_debug_messages.append("component workbook routing: no_cabinet fallback active")
    elif routing_mode == "single_cabinet":
        main_component_sheets = {
            "Component Marking": single_final_component_sheet,
        }
        expected_final_component_marking_sheet_names = ["Component Marking"]
        final_main_sheet_names = list(main_component_sheets.keys())
        final_debug_sheet_names = [
            "General information",
            "Component Marking",
            "Unused",
            "Developer Debug",
        ]
        debug_component_sheets = {
            "General information": _build_component_general_information_sheet(
                routing_mode=routing_mode,
                detected_cabinet_ids=sorted_cabinet_ids,
                total_input_rows=len(component_df),
                component_marking_rows=len(component_marking_df),
                unused_rows=len(unused_df),
                cabinet_source_row_counts=cabinet_source_row_counts,
                final_main_sheet_names=final_main_sheet_names,
                final_debug_sheet_names=final_debug_sheet_names,
            ),
            "Component Marking": localized_component_marking_sheet_df,
            "Unused": unused_export_df,
        }
        if use_single_cabinet_local_dataset:
            developer_debug_messages.append(
                "component workbook routing: single_cabinet local dataset active"
            )
        else:
            developer_debug_messages.append(
                "component workbook routing: single detected cabinet routed through single_cabinet mode"
            )
        developer_debug_messages.append(
            "component workbook routing: single_cabinet main workbook uses one final combined Component Marking sheet"
        )
        developer_debug_messages.append(
            "component workbook routing: moved Component Marking to debug workbook in single_cabinet mode"
        )
        developer_debug_messages.append(
            "component workbook routing: moved Unused to debug workbook in single_cabinet mode"
        )
        developer_debug_messages.append(
            "component workbook routing: removed standalone Component Strip and CM sheets from single_cabinet main workbook"
        )
    else:
        main_component_sheets = {
            **cabinet_final_component_sheets,
        }
        expected_final_component_marking_sheet_names = [
            _build_component_markings_workbook_sheet_name(cabinet_id, "Component Marking")
            for cabinet_id in sorted_cabinet_ids
        ]
        final_main_sheet_names = list(main_component_sheets.keys())
        final_debug_sheet_names = [
            "General information",
            "Component Marking",
            "Component Marking Raw",
            "Unused",
            "Developer Debug",
        ]
        debug_component_sheets = {
            "General information": _build_component_general_information_sheet(
                routing_mode=routing_mode,
                detected_cabinet_ids=sorted_cabinet_ids,
                total_input_rows=len(component_df),
                component_marking_rows=len(component_marking_df),
                unused_rows=len(unused_df),
                cabinet_source_row_counts=cabinet_source_row_counts,
                final_main_sheet_names=final_main_sheet_names,
                final_debug_sheet_names=final_debug_sheet_names,
            ),
            "Component Marking": localized_component_marking_sheet_df,
            "Component Marking Raw": component_marking_sheet_df,
            "Unused": unused_export_df,
        }
        developer_debug_messages.append(
            f"component workbook routing: cabinet mode active -> {routing_mode}"
        )
        developer_debug_messages.append(
            "component workbook routing: moved global Component Marking to debug workbook in multi_cabinet mode"
        )
        developer_debug_messages.append(
            "component workbook routing: moved raw cabinet-aware Component Marking view to debug workbook"
        )
        developer_debug_messages.append(
            "component workbook routing: moved Unused to debug workbook in multi_cabinet mode"
        )
        developer_debug_messages.append(
            "component workbook routing: multi_cabinet main workbook uses one final combined Component Marking sheet per cabinet"
        )
        developer_debug_messages.append(
            "component workbook routing: removed standalone Component Strip and CM sheets from multi_cabinet main workbook"
        )

    if "General information" in debug_component_sheets:
        developer_debug_messages.append("component debug workbook: added General information sheet")
    if "Component Marking" in debug_component_sheets:
        developer_debug_messages.append("component debug workbook: added Component Marking sheet")
    if "Component Marking Raw" in debug_component_sheets:
        developer_debug_messages.append("component debug workbook: added Component Marking Raw sheet")
    if "Unused" in debug_component_sheets:
        developer_debug_messages.append("component debug workbook: added Unused sheet")

    component_sheets = main_component_sheets
    final_main_sheet_names = list(component_sheets.keys())
    developer_debug_messages.append(
        "component workbook routing: final main sheets -> "
        + (", ".join(final_main_sheet_names) if final_main_sheet_names else "none")
    )
    developer_debug_messages.append(
        "component workbook routing: final debug sheets -> "
        + (", ".join(final_debug_sheet_names) if final_debug_sheet_names else "none")
    )
    _validate_component_routing_sheet_sets(
        routing_mode,
        final_main_sheet_names,
        final_debug_sheet_names,
        expected_final_component_marking_sheet_names,
    )
    developer_debug_messages.append(
        f"component workbook routing: {routing_mode} final routing applied successfully"
    )
    developer_debug_messages.append("component debug workbook: added Developer Debug sheet")
    developer_debug_messages.append("component debug workbook: general debug workbook generated")
    developer_debug_messages.append("component debug UI: verbose debug moved from UI to workbook")
    developer_debug_messages.append(
        "component debug workbook sheets -> "
        + (", ".join(final_debug_sheet_names) if final_debug_sheet_names else "none")
    )
    ui_component_debug_messages = ["Detailed component debug is available in the downloaded debug workbook."]

    debug_component_sheets["Developer Debug"] = _build_component_debug_messages_sheet(
        developer_debug_messages
    )
    for sheet_name, sheet_df in component_sheets.items():
        _validate_component_export_sheet(sheet_name, sheet_df)
    for sheet_name, sheet_df in debug_component_sheets.items():
        _validate_component_export_sheet(sheet_name, sheet_df)
    component_debug_workbook = _build_component_debug_workbook(debug_component_sheets)

    return {
        "sheets": component_sheets,
        "cabinet_map": cabinet_map,
        "user_info_messages": user_info_messages,
        "developer_debug_messages": ui_component_debug_messages,
        "debug_workbook": component_debug_workbook,
        "production_workbook": production_workbook_bytes,
        "production_filename": _build_component_production_filename(file_name),
        "source_file": file_name or "uploaded_file",
    }
