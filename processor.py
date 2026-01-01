from __future__ import annotations

import re
from io import BytesIO
from typing import Any, Dict, List, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


# -----------------------------------------------------------------------------
# Helpers
# -----------------------------------------------------------------------------
def _drop_unnamed_cols(df: pd.DataFrame) -> pd.DataFrame:
    return df.loc[:, ~df.columns.astype(str).str.match(r"^Unnamed")].copy()


def _ensure_cols(df: pd.DataFrame, cols: List[str]) -> None:
    missing = [c for c in cols if c not in df.columns]
    if missing:
        raise ValueError(f"Trūksta privalomų stulpelių: {', '.join(missing)}")


def _append_removed(parts: List[pd.DataFrame], df_part: pd.DataFrame, reason: str) -> None:
    if df_part.empty:
        return
    tmp = df_part.copy()
    tmp["Removed Reason"] = reason
    parts.append(tmp)


def _to_num(v: Any) -> float | None:
    x = pd.to_numeric(v, errors="coerce")
    return None if pd.isna(x) else float(x)


def _gs_int(v: Any) -> int | None:
    n = _to_num(v)
    return None if n is None else int(n)


def _gs_str(v: Any) -> str:
    n = _gs_int(v)
    return "" if n is None else str(n)


def _starts_with_any(s: pd.Series, prefixes: Tuple[str, ...]) -> pd.Series:
    return s.astype(str).str.startswith(prefixes, na=False)


def _next_free_gs(existing: set[int], preferred: int) -> int:
    g = preferred
    while g in existing:
        g += 1
    existing.add(g)
    return g


def _add_gs_prefix(df: pd.DataFrame) -> pd.DataFrame:
    """Prefix Name with +GS where missing."""
    gss = df["Group Sorting"].apply(_gs_str)
    has_gs = gss.ne("")
    not_prefixed = ~df["Name"].astype(str).str.startswith("+", na=False)
    df.loc[has_gs & not_prefixed, "Name"] = "+" + gss[has_gs & not_prefixed] + df.loc[
        has_gs & not_prefixed, "Name"
    ].astype(str)
    return df


def _apply_function_designation(df: pd.DataFrame) -> pd.DataFrame:
    def type_to_func(t: str) -> str:
        m = {
            "SE.A9F04604_ADV": "POWER",
            "WAGO.2002-1611/1000-541_ADV": "FUSES",
            "WAGO.2002-1611/1000-836_ADV": "FUSES",
            "SE.RGZE1S48M + SE.RXG22P7": "2POLE",
            "SE.RXZE2S114M + SE.RXM4GB2BD": "4POLE",
            "WAGO.2002-3201_ADV": "CONTROL",
            "WAGO.2002-3207_ADV": "CONTROL",
        }
        return m.get(t, "")

    func_col = df["Type"].astype(str).map(type_to_func).fillna("")
    hasf = func_col.astype(str).ne("")
    df.loc[hasf, "Name"] = "=" + func_col[hasf].astype(str) + df.loc[hasf, "Name"].astype(str)
    return df


def _allocate_category_gs(
    df: pd.DataFrame, missing_default: int | pd.Series | None = None, fallback_col: str | None = None
) -> pd.DataFrame:
    """
    Fill missing Group Sorting values without overwriting existing ones.

    `missing_default` supplies the value (or per-row series) for rows where
    Group Sorting is missing. If `fallback_col` is provided, its numeric value
    is used when both Group Sorting and `missing_default` are missing.
    """
    df = df.copy()

    gs_numeric = pd.to_numeric(df["Group Sorting"], errors="coerce")

    if isinstance(missing_default, pd.Series):
        default_series = missing_default.reindex(df.index)
    elif missing_default is None:
        default_series = pd.Series([None] * len(df), index=df.index, dtype="float")
    else:
        default_series = pd.Series(missing_default, index=df.index)

    if fallback_col is not None and fallback_col in df.columns:
        fallback_series = pd.to_numeric(df[fallback_col], errors="coerce")
        default_series = default_series.combine_first(fallback_series)

    resolved = gs_numeric.combine_first(default_series)
    df["Group Sorting"] = resolved
    df["_gs_sort"] = pd.to_numeric(resolved, errors="coerce").fillna(10**9).astype(int)

    return df


def _terminal_base_name(name: str) -> str:
    s = str(name)
    s = re.sub(r"^=[A-Z]+", "", s, flags=re.IGNORECASE)
    s = s.lstrip("=")
    s = re.sub(r"^\+\d+", "", s)
    return s


def _validate_terminal_uniqueness(df: pd.DataFrame) -> None:
    if df.empty:
        return

    base = df["Name"].astype(str).apply(_terminal_base_name)
    combo = df["Type"].astype(str)
    gs_num = pd.to_numeric(df["Group Sorting"], errors="coerce")

    dup_mask = (
        pd.concat({"base": base, "type": combo, "gs": gs_num}, axis=1)
        .groupby(["base", "type"])["gs"]
        .transform("nunique")
        > 1
    )
    if dup_mask.any():
        bad = df.loc[dup_mask, ["Name", "Type", "Group Sorting"]]
        raise ValueError(
            "Terminal invariant violated: multiple GS values for the same terminal (Name+Type):\n"
            f"{bad.to_string(index=False)}"
        )


def _extract_x_digits(name: str) -> str | None:
    m = re.search(r"-X(\d+)", str(name))
    if not m:
        return None
    digits = m.group(1)
    return digits if len(digits) >= 2 else None


def _extract_x_subgroup(name: str) -> str | None:
    digits = _extract_x_digits(name)
    return digits[:2] if digits is not None else None


def _x_number_endswith_14(name: str) -> bool:
    digits = _extract_x_digits(name)
    return bool(digits and digits.endswith("14"))


def _terminal_sort_key(name: str) -> int:
    """
    Rikiavimo raktas, kad X192A* galėtume laikyti "tarp" X1922 ir X1953.
    Prielaida: X192A3 = 1923 (t.y. 192 + A3).
    """
    s = str(name)

    m = re.search(r"-X(\d{4})\b", s)
    if m:
        return int(m.group(1))

    m = re.search(r"-X(\d{3})A(\d+)\b", s)  # X192A3
    if m:
        return int(m.group(1) + m.group(2))  # "192"+"3" => 1923

    m = re.search(r"-X(\d+)", s)
    if m:
        return int(m.group(1))

    return 10**9


def _normalize_terminal_name(name: str) -> str:
    """
    Normalizuoja terminalo pavadinimą deduplikavimui, nuimdama
    automatiškai pridėtą +GS prefiksą ir lygybės ženklą.
    """
    s = str(name).lstrip("=")
    return re.sub(r"^\+\d+", "", s)


# -----------------------------------------------------------------------------
# Terminal rule: X192A insertion inside one base GS group
# -----------------------------------------------------------------------------
def apply_x192a_terminal_gs_rules(df: pd.DataFrame) -> pd.DataFrame:
    """
    Taisyklė:

    Bazė = ORIGINALUS GS iš įkelto failo (df['_gs_orig']).

    Vienoje bazinėje GS (pvz. 2010), jei yra X192A*:
      - iki X192A paliekam esamam GS
      - visi X192A* perkelti į naują GS = next_free(base+1) (pvz. 2011)
      - visi terminalai po X192A (pagal X numerį) perkelti į dar sekantį GS (pvz. 2012)
    """
    terminal_types = {"WAGO.2002-3201_ADV"}
    is_terminal = df["Type"].astype(str).isin(terminal_types)

    gs_num = pd.to_numeric(df["Group Sorting"], errors="coerce")
    existing_gs: set[int] = set(gs_num.dropna().astype(int).tolist())

    # Bazė = originalus GS (tik terminalams)
    if "_gs_orig" not in df.columns:
        df["_gs_orig"] = gs_num

    work = df[is_terminal & df["_gs_orig"].notna()].copy()
    if work.empty:
        return df

    work["_base_gs"] = work["_gs_orig"].astype(int)
    work["_xkey"] = work["Name"].astype(str).apply(_terminal_sort_key)
    work["_is_x192a"] = work["Name"].astype(str).str.contains(r"-X192A\d+\b", regex=True, na=False)

    for base_gs, grp in work.groupby("_base_gs", sort=True):
        if not grp["_is_x192a"].any():
            continue

        gs_for_x192a = _next_free_gs(existing_gs, int(base_gs) + 1)
        max_x192a_key = int(grp.loc[grp["_is_x192a"], "_xkey"].max())
        after_mask = (grp["_xkey"] > max_x192a_key) & (~grp["_is_x192a"])

        gs_for_after = None
        if after_mask.any():
            gs_for_after = _next_free_gs(existing_gs, gs_for_x192a + 1)

        df.loc[grp.loc[grp["_is_x192a"]].index, "Group Sorting"] = gs_for_x192a
        if gs_for_after is not None:
            df.loc[grp.loc[after_mask].index, "Group Sorting"] = gs_for_after

    return df


# -----------------------------------------------------------------------------
# Main processing
# -----------------------------------------------------------------------------
def process_excel(file_bytes: bytes) -> Tuple[pd.DataFrame, pd.DataFrame, bytes, Dict[str, int]]:
    df = pd.read_excel(BytesIO(file_bytes), sheet_name=0, engine="openpyxl")
    df = _drop_unnamed_cols(df)
    input_rows = len(df)

    required = ["Name", "Type", "Quantity", "Group Sorting"]
    _ensure_cols(df, required)

    df["Name"] = df["Name"].astype(str)
    df["Type"] = df["Type"].astype(str)
    df["_name_orig"] = df["Name"]

    # ensure columns exist
    for col, default in [
        ("Accessories", ""),
        ("Quantity of accessories", 0),
        ("Accessories2", ""),
        ("Quantity of accessories2", 0),
        ("Designation", ""),
    ]:
        if col not in df.columns:
            df[col] = default

    removed_parts: List[pd.DataFrame] = []

    # -------------------------------------------------------------------------
    # STEP 0 – remove non-numeric Group Sorting (but not empty)
    # -------------------------------------------------------------------------
    gs_raw = df["Group Sorting"]
    gs_num = pd.to_numeric(gs_raw, errors="coerce")
    non_numeric = gs_raw.notna() & (gs_raw.astype(str).str.strip() != "") & gs_num.isna()
    _append_removed(removed_parts, df[non_numeric], "Removed: Group Sorting is not numeric")
    df = df[~non_numeric].copy()

    # Įsimenam ORIGINALŲ GS (tik skaičiai)
    df["_gs_orig"] = pd.to_numeric(df["Group Sorting"], errors="coerce")

    # -------------------------------------------------------------------------
    # STEP 1 – remove by Name prefixes
    # -------------------------------------------------------------------------
    remove_prefixes = ("-B", "-C", "-R", "-M", "-P", "-Q", "-S", "-W", "-T")
    m1 = _starts_with_any(df["Name"], remove_prefixes)
    _append_removed(removed_parts, df[m1], "Removed by Name prefix (-B/-C/-R/-M/-P/-Q/-S/-W/-T)")
    df = df[~m1].copy()

    # -------------------------------------------------------------------------
    # STEP 2 – invalid prefix highlight (not removed)
    # -------------------------------------------------------------------------
    valid_prefixes = ("-F", "-K", "-X")
    df["_highlight_invalid_prefix"] = ~_starts_with_any(df["Name"], valid_prefixes)

    # -------------------------------------------------------------------------
    # STEP 3 – Type allowlist (palik kaip pas tave projekte)
    # -------------------------------------------------------------------------
    allowed_types = {
        "2002-1611/1000-541",
        "2002-1611/1000-836",
        "2002-3201",
        "2002-3207",
        "RXZE2S114M",
        "RXM4GB2BD",
        "RXG22P7",
        "RXG22BD",
        "RGZE1S48M",
        "A9F04604",
        "RE17LCBM",
        "RE22R1AMR",
        "39.00.8.230.8240",
    }
    m3 = df["Type"].isin(allowed_types)
    _append_removed(removed_parts, df[~m3], "Removed by Type filter (not in allowed list)")
    df = df[m3].copy()

    # -------------------------------------------------------------------------
    # STEP 4 – map types (WAGO/SE) + _ADV
    # -------------------------------------------------------------------------
    wago_types = {
        "2002-1611/1000-541",
        "2002-1611/1000-836",
        "2002-3201",
        "2002-3207",
        "249-116",
        "2002-991",
        "2002-3292",
    }
    mw = df["Type"].isin(wago_types)
    df.loc[mw, "Type"] = "WAGO." + df.loc[mw, "Type"].astype(str)

    se_map = {
        "RGZE1S48M": "SE.RGZE1S48M",
        "RXG22P7": "SE.RXG22P7",
        "RXG22BD": "SE.RXG22BD",
        "RXM4GB2BD": "SE.RXM4GB2BD",
        "RXZE2S114M": "SE.RXZE2S114M",
        "A9F04604": "SE.A9F04604",
    }
    df["Type"] = df["Type"].replace(se_map)

    adv_map = {
        "WAGO.2002-3207": "WAGO.2002-3207_ADV",
        "WAGO.2002-3201": "WAGO.2002-3201_ADV",
        "WAGO.2002-3292": "WAGO.2002-3292_ADV",
        "WAGO.2002-991": "WAGO.2002-991_ADV",
        "WAGO.249-116": "WAGO.249-116_ADV",
        "WAGO.2002-1611/1000-541": "WAGO.2002-1611/1000-541_ADV",
        "WAGO.2002-1611/1000-836": "WAGO.2002-1611/1000-836_ADV",
        "SE.A9F04604": "SE.A9F04604_ADV",
    }
    df["Type"] = df["Type"].replace(adv_map)

    relay_map = {
        "RE17LCBM": "SE.RE17LCBM_ADV",
        "RE22R1AMR": "SE.RE22R1AMR_ADV",
        "39.00.8.230.8240": "FIN.39.00.8.230.8240_ADV",
    }
    df["Type"] = df["Type"].replace(relay_map)

    # -------------------------------------------------------------------------
    # STEP 4b – relay merge by Name (2POLE/4POLE)
    # -------------------------------------------------------------------------
    def _merge_relay(base_type: str, combined_type: str) -> None:
        nonlocal df
        mb = df["Type"].astype(str).eq(base_type)
        if not mb.any():
            return

        drop_idxs: List[int] = []
        # grupuojam tik pagal Name, kur bent viena eilutė yra base_type
        names_with_base = set(df.loc[mb, "Name"].astype(str).tolist())

        for name_val, grp in df[df["Name"].astype(str).isin(names_with_base)].groupby("Name", sort=False):
            idxs = grp.index.to_list()
            keep = None
            for i in idxs:
                if df.at[i, "Type"] == base_type:
                    keep = i
                    break
            if keep is None:
                keep = idxs[0]

            for i in idxs:
                if i != keep:
                    drop_idxs.append(i)

            df.at[keep, "Type"] = combined_type
            df.at[keep, "Quantity"] = 1

        if drop_idxs:
            _append_removed(removed_parts, df.loc[drop_idxs], "Removed: merged into relay combined row")
            df = df.drop(index=drop_idxs).copy()

    _merge_relay("SE.RGZE1S48M", "SE.RGZE1S48M + SE.RXG22P7")
    _merge_relay("SE.RXZE2S114M", "SE.RXZE2S114M + SE.RXM4GB2BD")

    # -------------------------------------------------------------------------
    # Category pipelines
    # -------------------------------------------------------------------------
    terminal_types = {"WAGO.2002-3201_ADV", "WAGO.2002-3207_ADV"}
    fuse_types = {"WAGO.2002-1611/1000-541_ADV", "WAGO.2002-1611/1000-836_ADV", "SE.A9F04604_ADV"}
    relay_types = set(df["Type"].unique()) - terminal_types - fuse_types

    terminal_df = df[df["Type"].isin(terminal_types)].copy()
    fuse_df = df[df["Type"].isin(fuse_types)].copy()
    relay_df = df[df["Type"].isin(relay_types)].copy()

    def process_relays(relay_data: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
        if relay_data.empty:
            return relay_data.copy(), pd.DataFrame(columns=list(relay_data.columns) + ["Removed Reason"])

        relay_data["_gs_sort"] = pd.to_numeric(relay_data["Group Sorting"], errors="coerce").fillna(10**9).astype(int)
        relay_data = _allocate_category_gs(relay_data, missing_default=1)
        return relay_data, pd.DataFrame(columns=list(relay_data.columns) + ["Removed Reason"])

    def process_fuses(fuse_data: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
        if fuse_data.empty:
            return fuse_data.copy(), pd.DataFrame(columns=list(fuse_data.columns) + ["Removed Reason"])

        fuse_data = fuse_data.copy()
        fuse_data["_orig_order"] = range(len(fuse_data))
        fuse_data["Accessories"] = fuse_data["Accessories"].fillna("")
        fuse_data["Accessories2"] = fuse_data["Accessories2"].fillna("")
        fuse_data["Quantity of accessories"] = (
            pd.to_numeric(fuse_data["Quantity of accessories"], errors="coerce").fillna(0).astype(int)
        )
        fuse_data["Quantity of accessories2"] = (
            pd.to_numeric(fuse_data["Quantity of accessories2"], errors="coerce").fillna(0).astype(int)
        )

        def _parse_fuse_key(name: str) -> dict | None:
            s = str(name)
            m_a = re.search(r"-F(\d{3})A(\d+)", s)
            if m_a:
                main = int(m_a.group(1))
                sub = int(m_a.group(2))
                return {"main": main, "sub": sub, "is_a": True}
            m = re.search(r"-F(\d+)(?:\.(\d+))?", s)
            if not m:
                return None
            return {"main": int(m.group(1)), "sub": int(m.group(2) or 0), "is_a": False}

        def _fuse_order_value(key: dict | None) -> float | None:
            if key is None:
                return None
            if key.get("is_a"):
                # F192A5 -> 1925 (matches terminal-style insertion)
                return int(f"{key['main']}{key['sub']}")
            return key["main"] + key["sub"] / 10

        fuse_data["_fuse_key"] = fuse_data["Name"].astype(str).apply(_parse_fuse_key)
        fuse_data["_fuse_order_val"] = fuse_data["_fuse_key"].apply(_fuse_order_value)

        def _fuse_sort_key(idx: int) -> Tuple[int, int, int]:
            order_val = fuse_data.at[idx, "_fuse_order_val"]
            orig = int(fuse_data.at[idx, "_orig_order"])
            if pd.isna(order_val):
                return (1, 10**9, orig)
            return (0, int(order_val), orig)

        fuse_types_target = {"WAGO.2002-1611/1000-541_ADV", "WAGO.2002-1611/1000-836_ADV"}
        is_541 = fuse_data["Type"].astype(str).eq("WAGO.2002-1611/1000-541_ADV")
        is_836 = fuse_data["Type"].astype(str).eq("WAGO.2002-1611/1000-836_ADV")

        existing_gs = pd.to_numeric(fuse_data["Group Sorting"], errors="coerce")
        existing_fuse_gs = sorted(set(int(x) for x in existing_gs.dropna() if 31 <= x <= 50))
        existing_non541_gs = sorted(
            set(int(x) for x in existing_gs[~is_541].dropna() if 31 <= x <= 50)
        )

        gs_541 = 31
        if 31 in existing_non541_gs:
            alternatives = [g for g in existing_fuse_gs if g != 31]
            if alternatives:
                gs_541 = min(alternatives)

        if is_541.any():
            fuse_data.loc[is_541, "Group Sorting"] = gs_541

        fuse_main = fuse_data["_fuse_key"].apply(lambda k: k.get("main") if isinstance(k, dict) else None)
        block32 = is_836 & fuse_main.apply(lambda m: m is not None and 900 <= m <= 999)
        fuse_data.loc[block32, "Group Sorting"] = 32

        base_gs = 33
        non_f9_mask = is_836 & (~block32)
        fuse_data.loc[non_f9_mask, "Group Sorting"] = base_gs

        f192a_mask = non_f9_mask & fuse_data["Name"].astype(str).str.contains(r"-F192A", regex=True, na=False)
        order_series = fuse_data["_fuse_order_val"]
        after_mask = pd.Series([False] * len(fuse_data), index=fuse_data.index)
        if f192a_mask.any():
            first_key = order_series[f192a_mask].min()
            last_key = order_series[f192a_mask].max()
            before_mask = non_f9_mask & (~f192a_mask) & order_series.notna() & (order_series < first_key)
            after_mask = non_f9_mask & (~f192a_mask) & order_series.notna() & (order_series > last_key)

            fuse_data.loc[f192a_mask, "Group Sorting"] = base_gs + 1
            if after_mask.any():
                fuse_data.loc[after_mask, "Group Sorting"] = base_gs + 2

        def _apply_accessories(block_mask: pd.Series, accessories: str | None, accessories2: str | None) -> None:
            if not block_mask.any():
                return
            idxs = block_mask[block_mask].index.tolist()
            last_idx = sorted(idxs, key=_fuse_sort_key)[-1]
            if accessories and str(fuse_data.at[last_idx, "Accessories"]).strip() == "":
                fuse_data.at[last_idx, "Accessories"] = accessories
                fuse_data.at[last_idx, "Quantity of accessories"] = 1
            if accessories2 and str(fuse_data.at[last_idx, "Accessories2"]).strip() == "":
                fuse_data.at[last_idx, "Accessories2"] = accessories2
                fuse_data.at[last_idx, "Quantity of accessories2"] = 1

        last_836_block = None
        if f192a_mask.any():
            if after_mask.any():
                last_836_block = after_mask
            else:
                last_836_block = f192a_mask
        else:
            last_836_block = non_f9_mask

        _apply_accessories(is_541, "WAGO.2002-991_ADV", "WAGO.249-116_ADV")
        _apply_accessories(block32, "WAGO.2002-991_ADV", None)
        _apply_accessories(last_836_block, "WAGO.2002-991_ADV", None)

        a9f_mask = fuse_data["Type"].astype(str).eq("SE.A9F04604_ADV")
        missing_defaults = pd.Series(None, index=fuse_data.index, dtype="float")
        missing_defaults.loc[a9f_mask] = 0
        fuse_data["_terminal_sort"] = fuse_data["_fuse_order_val"].where(
            fuse_data["_fuse_order_val"].notna(), 10**9 + fuse_data["_orig_order"]
        )
        fuse_data = _allocate_category_gs(fuse_data, missing_default=missing_defaults)
        fuse_data["_gs_sort"] = pd.to_numeric(fuse_data["Group Sorting"], errors="coerce").fillna(10**9).astype(int)
        fuse_data = fuse_data.drop(columns=["_orig_order", "_fuse_key", "_fuse_order_val"], errors="ignore")

        return fuse_data, pd.DataFrame(columns=list(fuse_data.columns) + ["Removed Reason"])

    def process_terminals(term_data: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
        removed_local: List[pd.DataFrame] = []
        if term_data.empty:
            empty_removed = pd.DataFrame(columns=list(term_data.columns) + ["Removed Reason"])
            return term_data.copy(), empty_removed

        term_data = apply_x192a_terminal_gs_rules(term_data)

        term_data["Quantity"] = pd.to_numeric(term_data["Quantity"], errors="coerce").fillna(0)

        rows: List[dict] = []
        for _, r in term_data.iterrows():
            d = r.to_dict()
            t = str(d.get("Type", ""))
            qty = pd.to_numeric(d.get("Quantity", 0), errors="coerce")

            if t in terminal_types and pd.notna(qty) and qty > 1:
                n = int(qty)
                base = d.copy()
                base["Quantity"] = 1
                base["Designation"] = ""
                rows.append(base)

                for i in range(1, n):
                    rr = d.copy()
                    rr["Quantity"] = 1
                    rr["Designation"] = str(i)
                    rr["Accessories"] = ""
                    rr["Quantity of accessories"] = 0
                    rr["Accessories2"] = ""
                    rr["Quantity of accessories2"] = 0
                    rows.append(rr)
            else:
                d["Designation"] = d.get("Designation", "")
                rows.append(d)

        term_data = pd.DataFrame(rows)

        # PE mapping and /3 reduction
        pe_mask = term_data["Type"].astype(str).str.contains("2002-3207", na=False)
        pe_with_base = pe_mask & term_data["_gs_orig"].notna()

        if pe_with_base.any():
            base_values = term_data.loc[pe_with_base, "_gs_orig"].astype(int)
            unique_bases = sorted(base_values.unique())
            base_to_pe = {b: i + 1 for i, b in enumerate(unique_bases)}

            for idx in term_data.loc[pe_with_base].index:
                base = int(term_data.at[idx, "_gs_orig"])
                pe_id = base_to_pe[base]
                term_data.at[idx, "Name"] = f"+{base}-PE{pe_id}"
                term_data.at[idx, "Group Sorting"] = base

        name_s = term_data["Name"].astype(str)
        pe_name_mask = name_s.str.contains(r"-PE\d+\b", regex=True, na=False)
        if pe_name_mask.any():
            keep: List[int] = []
            for pe_name, grp in term_data[pe_name_mask].groupby("Name", sort=False):
                idxs = grp.index.to_list()
                cnt = len(idxs)
                need = (cnt + 2) // 3  # ceil(cnt/3)

                for offset, idx in enumerate(idxs[:need]):
                    term_data.at[idx, "Designation"] = "" if offset == 0 else str(offset)

                removed_idxs = idxs[need:]
                if removed_idxs:
                    _append_removed(
                        removed_local,
                        term_data.loc[removed_idxs],
                        "Removed: PE reduction ceil(count/3)",
                    )

                keep.extend(idxs[:need])

            term_data = term_data[(~pe_name_mask) | (term_data.index.isin(keep))].copy()

        # Accessories
        term_data["Accessories"] = term_data["Accessories"].fillna("")
        term_data["Accessories2"] = term_data["Accessories2"].fillna("")
        term_data["Quantity of accessories"] = (
            pd.to_numeric(term_data["Quantity of accessories"], errors="coerce").fillna(0).astype(int)
        )
        term_data["Quantity of accessories2"] = (
            pd.to_numeric(term_data["Quantity of accessories2"], errors="coerce").fillna(0).astype(int)
        )

        # current GS for ordering
        gs_now = pd.to_numeric(term_data["Group Sorting"], errors="coerce").fillna(10**9).astype(int)
        term_data["_gs_sort"] = gs_now

        gs_base = pd.to_numeric(term_data.get("_gs_orig", pd.Series([None] * len(term_data))), errors="coerce")

        def _place_cover(candidate_idxs: List[int], gs_orig_value: int | None, subgroup_key: str | None) -> None:
            for idx in reversed(candidate_idxs):
                if str(term_data.at[idx, "Accessories"]).strip() == "":
                    term_data.at[idx, "Accessories"] = "WAGO.2002-3292_ADV"
                    term_data.at[idx, "Quantity of accessories"] = 1
                    return
            raise ValueError(
                f"Cannot place terminal cover (gs_orig={gs_orig_value}, subgroup={subgroup_key}): no empty Accessories slot"
            )

        for base in (1010, 1110):
            m = term_data["Type"].astype(str).isin(terminal_types) & gs_base.eq(base)
            if not m.any():
                continue
            tmp = term_data[m].copy()
            tmp["_sub"] = tmp["_name_orig"].astype(str).apply(_extract_x_subgroup)
            for _, grp in tmp.groupby("_sub", sort=False, dropna=False):
                idxs = grp.index.to_list()
                _place_cover(idxs, base, grp["_sub"].iloc[0])

        m1030 = term_data["Type"].astype(str).isin(terminal_types) & gs_base.eq(1030)
        if m1030.any():
            tmp = term_data[m1030].copy()
            tmp["_sub"] = tmp["_name_orig"].astype(str).apply(_extract_x_subgroup)
            for _, grp in tmp.groupby("_sub", sort=False, dropna=False):
                target_idx = None
                idxs = grp.index.to_list()
                for idx in idxs:
                    if _x_number_endswith_14(term_data.at[idx, "_name_orig"]):
                        target_idx = idx
                        break
                if target_idx is None:
                    target_idx = idxs[-1]
                target_pos = idxs.index(target_idx)
                candidate_idxs = idxs[: target_pos + 1]
                _place_cover(candidate_idxs, 1030, grp["_sub"].iloc[0])

        other_mask = (
            term_data["Type"].astype(str).isin(terminal_types) & gs_base.notna() & (~gs_base.isin([1010, 1110, 1030]))
        )
        if other_mask.any():
            xkey = term_data["Name"].astype(str).apply(_terminal_sort_key)
            for _, grp_idxs in term_data[other_mask].groupby(gs_base[other_mask], sort=True).groups.items():
                idxs = list(grp_idxs)
                sorted_idxs = sorted(
                    idxs,
                    key=lambda i: (
                        int(term_data.at[i, "_gs_sort"]),
                        int(xkey.loc[i]),
                        str(term_data.at[i, "Designation"]),
                    ),
                )
                gs_val = term_data.at[sorted_idxs[0], "_gs_orig"]
                _place_cover(sorted_idxs, gs_val if pd.notna(gs_val) else None, None)

        # remove duplicate terminals within same original GS
        term_data["_terminal_sort"] = term_data["Name"].astype(str).apply(_terminal_sort_key)

        is_terminal_mask = term_data["Type"].astype(str).isin(terminal_types)
        if is_terminal_mask.any():
            dup_idxs: List[int] = []
            term_df = term_data[is_terminal_mask & term_data["_gs_orig"].notna()].copy()
            term_df["_norm_term_name"] = term_df["Name"].apply(_normalize_terminal_name)

            group_keys = ["_gs_orig", "Type", "_norm_term_name", "Designation"]
            for _, grp in term_df.groupby(group_keys, sort=False):
                keep_idx = grp.sort_values(["_gs_sort", "_terminal_sort", "Name"], kind="stable").index[0]
                dup_idxs.extend(i for i in grp.index if i != keep_idx)

            if dup_idxs:
                _append_removed(
                    removed_local, term_data.loc[dup_idxs], "Removed: duplicate terminal within original GS"
                )
                term_data = term_data.drop(index=dup_idxs).copy()

        term_data = _allocate_category_gs(term_data, missing_default=None, fallback_col="_gs_orig")
        term_data = _add_gs_prefix(term_data)

        # Terminal duplicate bugfix across GS
        term_data["_terminal_base"] = term_data["Name"].astype(str).apply(_terminal_base_name)
        term_data["_gs_num_final"] = pd.to_numeric(term_data["Group Sorting"], errors="coerce")
        duplicate_idxs: List[int] = []
        for (_, _type), grp in term_data.groupby(["_terminal_base", "Type"], sort=False):
            if grp["_gs_num_final"].nunique() <= 1:
                continue
            keep_idx = grp.sort_values(["_gs_num_final", "_terminal_sort", "Name"], kind="stable").index[0]
            duplicate_idxs.extend(i for i in grp.index if i != keep_idx)

        if duplicate_idxs:
            _append_removed(
                removed_local,
                term_data.loc[duplicate_idxs],
                "Removed: duplicate terminal (Name+Type) across multiple GS",
            )
            term_data = term_data.drop(index=duplicate_idxs).copy()

        term_data.drop(columns=["_terminal_base", "_gs_num_final"], inplace=True, errors="ignore")
        _validate_terminal_uniqueness(term_data)

        return term_data, (pd.concat(removed_local, ignore_index=True) if removed_local else pd.DataFrame())

    relay_clean, relay_removed = process_relays(relay_df)
    fuse_clean, fuse_removed = process_fuses(fuse_df)
    terminal_clean, terminal_removed = process_terminals(terminal_df)

    cleaned_df = pd.concat([relay_clean, fuse_clean, terminal_clean], ignore_index=True)
    removed_df_parts = [p for p in [relay_removed, fuse_removed, terminal_removed] if p is not None and not p.empty]
    removed_df = pd.concat(removed_parts + removed_df_parts, ignore_index=True) if (removed_parts or removed_df_parts) else pd.DataFrame()
    if removed_df.empty:
        removed_df = pd.DataFrame(columns=list(df.columns) + ["Removed Reason"])
    elif "Removed Reason" not in removed_df.columns:
        removed_df["Removed Reason"] = ""

    cleaned_df = _add_gs_prefix(cleaned_df)
    cleaned_df = _apply_function_designation(cleaned_df)

    _validate_terminal_uniqueness(cleaned_df[cleaned_df["Type"].astype(str).isin(terminal_types)])

    # Final sort
    is_terminal_sorted = cleaned_df["Type"].astype(str).isin(terminal_types)
    if "_terminal_sort" in cleaned_df.columns:
        cleaned_df["_terminal_sort"] = cleaned_df["_terminal_sort"].fillna(
            cleaned_df["Name"].astype(str).apply(_terminal_sort_key)
        )
    else:
        cleaned_df["_terminal_sort"] = cleaned_df["Name"].astype(str).apply(_terminal_sort_key)
    cleaned_df["_terminal_sort_order"] = is_terminal_sorted.map({True: 0, False: 1})
    cleaned_df["_gs_sort"] = pd.to_numeric(cleaned_df["Group Sorting"], errors="coerce").fillna(10**9).astype(int)

    cleaned_df = cleaned_df.sort_values(
        ["_gs_sort", "_terminal_sort_order", "_terminal_sort", "Name"],
        kind="stable",
    )

    # Single secondary accessory on the globally last WAGO.2002-3201_ADV terminal
    mask_primary_terminal = cleaned_df["Type"].astype(str).eq("WAGO.2002-3201_ADV")
    if mask_primary_terminal.any():
        last_idx = cleaned_df[mask_primary_terminal].index[-1]
        if str(cleaned_df.at[last_idx, "Accessories2"]).strip() == "":
            cleaned_df.at[last_idx, "Accessories2"] = "WAGO.249-116_ADV"
            cleaned_df.at[last_idx, "Quantity of accessories2"] = 1

    cleaned_df = cleaned_df.drop(columns=["_gs_sort", "_terminal_sort", "_terminal_sort_order"], errors="ignore")

    # Workbook output
    wb = Workbook()
    ws_c = wb.active
    ws_c.title = "Cleaned"
    ws_r = wb.create_sheet("Removed")

    out_clean = cleaned_df.drop(columns=["_highlight_invalid_prefix"], errors="ignore")
    out_clean = out_clean.drop(columns=["_gs_orig", "_name_orig"], errors="ignore")
    for row in dataframe_to_rows(out_clean, index=False, header=True):
        ws_c.append(row)

    for r in range(2, ws_c.max_row + 1):
        c = ws_c.cell(row=r, column=1)
        c.number_format = "@"
        c.data_type = "s"

    if "_highlight_invalid_prefix" in cleaned_df.columns:
        flags = cleaned_df["_highlight_invalid_prefix"].tolist()
        for excel_row, bad in enumerate(flags, start=2):
            if bool(bad):
                for col in range(1, ws_c.max_column + 1):
                    ws_c.cell(row=excel_row, column=col).fill = YELLOW_FILL

    removed_out = removed_df.drop(
        columns=["_gs_orig", "_name_orig", "_terminal_sort", "_terminal_sort_order"], errors="ignore"
    )
    for row in dataframe_to_rows(removed_out, index=False, header=True):
        ws_r.append(row)

    if ws_r.max_row >= 1:
        headers = [ws_r.cell(row=1, column=c).value for c in range(1, ws_r.max_column + 1)]
        if "Name" in headers:
            name_col = headers.index("Name") + 1
            for r in range(2, ws_r.max_row + 1):
                c = ws_r.cell(row=r, column=name_col)
                c.number_format = "@"
                c.data_type = "s"

    bio = BytesIO()
    wb.save(bio)
    out_bytes = bio.getvalue()

    stats = {
        "input_rows": input_rows,
        "cleaned_rows": len(out_clean),
        "removed_rows": len(removed_df),
    }

    return out_clean, removed_df, out_bytes, stats
